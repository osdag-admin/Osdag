# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'ui_design_preferences.ui'
#
# Created by: PyQt5 UI code generator 5.10.1
#
# WARNING! All changes made in this file will be lost!
from Common import *
from utils.common.component import Section,I_sectional_Properties
from utils.common.component import *
from utils.common.other_standards import *
from design_type.connection.fin_plate_connection import FinPlateConnection
from design_type.connection.connection import Connection

from PyQt5.QtWidgets import QMessageBox, qApp
from PyQt5.QtGui import QDoubleValidator, QIntValidator, QPixmap, QPalette
from PyQt5.QtCore import QFile, pyqtSignal, QTextStream, Qt, QIODevice,pyqtSlot
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMainWindow, QDialog, QFontDialog, QApplication, QFileDialog, QColorDialog
from PyQt5.QtCore import QFile, pyqtSignal, QTextStream, Qt, QIODevice
from PyQt5.QtCore import QRegExp
from PyQt5.QtGui import QBrush
from PyQt5.QtGui import QColor
from PyQt5.QtGui import QDoubleValidator, QIntValidator, QPixmap, QPalette
from PyQt5.QtGui import QTextCharFormat
from PyQt5.QtGui import QTextCursor
from PyQt5.QtWidgets import QMainWindow, QDialog, QFontDialog, QApplication, QFileDialog, QColorDialog
from PyQt5.QtGui import QStandardItem
import os
import json
import logging
from drawing_2D.Svg_Window import SvgWindow
import sys
import sqlite3
import shutil
import openpyxl


class Ui_Dialog(object):
    def setupUi(self, DesignPreferences, main, input_dictionary):
        DesignPreferences.setObjectName("DesignPreferences")
        #DesignPreferences.resize(1170,870)
        self.gridLayout_5 = QtWidgets.QGridLayout(DesignPreferences)
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.gridLayout_2 = QtWidgets.QGridLayout()
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.btn_defaults = QtWidgets.QPushButton(DesignPreferences)
        self.btn_defaults.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        font = QtGui.QFont()
        font.setFamily("Arial")
        self.btn_defaults.setFont(font)
        self.btn_defaults.setObjectName("btn_defaults")
        self.gridLayout_2.addWidget(self.btn_defaults, 0, 1, 1, 1)
        self.btn_save = QtWidgets.QPushButton(DesignPreferences)

        font = QtGui.QFont()
        font.setFamily("Arial")
        self.btn_save.setFont(font)
        self.btn_save.setObjectName("btn_save")
        self.gridLayout_2.addWidget(self.btn_save, 0, 2, 1, 1)
        self.btn_close = QtWidgets.QPushButton(DesignPreferences)

        self.btn_close.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        font = QtGui.QFont()
        font.setFamily("Arial")
        self.btn_close.setFont(font)
        self.btn_close.setObjectName("btn_close")
        self.gridLayout_2.addWidget(self.btn_close, 0, 3, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(28, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem, 0, 0, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem1, 0, 4, 1, 1)
        self.gridLayout_5.addLayout(self.gridLayout_2, 1, 0, 1, 1)
        self.tabWidget = QtWidgets.QTabWidget(DesignPreferences)
        font = QtGui.QFont()
        font.setFamily("Arial")
        self.tabWidget.setFont(font)
        self.tabWidget.setObjectName("tabWidget")

###################################################################
######################################################################
######################################################################
####################################################################

#START

        # tab_index = 0
        # for tab_details in main.tab_list(main):
        #     tab_name = tab_details[0]
        #     tab_elements = tab_details[1]
        #     tab = QtWidgets.QWidget()
        #     tab.setObjectName(tab_name)
        #     elements = tab_elements()
        _translate = QtCore.QCoreApplication.translate
        #     i = 0
        #     j = 6
        #     for element in elements:
        #         lable = element[1]
        #         type = element[2]
        #         # value = option[4]
        #         if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
        #             l = QtWidgets.QLabel(tab)
        #             if lable in [KEY_DISP_SUPTNGSEC_THERMAL_EXP]:
        #                 l.setGeometry(QtCore.QRect(3 + j, 10 + i, 165, 28))
        #                 i = i + 10
        #             else:
        #                 l.setGeometry(QtCore.QRect(3 + j, 10 + i, 165, 22))
        #             font = QtGui.QFont()
        #             font.setPointSize(9)
        #             if lable in [KEY_DISP_SUPTNGSEC_DESIGNATION, KEY_DISP_SUPTNGSEC_TYPE, KEY_DISP_SUPTNGSEC_SOURCE]:
        #                 font.setWeight(75)
        #             else:
        #                 font.setWeight(50)
        #             l.setFont(font)
        #             l.setObjectName(element[0] + "_label")
        #             l.setText(_translate("MainWindow", "<html><head/><body><p>" + lable + "</p></body></html>"))
        #             l.setAlignment(QtCore.Qt.AlignCenter)
        #
        #         if type == TYPE_COMBOBOX:
        #             combo = QtWidgets.QComboBox(tab)
        #             combo.setGeometry(QtCore.QRect(170 + j, 10 + i, 130, 22))
        #             font = QtGui.QFont()
        #             font.setPointSize(9)
        #             font.setBold(False)
        #             font.setWeight(50)
        #             combo.setFont(font)
        #             combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
        #             combo.setMaxVisibleItems(5)
        #             combo.setObjectName(element[0])
        #             for item in element[3]:
        #                 combo.addItem(item)
        #
        #         if type == TYPE_TITLE:
        #             q = QtWidgets.QLabel(tab)
        #             q.setGeometry(QtCore.QRect(j, 10 + i, 155, 35))
        #             font = QtGui.QFont()
        #             font.setPointSize(10)
        #             q.setFont(font)
        #             q.setObjectName("_title")
        #             q.setText(_translate("MainWindow",
        #                                  "<html><head/><body><p><span style=\" font-weight:600;\">" + lable + "</span></p></body></html>"))
        #
        #         if type == TYPE_TEXTBOX:
        #             r = QtWidgets.QLineEdit(tab)
        #             r.setGeometry(QtCore.QRect(170 + j, 10 + i, 130, 22))
        #             font = QtGui.QFont()
        #             font.setPointSize(9)
        #             font.setBold(False)
        #             font.setWeight(50)
        #             r.setFont(font)
        #             r.setObjectName(element[0])
        #             if element[0] in [KEY_SUPTNGSEC_DEPTH, KEY_SUPTNGSEC_FLANGE_W, KEY_SUPTNGSEC_FLANGE_T,
        #                               KEY_SUPTNGSEC_WEB_T, KEY_SUPTDSEC_DEPTH, KEY_SUPTDSEC_FLANGE_W,
        #                               KEY_SUPTDSEC_FLANGE_T, KEY_SUPTDSEC_WEB_T]:
        #                 r.setValidator(QDoubleValidator())
        #
        #         if type == TYPE_IMAGE:
        #             im = QtWidgets.QLabel(tab)
        #             im.setGeometry(QtCore.QRect(60 + j, 30 + i, 200, 300))
        #             im.setObjectName(element[0])
        #             im.setScaledContents(True)
        #             image = QPixmap("./ResourceFiles/images/Columns_Beams.png")
        #             im.setPixmap(image)
        #             i = i + 300
        #
        #         if type == TYPE_BREAK:
        #             j = j + 310
        #             i = -30
        #
        #         if type == TYPE_ENTER:
        #             pass
        #
        #         i = i + 30
        #     pushButton_Add = QtWidgets.QPushButton(tab)
        #     pushButton_Add.setObjectName(str("pushButton_Add_"+tab_name))
        #     pushButton_Add.setGeometry(QtCore.QRect(6, 500, 160, 27))
        #     font = QtGui.QFont()
        #     font.setPointSize(9)
        #     font.setBold(False)
        #     font.setWeight(50)
        #     pushButton_Add.setFont(font)
        #     pushButton_Add.setText("Add")
        #
        #     pushButton_Clear = QtWidgets.QPushButton(tab)
        #     pushButton_Clear.setObjectName(str("pushButton_Clear_"+tab_name))
        #     pushButton_Clear.setGeometry(QtCore.QRect(180, 500, 160, 27))
        #     font = QtGui.QFont()
        #     font.setPointSize(9)
        #     font.setBold(False)
        #     font.setWeight(50)
        #     pushButton_Clear.setFont(font)
        #     pushButton_Clear.setText("Clear")
        #
        #     pushButton_Import = QtWidgets.QPushButton(tab)
        #     pushButton_Import.setObjectName(str("pushButton_Import_"+tab_name))
        #     pushButton_Import.setGeometry(QtCore.QRect(770, 500, 160, 27))
        #     font = QtGui.QFont()
        #     font.setPointSize(9)
        #     font.setBold(False)
        #     font.setWeight(50)
        #     pushButton_Import.setFont(font)
        #     pushButton_Import.setText("Import xlsx file")
        #
        #     pushButton_Download = QtWidgets.QPushButton(tab)
        #     pushButton_Download.setObjectName(str("pushButton_Download_"+tab_name))
        #     pushButton_Download.setGeometry(QtCore.QRect(600, 500, 160, 27))
        #     font = QtGui.QFont()
        #     font.setPointSize(9)
        #     font.setBold(False)
        #     font.setWeight(50)
        #     pushButton_Download.setFont(font)
        #     pushButton_Download.setText("Download xlsx file")
        #
        #     self.tabWidget.addTab(tab, "")
        #     self.tabWidget.setTabText(tab_index, tab_name)
        #     tab_index += 1





#END

####################################################################
####################################################################
#####################################################################





        #
        #
        #
        # self.tab_Column = QtWidgets.QWidget()
        # self.tab_Column.setObjectName("tab_Column")
        #
        # '''
        # @author: Umair
        # '''
        #
        # supporting_section_list = main.supporting_section_values(self)
        # _translate = QtCore.QCoreApplication.translate
        # i = 0
        # j = 6
        # for element in supporting_section_list:
        #     lable = element[1]
        #     type = element[2]
        #     # value = option[4]
        #     if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
        #         l = QtWidgets.QLabel(self.tab_Column)
        #         if lable in [KEY_DISP_SUPTNGSEC_THERMAL_EXP]:
        #             l.setGeometry(QtCore.QRect(3 + j, 10 + i, 165, 28))
        #             i = i + 10
        #         else:
        #             l.setGeometry(QtCore.QRect(3 + j, 10 + i, 165, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         if lable in [KEY_DISP_SUPTNGSEC_DESIGNATION, KEY_DISP_SUPTNGSEC_TYPE, KEY_DISP_SUPTNGSEC_SOURCE]:
        #             font.setWeight(75)
        #         else:
        #             font.setWeight(50)
        #         l.setFont(font)
        #         l.setObjectName(element[0] + "_label")
        #         l.setText(_translate("MainWindow", "<html><head/><body><p>" + lable + "</p></body></html>"))
        #         l.setAlignment(QtCore.Qt.AlignCenter)
        #
        #     if type == TYPE_COMBOBOX:
        #         combo = QtWidgets.QComboBox(self.tab_Column)
        #         combo.setGeometry(QtCore.QRect(170 + j, 10 + i, 130, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         combo.setFont(font)
        #         combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
        #         combo.setMaxVisibleItems(5)
        #         combo.setObjectName(element[0])
        #         for item in element[3]:
        #             combo.addItem(item)
        #
        #     if type == TYPE_TITLE:
        #         q = QtWidgets.QLabel(self.tab_Column)
        #         q.setGeometry(QtCore.QRect(j, 10 + i, 155, 35))
        #         font = QtGui.QFont()
        #         font.setPointSize(10)
        #         q.setFont(font)
        #         q.setObjectName("_title")
        #         q.setText(_translate("MainWindow",
        #                              "<html><head/><body><p><span style=\" font-weight:600;\">" + lable + "</span></p></body></html>"))
        #
        #     if type == TYPE_TEXTBOX:
        #         r = QtWidgets.QLineEdit(self.tab_Column)
        #         r.setGeometry(QtCore.QRect(170 + j, 10 + i, 130, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         r.setFont(font)
        #         r.setObjectName(element[0])
        #         if element[0] in [KEY_SUPTNGSEC_DEPTH, KEY_SUPTNGSEC_FLANGE_W, KEY_SUPTNGSEC_FLANGE_T, KEY_SUPTNGSEC_WEB_T]:
        #             r.setValidator(QDoubleValidator())
        #
        #     if type == TYPE_IMAGE:
        #         im = QtWidgets.QLabel(self.tab_Column)
        #         im.setGeometry(QtCore.QRect(60 + j, 30 + i, 200, 300))
        #         im.setObjectName(element[0])
        #         im.setScaledContents(True)
        #         image = QPixmap("./ResourceFiles/images/Columns_Beams.png")
        #         im.setPixmap(image)
        #         i = i + 300
        #
        #     if type == TYPE_BREAK:
        #         j = j + 310
        #         i = -30
        #
        #     if type == TYPE_ENTER:
        #         pass
        #
        #     i = i + 30
        # pushButton_Add_Column = QtWidgets.QPushButton(self.tab_Column)
        # pushButton_Add_Column.setObjectName("pushButton_Add_Column")
        # pushButton_Add_Column.setGeometry(QtCore.QRect(6, 500, 160, 27))
        # font = QtGui.QFont()
        # font.setPointSize(9)
        # font.setBold(False)
        # font.setWeight(50)
        # pushButton_Add_Column.setFont(font)
        # pushButton_Add_Column.setText("Add")
        #
        # pushButton_Clear_Column = QtWidgets.QPushButton(self.tab_Column)
        # pushButton_Clear_Column.setObjectName("pushButton_Clear_Column")
        # pushButton_Clear_Column.setGeometry(QtCore.QRect(180, 500, 160, 27))
        # font = QtGui.QFont()
        # font.setPointSize(9)
        # font.setBold(False)
        # font.setWeight(50)
        # pushButton_Clear_Column.setFont(font)
        # pushButton_Clear_Column.setText("Clear")
        #
        # pushButton_Import_Column = QtWidgets.QPushButton(self.tab_Column)
        # pushButton_Import_Column.setObjectName("pushButton_Import_Column")
        # pushButton_Import_Column.setGeometry(QtCore.QRect(770, 500, 160, 27))
        # font = QtGui.QFont()
        # font.setPointSize(9)
        # font.setBold(False)
        # font.setWeight(50)
        # pushButton_Import_Column.setFont(font)
        # pushButton_Import_Column.setText("Import xlsx file")
        #
        # pushButton_Download_Column = QtWidgets.QPushButton(self.tab_Column)
        # pushButton_Download_Column.setObjectName("pushButton_Download_Column")
        # pushButton_Download_Column.setGeometry(QtCore.QRect(600, 500, 160, 27))
        # font = QtGui.QFont()
        # font.setPointSize(9)
        # font.setBold(False)
        # font.setWeight(50)
        # pushButton_Download_Column.setFont(font)
        # pushButton_Download_Column.setText("Download xlsx file")
        #
        # self.tabWidget.addTab(self.tab_Column, "")
        # self.tab_Beam = QtWidgets.QWidget()
        # self.tab_Beam.setObjectName("tab_Beam")
        #
        # supported_section_list = main.supported_section_values(self)
        # _translate = QtCore.QCoreApplication.translate
        # i = 0
        # j = 6
        # for element in supported_section_list:
        #     lable = element[1]
        #     type = element[2]
        #     # value = option[4]
        #     if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
        #         l = QtWidgets.QLabel(self.tab_Beam)
        #         if lable in [KEY_DISP_SUPTDSEC_THERMAL_EXP]:
        #             l.setGeometry(QtCore.QRect(3 + j, 10 + i, 165, 28))
        #             i = i + 10
        #         else:
        #             l.setGeometry(QtCore.QRect(3 + j, 10 + i, 165, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         if lable in [KEY_DISP_SUPTNGSEC_DESIGNATION, KEY_DISP_SUPTNGSEC_TYPE, KEY_DISP_SUPTNGSEC_SOURCE]:
        #             font.setWeight(75)
        #         else:
        #             font.setWeight(50)
        #         l.setFont(font)
        #         l.setObjectName(element[0] + "_label")
        #         l.setText(_translate("MainWindow", "<html><head/><body><p>" + lable + "</p></body></html>"))
        #         l.setAlignment(QtCore.Qt.AlignCenter)
        #
        #     if type == TYPE_COMBOBOX:
        #         combo = QtWidgets.QComboBox(self.tab_Beam)
        #         combo.setGeometry(QtCore.QRect(170 + j, 10 + i, 130, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         combo.setFont(font)
        #         combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
        #         combo.setMaxVisibleItems(5)
        #         combo.setObjectName(element[0])
        #         for item in element[3]:
        #             combo.addItem(item)
        #
        #     if type == TYPE_TITLE:
        #         q = QtWidgets.QLabel(self.tab_Beam)
        #         q.setGeometry(QtCore.QRect(j, 10 + i, 155, 35))
        #         font = QtGui.QFont()
        #         font.setPointSize(10)
        #         q.setFont(font)
        #         q.setObjectName("_title")
        #         q.setText(_translate("MainWindow",
        #                              "<html><head/><body><p><span style=\" font-weight:600;\">" + lable + "</span></p></body></html>"))
        #
        #     if type == TYPE_TEXTBOX:
        #         r = QtWidgets.QLineEdit(self.tab_Beam)
        #         r.setGeometry(QtCore.QRect(170 + j, 10 + i, 130, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         r.setFont(font)
        #         r.setObjectName(element[0])
        #         if element[0] in [KEY_SUPTDSEC_DEPTH, KEY_SUPTDSEC_FLANGE_W, KEY_SUPTDSEC_FLANGE_T,
        #                        KEY_SUPTDSEC_WEB_T]:
        #             r.setValidator(QDoubleValidator())
        #
        #     if type == TYPE_IMAGE:
        #         im = QtWidgets.QLabel(self.tab_Beam)
        #         im.setGeometry(QtCore.QRect(60 + j, 30 + i, 200, 300))
        #         im.setObjectName(element[0])
        #         im.setScaledContents(True)
        #         image = QPixmap("./ResourceFiles/images/Columns_Beams.png")
        #         im.setPixmap(image)
        #         i = i + 300
        #
        #     if type == TYPE_BREAK:
        #         j = j + 310
        #         i = -30
        #
        #     if type == TYPE_ENTER:
        #         pass
        #
        #     i = i + 30
        # pushButton_Add_Beam = QtWidgets.QPushButton(self.tab_Beam)
        # pushButton_Add_Beam.setObjectName("pushButton_Add_Beam")
        # pushButton_Add_Beam.setGeometry(QtCore.QRect(6, 500, 160, 27))
        # font = QtGui.QFont()
        # font.setPointSize(9)
        # font.setBold(False)
        # font.setWeight(50)
        # pushButton_Add_Beam.setFont(font)
        # pushButton_Add_Beam.setText("Add")
        #
        # pushButton_Clear_Beam = QtWidgets.QPushButton(self.tab_Beam)
        # pushButton_Clear_Beam.setObjectName("pushButton_Clear_Beam")
        # pushButton_Clear_Beam.setGeometry(QtCore.QRect(180, 500, 160, 27))
        # font = QtGui.QFont()
        # font.setPointSize(9)
        # font.setBold(False)
        # font.setWeight(50)
        # pushButton_Clear_Beam.setFont(font)
        # pushButton_Clear_Beam.setText("Clear")
        #
        # pushButton_Import_Beam = QtWidgets.QPushButton(self.tab_Beam)
        # pushButton_Import_Beam.setObjectName("pushButton_Import_Beam")
        # pushButton_Import_Beam.setGeometry(QtCore.QRect(770, 500, 160, 27))
        # font = QtGui.QFont()
        # font.setPointSize(9)
        # font.setBold(False)
        # font.setWeight(50)
        # pushButton_Import_Beam.setFont(font)
        # pushButton_Import_Beam.setText("Import xlsx file")
        #
        # pushButton_Download_Beam = QtWidgets.QPushButton(self.tab_Beam)
        # pushButton_Download_Beam.setObjectName("pushButton_Download_Beam")
        # pushButton_Download_Beam.setGeometry(QtCore.QRect(600, 500, 160, 27))
        # font = QtGui.QFont()
        # font.setPointSize(9)
        # font.setBold(False)
        # font.setWeight(50)
        # pushButton_Download_Beam.setFont(font)
        # pushButton_Download_Beam.setText("Download xlsx file")
        #
        # self.tabWidget.addTab(self.tab_Beam, "")

###################################################################
######################################################################
######################################################################
####################################################################

# START

        tab_index = 0
        dialog_height = 500
        buttons = []
        for tab_details in main.tab_list(main):
            tab_name = tab_details[0]
            tab_elements = tab_details[2]
            tab_type = tab_details[1]
            if tab_type == TYPE_TAB_1:
                tab = QtWidgets.QWidget()
                tab.setObjectName(tab_name)
                elements = tab_elements(main, input_dictionary)
                _translate = QtCore.QCoreApplication.translate
                i = 0
                j = 6
                labels = []
                combo_text = []
                for element in elements:
                    lable = element[1]
                    type = element[2]
                    # value = option[4]
                    if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
                        l = QtWidgets.QLabel(tab)
                        l.move(3 + j, 10 + i)
                        font = QtGui.QFont()
                        font.setPointSize(9)
                        if lable in [KEY_DISP_DESIGNATION, 'Type', 'Source']:
                            font.setWeight(75)
                        else:
                            font.setWeight(50)
                        l.setFont(font)
                        l.setObjectName(element[0] + "_label")
                        l.setText(_translate("MainWindow", "<html><head/><body><p>" + lable + "</p></body></html>"))
                        l.setAlignment(QtCore.Qt.AlignLeft)
                        l.resize(l.sizeHint().width(), l.sizeHint().height())
                        labels.append((l, l.sizeHint().width()+3+j))
                    if type == TYPE_COMBOBOX:
                        combo = QtWidgets.QComboBox(tab)
                        combo.setGeometry(QtCore.QRect(l.sizeHint().width() + 10 + j, 10 + i, 130, 22))
                        font = QtGui.QFont()
                        font.setPointSize(9)
                        font.setBold(False)
                        font.setWeight(50)
                        combo.setFont(font)
                        combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
                        combo.setMaxVisibleItems(5)
                        combo.setObjectName(element[0])
                        for item in element[3]:
                            combo.addItem(item)
                        if input_dictionary:
                            combo.setCurrentText(str(element[4]))
                        combo.view().setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
                        width = combo.minimumSizeHint().width()
                        combo.view().setMinimumWidth(width)
                        combo_text.append((combo,l.sizeHint().width() + 10 + j,10+i))
                    if type == TYPE_TITLE:
                        q = QtWidgets.QLabel(tab)
                        q.setGeometry(QtCore.QRect(j, 10 + i, 155, 35))
                        font = QtGui.QFont()
                        font.setPointSize(10)
                        q.setFont(font)
                        q.setObjectName("_title")
                        q.setText(_translate("MainWindow",
                                             "<html><head/><body><p><span style=\" font-weight:600;\">" + lable + "</span></p></body></html>"))
                        q.resize(q.sizeHint().width(), q.sizeHint().height())

                    if type == TYPE_TEXTBOX:
                        r = QtWidgets.QLineEdit(tab)
                        r.setGeometry(QtCore.QRect(QtCore.QRect(l.sizeHint().width() + 10 + j, 10 + i, 130, 22)))
                        font = QtGui.QFont()
                        font.setPointSize(9)
                        font.setBold(False)
                        font.setWeight(50)
                        r.setFont(font)
                        r.setObjectName(element[0])
                        if element[0] in ['Label_1', 'Label_2', 'Label_3', 'Label_4']:
                            r.setValidator(QDoubleValidator())
                        if input_dictionary:
                            r.setText(str(element[4]))
                        combo_text.append((r,l.sizeHint().width() + 10 + j,10+i))
                    if type == TYPE_IMAGE:
                        im = QtWidgets.QLabel(tab)
                        im.setGeometry(QtCore.QRect(60 + j, 30 + i, 200, 300))
                        im.setObjectName(element[0])
                        im.setScaledContents(True)
                        image = QPixmap(element[4])
                        im.setPixmap(image)
                        i = i + 300
                        im.resize(im.sizeHint().width(), im.sizeHint().height())

                    if type == TYPE_BREAK:
                        ki = -1
                        for item,size in labels:
                            item.resize(size,item.sizeHint().height())
                            ki = max(ki,size)
                        labels =[]

                        for item in combo_text:
                            x,y = item[1], item[2]
                            item[0].move(ki+10,y)
                        combo_text = []
                        dialog_height = max(dialog_height, i)
                        j = j + 400
                        i = -30

                    if type == TYPE_ENTER:
                        pass

                    i = i + 30

                pushButton_Add = QtWidgets.QPushButton(tab)
                pushButton_Add.setObjectName(str("pushButton_Add_" + tab_name))
                pushButton_Add.setGeometry(QtCore.QRect(200, i + 70, 160, 27))
                font = QtGui.QFont()
                font.setPointSize(9)
                font.setBold(False)
                font.setWeight(50)
                pushButton_Add.setFont(font)
                pushButton_Add.setText("Add")
                buttons.append((pushButton_Add,200,i+70))
                #pushButton_Add.resize(pushButton_Add.sizeHint().width() + 10, pushButton_Add.sizeHint().height() + 7)

                pushButton_Clear = QtWidgets.QPushButton(tab)
                pushButton_Clear.setObjectName(str("pushButton_Clear_" + tab_name))
                pushButton_Clear.setGeometry(QtCore.QRect(400, i + 70, 160, 27))
                font = QtGui.QFont()
                font.setPointSize(9)
                font.setBold(False)
                font.setWeight(50)
                pushButton_Clear.setFont(font)
                pushButton_Clear.setText("Clear")
                buttons.append((pushButton_Clear,400,i+70))
                #pushButton_Clear.resize(pushButton_Clear.sizeHint().width() + 10, pushButton_Clear.sizeHint().height() + 7)

                pushButton_Import = QtWidgets.QPushButton(tab)
                pushButton_Import.setObjectName(str("pushButton_Import_" + tab_name))
                pushButton_Import.setGeometry(QtCore.QRect(600, i + 70, 160, 27))
                font = QtGui.QFont()
                font.setPointSize(9)
                font.setBold(False)
                font.setWeight(50)
                pushButton_Import.setFont(font)
                pushButton_Import.setText("Import xlsx file")
                buttons.append((pushButton_Import,600,i+70))
                #pushButton_Import.resize(pushButton_Import.sizeHint().width() + 10, pushButton_Import.sizeHint().height() + 7)

                pushButton_Download = QtWidgets.QPushButton(tab)
                pushButton_Download.setObjectName(str("pushButton_Download_" + tab_name))
                pushButton_Download.setGeometry(QtCore.QRect(800, i + 70, 160, 27))
                font = QtGui.QFont()
                font.setPointSize(9)
                font.setBold(False)
                font.setWeight(50)
                pushButton_Download.setFont(font)
                pushButton_Download.setText("Download xlsx file")
                buttons.append((pushButton_Download,800,i+70))
                #pushButton_Download.resize(pushButton_Download.sizeHint().width() + 10, pushButton_Download.sizeHint().height() + 7)

                if combo_text and labels:
                    ki = -1
                    for item,size in labels:
                        item.resize(size,item.sizeHint().height())
                        ki = max(ki,size)

                    for item in combo_text:
                        x,y = item[1], item[2]
                        item[0].move(ki+10,y)

                self.tabWidget.addTab(tab, "")
                self.tabWidget.setTabText(tab_index, tab_name)
                tab_index += 1
                #dialog_height = max(i+70, dialog_height)
            elif tab_type == TYPE_TAB_2:
                labels = []
                combo_text = []
                tab = QtWidgets.QWidget()
                tab.setObjectName(tab_name)
                elements = tab_elements(main, input_dictionary)
                label_1 = QtWidgets.QLabel(tab)
                font = QtGui.QFont()
                font.setFamily("Arial")
                font.setWeight(75)
                label_1.setFont(font)
                label_1.setObjectName("label_1")
                label_1.setGeometry(QtCore.QRect(10, 10, 130, 22))
                label_1.setText("Inputs")
                _translate = QtCore.QCoreApplication.translate
                i = 30
                j = 6
                for element in elements:
                    lable = element[1]
                    type = element[2]
                    # value = option[4]
                    if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
                        l = QtWidgets.QLabel(tab)
                        if lable in [KEY_DISP_DP_DETAILING_GAP, KEY_DISP_DP_DETAILING_CORROSIVE_INFLUENCES,
                                     KEY_DISP_DP_ANCHOR_BOLT_FRICTION]:
                            l.setGeometry(QtCore.QRect(3 + j, 10 + i, 165, 28))
                            i = i + 10
                        else:
                            l.setGeometry(QtCore.QRect(3 + j, 10 + i, 165, 22))
                        font = QtGui.QFont()
                        font.setPointSize(9)
                        font.setWeight(50)
                        l.setFont(font)
                        l.setObjectName(element[0] + "_label")
                        l.setText(_translate("MainWindow", "<html><head/><body><p>" + lable + "</p></body></html>"))
                        l.setAlignment(QtCore.Qt.AlignLeft)
                        l.resize(l.sizeHint().width(), l.sizeHint().width())
                        labels.append((l, l.sizeHint().width()+3+j))
                    if type == TYPE_COMBOBOX:
                        combo = QtWidgets.QComboBox(tab)
                        combo.setGeometry(QtCore.QRect(l.sizeHint().width() + 10 + j, 10 + i, 270, 22))
                        font = QtGui.QFont()
                        font.setPointSize(9)
                        font.setBold(False)
                        font.setWeight(50)
                        combo.setFont(font)
                        combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
                        combo.setMaxVisibleItems(5)
                        combo.setObjectName(element[0])
                        for item in element[3]:
                            combo.addItem(item)
                        # if element[0] == KEY_DP_BOLT_SLIP_FACTOR:
                        #     combo.setCurrentIndex(4)
                        if element[0] == KEY_DP_DESIGN_METHOD:
                            combo.model().item(1).setEnabled(False)
                            combo.model().item(2).setEnabled(False)
                        if input_dictionary:
                            combo.setCurrentText(str(element[4]))
                        combo.view().setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
                        width = combo.minimumSizeHint().width()
                        combo.view().setMinimumWidth(width)
                        combo_text.append((combo,l.sizeHint().width() + 10 + j,10+i))
                    if type == TYPE_TITLE:
                        q = QtWidgets.QLabel(tab)
                        q.setGeometry(QtCore.QRect(j, 10 + i, 155, 35))
                        font = QtGui.QFont()
                        font.setPointSize(10)
                        q.setFont(font)
                        q.setObjectName("_title")
                        q.setText(_translate("MainWindow",
                                             "<html><head/><body><p><span style=\" font-weight:600;\">" + lable + "</span></p></body></html>"))
                    if type == TYPE_NOTE:
                        q = QtWidgets.QLabel(tab)
                        q.setGeometry(QtCore.QRect(j, 55 + i, 355, 35))
                        font = QtGui.QFont()
                        font.setPointSize(10)
                        q.setFont(font)
                        q.setObjectName("_title")
                        q.setText(_translate("MainWindow",
                                             "<html><head/><body><p><span style=\" font-weight:600;\">" + lable + "</span></p></body></html>"))
                        q.resize(q.sizeHint().width(),q.sizeHint().height())
                    if type == TYPE_TEXTBOX:
                        r = QtWidgets.QLineEdit(tab)
                        r.setGeometry(QtCore.QRect(l.sizeHint().width() + 10 + j, 10 + i, 270, 22))
                        font = QtGui.QFont()
                        font.setPointSize(9)
                        font.setBold(False)
                        font.setWeight(50)
                        r.setFont(font)
                        r.setObjectName(element[0])
                        if element[3]:
                            r.setText(element[3])
                        dbl_validator = QDoubleValidator()
                        if element[0] in [KEY_DP_BOLT_MATERIAL_G_O, KEY_DP_WELD_MATERIAL_G_O]:
                            r.setValidator(dbl_validator)
                            r.setMaxLength(7)
                        if input_dictionary:
                            r.setText(str(element[4]))
                        combo_text.append((r,l.sizeHint().width() + 10 + j,10+i))
                    if type == TYPE_IMAGE:
                        im = QtWidgets.QLabel(tab)
                        im.setGeometry(QtCore.QRect(60 + j, 30 + i, 200, 300))
                        im.setObjectName(element[0])
                        im.setScaledContents(True)
                        image = QPixmap("./ResourceFiles/images/Columns_Beams.png")
                        im.setPixmap(image)
                        im.resize(im.sizeHint().width(),im.sizeHint().height())
                        i = i + 300

                    if type == TYPE_BREAK:

                        ki = -1
                        for item,size in labels:
                            item.resize(size,item.sizeHint().height())
                            ki = max(ki,size)
                        labels =[]

                        for item in combo_text:
                            x,y = item[1], item[2]
                            item[0].move(ki+10,y)
                        combo_text = []

                        j = j + 400
                        i = -30

                    if type == TYPE_ENTER:
                        i = i + 100

                    if type == TYPE_TEXT_BROWSER:
                        label_3 = QtWidgets.QLabel(tab)
                        font = QtGui.QFont()
                        font.setFamily("Arial")
                        font.setWeight(75)
                        label_3.setFont(font)
                        label_3.setObjectName("label_3")
                        label_3.setGeometry(QtCore.QRect(650, 10, 130, 22))
                        label_3.setText("Description")
                        textBrowser = QtWidgets.QTextBrowser(tab)
                        textBrowser.setMinimumSize(QtCore.QSize(210, 320))
                        textBrowser.setObjectName(element[0])
                        textBrowser.setGeometry(QtCore.QRect(650, 40, 480, 450))
                        textBrowser.setHtml(_translate("DesignPreferences", element[3]))
                        textBrowser.horizontalScrollBar().setVisible(False)

                    i = i + 30

                if combo_text and labels:
                    ki = -1
                    for item,size in labels:
                        item.resize(size,item.sizeHint().height())
                        ki = max(ki,size)

                    for item in combo_text:
                        x,y = item[1], item[2]
                        item[0].move(ki+10,y)

                self.tabWidget.addTab(tab, "")
                self.tabWidget.setTabText(tab_index, tab_name)
                tab_index += 1
        dialog_height += 170   # 70 for buttons and 100 for whitespaces
        DesignPreferences.resize(1170,dialog_height)  # Width of Design Preference Dialog is not set automatically. You have to adjust the width manually.
        for item in buttons:
            item[0].move(item[1],dialog_height-150)

# END

####################################################################
####################################################################
#####################################################################
        # self.tab_Bolt = QtWidgets.QWidget()
        # self.tab_Bolt.setObjectName("tab_Bolt")
        #
        # label_1 = QtWidgets.QLabel(self.tab_Bolt)
        # font = QtGui.QFont()
        # font.setFamily("Arial")
        # font.setWeight(75)
        # label_1.setFont(font)
        # label_1.setObjectName("label_1")
        # label_1.setGeometry(QtCore.QRect(10, 10, 130, 22))
        # label_1.setText("Inputs")
        # label_3 = QtWidgets.QLabel(self.tab_Bolt)
        # font = QtGui.QFont()
        # font.setFamily("Arial")
        # font.setWeight(75)
        # label_3.setFont(font)
        # label_3.setObjectName("label_3")
        # label_3.setGeometry(QtCore.QRect(400, 10, 130, 22))
        # label_3.setText("Description")
        # label_4 = QtWidgets.QLabel(self.tab_Bolt)
        # font = QtGui.QFont()
        # font.setFamily("Arial")
        # font.setWeight(75)
        # label_4.setFont(font)
        # label_4.setObjectName("label_4")
        # label_4.setGeometry(QtCore.QRect(10, 400, 400, 50))
        # label_4.setText("NOTE : If slip is permitted under the design load, design the bolt as"
        #                 "<br>a bearing bolt and select corresponding bolt grade.")
        # textBrowser = QtWidgets.QTextBrowser(self.tab_Bolt)
        # textBrowser.setMinimumSize(QtCore.QSize(210, 320))
        # textBrowser.setObjectName("textBrowser")
        # textBrowser.setGeometry(QtCore.QRect(400, 40, 520, 450))
        # textBrowser.setHtml(_translate("DesignPreferences",
        #                                     "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
        #                                     "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
        #                                     "p, li { white-space: pre-wrap; }\n"
        #                                     "</style></head><body style=\" font-family:\'Arial\'; font-size:8.25pt; font-weight:400; font-style:normal;\">\n"
        #                                     "<table border=\"0\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px;\" cellspacing=\"2\" cellpadding=\"0\">\n"
        #                                     "<tr>\n"
        #                                     "<td colspan=\"3\">\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">IS 800 Table 20 Typical Average Values for Coefficient of Friction (</span><span style=\" font-family:\'Calibri,sans-serif\'; font-size:9pt;\">µ</span><span style=\" font-family:\'Calibri,sans-serif\'; font-size:9pt; vertical-align:sub;\">f</span><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">)</span></p></td></tr></table>\n"
        #                                     "<p align=\"justify\" style=\"-qt-paragraph-type:empty; margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS Shell Dlg 2\'; font-size:8pt;\"><br /></p>\n"
        #                                     "<table border=\"0\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px;\" cellspacing=\"2\" cellpadding=\"0\">\n"
        #                                     "<tr>\n"
        #                                     "<td width=\"26\"></td>\n"
        #                                     "<td width=\"383\">\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Treatment of Surfaces</span></p></td>\n"
        #                                     "<td width=\"78\">\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  µ_f</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">i)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Surfaces not treated</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.2</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">ii)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Surfaces blasted with short or grit with any loose rust removed, no pitting</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.5</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">iii)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Surfaces blasted with short or grit and hot-dip galvanized</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.1</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">iv)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Surfaces blasted with short or grit and spray - metallized with zinc (thickness 50-70 µm)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.25</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">v)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Surfaces blasted with shot or grit and painted with ethylzinc silicate coat (thickness 30-60 µm)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.3</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">vi)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Sand blasted surface, after light rusting</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.52</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">vii)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Surfaces blasted with shot or grit and painted with ethylzinc silicate coat (thickness 60-80 µm)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.3</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">viii)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Surfaces blasted with shot or grit and painted with alcalizinc silicate coat (thickness 60-80 µm)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.3</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">ix)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Surfaces blasted with shot or grit and spray metallized with aluminium (thickness &gt;50 µm)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.5</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">x)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Clean mill scale</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.33</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">xi)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Sand blasted surface</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.48</span></p></td></tr>\n"
        #                                     "<tr>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">xii)</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Red lead painted surface</span></p></td>\n"
        #                                     "<td>\n"
        #                                     "<p align=\"justify\" style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">  0.1</span></p>\n"
        #                                     "<p align=\"justify\" style=\"-qt-paragraph-type:empty; margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS Shell Dlg 2\'; font-size:8pt;\"><br /></p></td></tr></table></body></html>"))
        #
        # bolt_list = Connection.bolt_values()
        # _translate = QtCore.QCoreApplication.translate
        # i = 40
        # for element in bolt_list:
        #     lable = element[1]
        #     type = element[2]
        #     # value = option[4]
        #     if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
        #         l = QtWidgets.QLabel(self.tab_Bolt)
        #         l.setGeometry(QtCore.QRect(6, 10 + i, 185, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setWeight(50)
        #         l.setFont(font)
        #         l.setObjectName(element[0] + "_label")
        #         l.setText(_translate("MainWindow", "<html><head/><body><p>" + lable + "</p></body></html>"))
        #         l.setAlignment(QtCore.Qt.AlignCenter)
        #
        #     if type == TYPE_COMBOBOX:
        #         combo = QtWidgets.QComboBox(self.tab_Bolt)
        #         combo.setGeometry(QtCore.QRect(230, 10 + i, 130, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         combo.setFont(font)
        #         combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
        #         combo.setMaxVisibleItems(5)
        #         combo.setObjectName(element[0])
        #         for item in element[3]:
        #             combo.addItem(item)
        #         if element[0] == KEY_DP_BOLT_SLIP_FACTOR:
        #             combo.setCurrentIndex(4)
        #
        #     if type == TYPE_TITLE:
        #         q = QtWidgets.QLabel(self.tab_Bolt)
        #         q.setGeometry(QtCore.QRect(3, 10 + i, 300, 35))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         q.setFont(font)
        #         q.setObjectName("_title")
        #         q.setText(_translate("MainWindow",
        #                              "<html><head/><body><p><span style=\" font-weight:600;\">" + lable + "</span></p></body></html>"))
        #
        #     if type == TYPE_TEXTBOX:
        #         r = QtWidgets.QLineEdit(self.tab_Bolt)
        #         r.setGeometry(QtCore.QRect(230, 10 + i, 130, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         r.setFont(font)
        #         r.setObjectName(element[0])
        #         if element[3]:
        #             r.setText(element[3])
        #         dbl_validator = QDoubleValidator()
        #         if element[0] == KEY_DP_BOLT_MATERIAL_G_O:
        #             r.setValidator(dbl_validator)
        #             r.setMaxLength(7)
        #
        #     if type == TYPE_ENTER:
        #         i = i + 100
        #
        #     i = i + 30
        #
        # self.tabWidget.addTab(self.tab_Bolt, "")
        # self.tab_Weld = QtWidgets.QWidget()
        # self.tab_Weld.setObjectName("tab_Weld")
        #
        # label_1 = QtWidgets.QLabel(self.tab_Weld)
        # font = QtGui.QFont()
        # font.setFamily("Arial")
        # font.setWeight(75)
        # label_1.setFont(font)
        # label_1.setObjectName("label_1")
        # label_1.setGeometry(QtCore.QRect(10, 10, 130, 22))
        # label_1.setText("Inputs")
        # label_3 = QtWidgets.QLabel(self.tab_Weld)
        # font = QtGui.QFont()
        # font.setFamily("Arial")
        # font.setWeight(75)
        # label_3.setFont(font)
        # label_3.setObjectName("label_3")
        # label_3.setGeometry(QtCore.QRect(400, 10, 130, 22))
        # label_3.setText("Description")
        # textBrowser = QtWidgets.QTextBrowser(self.tab_Weld)
        # textBrowser.setMinimumSize(QtCore.QSize(210, 320))
        # textBrowser.setObjectName("textBrowser")
        # textBrowser.setGeometry(QtCore.QRect(400, 40, 520, 450))
        # textBrowser.setHtml(_translate("DesignPreferences",
        #                                "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
        #                                "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
        #                                "p, li { white-space: pre-wrap; }\n"
        #                                "</style></head><body style=\" font-family:\'Arial\'; font-size:8.25pt; font-weight:400; font-style:normal;\">\n"
        #                                "<p align=\"justify\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Shop weld takes a material safety factor of 1.25</span></p>\n"
        #                                "<p align=\"justify\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Field weld takes a material safety factor of 1.5</span></p>\n"
        #                                "<p align=\"justify\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">(IS 800 - cl. 5. 4. 1 or Table 5)</span></p></body></html>"))
        # weld_list = Connection.weld_values()
        # _translate = QtCore.QCoreApplication.translate
        # i = 40
        # for element in weld_list:
        #     lable = element[1]
        #     type = element[2]
        #     if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
        #         l = QtWidgets.QLabel(self.tab_Weld)
        #         l.setGeometry(QtCore.QRect(6, 10 + i, 185, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setWeight(50)
        #         l.setFont(font)
        #         l.setObjectName(element[0] + "_label")
        #         l.setText(_translate("MainWindow", "<html><head/><body><p>" + lable + "</p></body></html>"))
        #         l.setAlignment(QtCore.Qt.AlignCenter)
        #
        #     if type == TYPE_COMBOBOX:
        #         combo = QtWidgets.QComboBox(self.tab_Weld)
        #         combo.setGeometry(QtCore.QRect(230, 10 + i, 130, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         combo.setFont(font)
        #         combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
        #         combo.setMaxVisibleItems(5)
        #         combo.setObjectName(element[0])
        #         for item in element[3]:
        #             combo.addItem(item)
        #
        #     if type == TYPE_TEXTBOX:
        #         r = QtWidgets.QLineEdit(self.tab_Weld)
        #         r.setGeometry(QtCore.QRect(230, 10 + i, 130, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         r.setFont(font)
        #         r.setObjectName(element[0])
        #         if element[3]:
        #             r.setText(element[3])
        #         dbl_validator = QDoubleValidator()
        #         if element[0] == KEY_DP_WELD_MATERIAL_G_O:
        #             r.setValidator(dbl_validator)
        #             r.setMaxLength(7)
        #
        #     i = i + 40
        # self.tabWidget.addTab(self.tab_Weld, "")
        # self.tab_Detailing = QtWidgets.QWidget()
        # self.tab_Detailing.setObjectName("tab_Detailing")
        #
        # label_1 = QtWidgets.QLabel(self.tab_Detailing)
        # font = QtGui.QFont()
        # font.setFamily("Arial")
        # font.setWeight(75)
        # label_1.setFont(font)
        # label_1.setObjectName("label_1")
        # label_1.setGeometry(QtCore.QRect(10, 10, 130, 22))
        # label_1.setText("Inputs")
        # label_3 = QtWidgets.QLabel(self.tab_Detailing)
        # font = QtGui.QFont()
        # font.setFamily("Arial")
        # font.setWeight(75)
        # label_3.setFont(font)
        # label_3.setObjectName("label_3")
        # label_3.setGeometry(QtCore.QRect(470, 10, 130, 22))
        # label_3.setText("Description")
        # textBrowser = QtWidgets.QTextBrowser(self.tab_Detailing)
        # textBrowser.setMinimumSize(QtCore.QSize(210, 320))
        # textBrowser.setObjectName("textBrowser")
        # textBrowser.setGeometry(QtCore.QRect(470, 40, 450, 450))
        # textBrowser.setHtml(_translate("DesignPreferences",
        #                                "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
        #                                "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
        #                                "p, li { white-space: pre-wrap; }\n"
        #                                "</style></head><body style=\" font-family:\'Arial\'; font-size:8.25pt; font-weight:400; font-style:normal;\">\n"
        #                                "<p align=\"justify\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">The minimum edge and end distances from the centre of any hole to the nearest edge of a plate shall not be less than </span><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt; font-weight:600;\">1.7</span><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\"> times the hole diameter in case of </span><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt; font-weight:600;\">[a- sheared or hand flame cut edges] </span><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">and </span><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt; font-weight:600;\">1.5 </span><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">times the hole diameter in case of </span><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt; font-weight:600;\">[b - Rolled, machine-flame cut, sawn and planed edges]</span><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\"> (IS 800 - cl. 10. 2. 4. 2)</span></p>\n"
        #                                "<p align=\"justify\" style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'Calibri\'; font-size:8pt; vertical-align:middle;\"><br /></p>\n"
        #                                "<p align=\"justify\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">This gap should include the tolerance value of 5mm. So if the assumed clearance is 5mm, then the gap should be = 10mm (= 5mm {clearance} + 5 mm{tolerance})</span></p>\n"
        #                                "<p align=\"justify\" style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'Calibri\'; font-size:8pt;\"><br /></p>\n"
        #                                "<p align=\"justify\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt;\">Specifying whether the members are exposed to corrosive influences, here, only affects the calculation of the maximum edge distance as per cl. 10.2.4.3</span></p>\n"
        #                                "<p align=\"justify\" style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-family:\'MS Shell Dlg 2\'; font-size:8pt;\"><br /></p></body></html>"))
        #
        # detailing_list = Connection.detailing_values()
        # _translate = QtCore.QCoreApplication.translate
        # i = 40
        # for element in detailing_list:
        #     lable = element[1]
        #     type = element[2]
        #     # value = option[4]
        #     if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
        #         l = QtWidgets.QLabel(self.tab_Detailing)
        #         l.setGeometry(QtCore.QRect(6, 10 + i, 174, 30))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setWeight(50)
        #         l.setFont(font)
        #         l.setObjectName(element[0] + "_label")
        #         l.setText(_translate("MainWindow", "<html><head/><body><p>" + lable + "</p></body></html>"))
        #         l.setAlignment(QtCore.Qt.AlignCenter)
        #
        #     if type == TYPE_COMBOBOX:
        #         combo = QtWidgets.QComboBox(self.tab_Detailing)
        #         combo.setGeometry(QtCore.QRect(180, 10 + i, 270, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         combo.setFont(font)
        #         combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
        #         combo.setMaxVisibleItems(5)
        #         combo.setObjectName(element[0])
        #         for item in element[3]:
        #             combo.addItem(item)
        #
        #     if type == TYPE_TITLE:
        #         q = QtWidgets.QLabel(self.tab_Detailing)
        #         q.setGeometry(QtCore.QRect(3, 10 + i, 300, 35))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         q.setFont(font)
        #         q.setObjectName("_title")
        #         q.setText(_translate("MainWindow",
        #                              "<html><head/><body><p><span style=\" font-weight:600;\">" + lable + "</span></p></body></html>"))
        #
        #     if type == TYPE_TEXTBOX:
        #         r = QtWidgets.QLineEdit(self.tab_Detailing)
        #         r.setGeometry(QtCore.QRect(180, 10 + i, 270, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         r.setFont(font)
        #         r.setObjectName(element[0])
        #         if element[3]:
        #             r.setText(element[3])
        #
        #     i = i + 40
        #
        # self.tabWidget.addTab(self.tab_Detailing, "")
        # self.tab_Design = QtWidgets.QWidget()
        # self.tab_Design.setObjectName("tab_Design")
        #
        # design_list = Connection.design_values()
        # _translate = QtCore.QCoreApplication.translate
        # i = 40
        # for element in design_list:
        #     lable = element[1]
        #     type = element[2]
        #     # value = option[4]
        #     if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
        #         l = QtWidgets.QLabel(self.tab_Design)
        #         l.setGeometry(QtCore.QRect(6, 10 + i, 174, 30))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setWeight(50)
        #         l.setFont(font)
        #         l.setObjectName(element[0] + "_label")
        #         l.setText(_translate("MainWindow", "<html><head/><body><p>" + lable + "</p></body></html>"))
        #         l.setAlignment(QtCore.Qt.AlignCenter)
        #
        #     if type == TYPE_COMBOBOX:
        #         combo = QtWidgets.QComboBox(self.tab_Design)
        #         combo.setGeometry(QtCore.QRect(180, 10 + i, 270, 22))
        #         font = QtGui.QFont()
        #         font.setPointSize(9)
        #         font.setBold(False)
        #         font.setWeight(50)
        #         combo.setFont(font)
        #         combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
        #         combo.setMaxVisibleItems(5)
        #         combo.setObjectName(element[0])
        #         for item in element[3]:
        #             combo.addItem(item)
        #         if element[0] == KEY_DP_DESIGN_METHOD:
        #             combo.model().item(1).setEnabled(False)
        #             combo.model().item(2).setEnabled(False)
        #
        # self.tabWidget.addTab(self.tab_Design, "")
        self.gridLayout_5.addWidget(self.tabWidget, 0, 0, 1, 1)

        self.retranslateUi(DesignPreferences)
        self.tabWidget.setCurrentIndex(2)
        QtCore.QMetaObject.connectSlotsByName(DesignPreferences)
        module = main.module_name(main)

        if module in [KEY_DISP_FINPLATE, KEY_DISP_ENDPLATE, KEY_DISP_CLEATANGLE, KEY_DISP_SEATED_ANGLE, KEY_DISP_BCENDPLATE]:

            pushButton_Clear_Column = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_COLSEC)
            pushButton_Clear_Column.clicked.connect(lambda: self.clear_tab("Column"))
            pushButton_Add_Column = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_COLSEC)
            pushButton_Add_Column.clicked.connect(self.add_tab_column)
            pushButton_Clear_Beam = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_BEAMSEC)
            pushButton_Clear_Beam.clicked.connect(lambda: self.clear_tab("Beam"))
            pushButton_Add_Beam = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_BEAMSEC)
            pushButton_Add_Beam.clicked.connect(self.add_tab_beam)
            if module== KEY_DISP_CLEATANGLE:
                pushButton_Clear_Angle = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + DISP_TITLE_CLEAT)
                pushButton_Clear_Angle.clicked.connect(lambda: self.clear_tab("Angle"))
                pushButton_Add_Angle = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + DISP_TITLE_CLEAT)
                pushButton_Add_Angle.clicked.connect(self.add_tab_angle)
            if module == KEY_DISP_SEATED_ANGLE:
                pushButton_Clear_Angle = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_SEATED_ANGLE)
                pushButton_Clear_Angle.clicked.connect(lambda: self.clear_tab("Angle"))
                pushButton_Add_Angle = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_SEATED_ANGLE)
                pushButton_Add_Angle.clicked.connect(self.add_tab_angle)

        if module == KEY_DISP_COLUMNCOVERPLATE or module == KEY_DISP_COLUMNCOVERPLATEWELD or module == KEY_DISP_COLUMNENDPLATE:
            pushButton_Clear_Column = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_COLSEC)
            pushButton_Clear_Column.clicked.connect(lambda: self.clear_tab("Column"))
            pushButton_Add_Column = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_COLSEC)
            pushButton_Add_Column.clicked.connect(self.add_tab_column)


        if module == KEY_DISP_BEAMCOVERPLATE or module == KEY_DISP_BEAMCOVERPLATEWELD or module == KEY_DISP_BEAMENDPLATE:
            pushButton_Clear_Beam = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_BEAMSEC)
            pushButton_Clear_Beam.clicked.connect(lambda: self.clear_tab("Beam"))
            pushButton_Add_Beam = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_BEAMSEC)
            pushButton_Add_Beam.clicked.connect(self.add_tab_beam)

        if module == KEY_DISP_COMPRESSION:
            pushButton_Clear_Column = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_COLSEC)
            pushButton_Clear_Column.clicked.connect(lambda: self.clear_tab("Column"))
            pushButton_Add_Column = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_COLSEC)
            pushButton_Add_Column.clicked.connect(self.add_tab_column)
            pushButton_Clear_Beam = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_BEAMSEC)
            pushButton_Clear_Beam.clicked.connect(lambda: self.clear_tab("Beam"))
            pushButton_Add_Beam = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_BEAMSEC)
            pushButton_Add_Beam.clicked.connect(self.add_tab_beam)

        if module == KEY_DISP_BASE_PLATE:
            pushButton_Clear_Column = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_COLSEC)
            pushButton_Clear_Column.clicked.connect(lambda: self.clear_tab("Column"))
            pushButton_Add_Column = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_COLSEC)
            pushButton_Add_Column.clicked.connect(self.add_tab_column)

        if module == KEY_DISP_TENSION_BOLTED or module == KEY_DISP_TENSION_WELDED:
            pushButton_Clear_Angle = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + DISP_TITLE_ANGLE)
            pushButton_Clear_Angle.clicked.connect(lambda: self.clear_tab("Angle"))
            pushButton_Add_Angle = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + DISP_TITLE_ANGLE)
            pushButton_Add_Angle.clicked.connect(self.add_tab_angle)
            pushButton_Clear_Channel = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Clear_" + DISP_TITLE_CHANNEL)
            pushButton_Clear_Channel.clicked.connect(lambda: self.clear_tab("Channel"))
            pushButton_Add_Channel = self.tabWidget.findChild(QtWidgets.QWidget, "pushButton_Add_" + DISP_TITLE_CHANNEL)
            pushButton_Add_Channel.clicked.connect(self.add_tab_channel)

    def clear_tab(self, tab_name):
        '''
        @author: Umair
        '''
        if tab_name == "Column":
            tab_Column = self.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_COLSEC)
            tab = tab_Column
        elif tab_name == "Beam":
            tab_Beam = self.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_BEAMSEC)
            tab = tab_Beam
        elif tab_name == "Angle":
            tab_Angle = self.tabWidget.findChild(QtWidgets.QWidget, DISP_TITLE_ANGLE)
            tab = tab_Angle
        elif tab_name == "Channel":
            tab_Channel = self.tabWidget.findChild(QtWidgets.QWidget, DISP_TITLE_CHANNEL)
            tab = tab_Channel

        for c in tab.children():
            if isinstance(c, QtWidgets.QComboBox):
                c.setCurrentIndex(0)
            elif isinstance(c, QtWidgets.QLineEdit):
                c.clear()

    def add_tab_column(self):
        '''
        @author: Umair
        '''
        tab_Column = self.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_COLSEC)
        name = self.tabWidget.tabText(self.tabWidget.indexOf(tab_Column))
        if name in [KEY_DISP_COLSEC, KEY_DISP_SECSIZE]:
            table = "Columns"
        elif name == KEY_DISP_PRIBM:
            table = "Beams"
        else:
            pass

        for ch in tab_Column.children():
            if isinstance(ch, QtWidgets.QLineEdit) and ch.text() == "":
                QMessageBox.information(QMessageBox(), 'Warning', 'Please Fill all missing parameters!')
                add_col = tab_Column.findChild(QtWidgets.QWidget, 'pushButton_Add_'+KEY_DISP_COLSEC)
                add_col.setDisabled(True)
                break
            elif isinstance(ch, QtWidgets.QLineEdit) and ch.text() != "":
                if ch.objectName() == KEY_SECSIZE or ch.objectName() == KEY_SUPTNGSEC:
                    Designation_c = ch.text()
                elif ch.objectName() == 'Label_23':
                    Source_c = ch.text()
                elif ch.objectName() == 'Label_1':
                    D_c = float(ch.text())
                elif ch.objectName() == 'Label_2':
                    B_c = float(ch.text())
                elif ch.objectName() == 'Label_3':
                    T_c = float(ch.text())
                elif ch.objectName() == 'Label_4':
                    tw_c = float(ch.text())
                elif ch.objectName() == 'Label_5':
                    FlangeSlope_c = float(ch.text())
                elif ch.objectName() == 'Label_6':
                    R1_c = float(ch.text())
                elif ch.objectName() == 'Label_7':
                    R2_c = float(ch.text())
                elif ch.objectName() == 'Label_11':
                    Mass_c = float(ch.text())
                elif ch.objectName() == 'Label_12':
                    Area_c = float(ch.text())
                elif ch.objectName() == 'Label_13':
                    Iz_c = float(ch.text())
                elif ch.objectName() == 'Label_14':
                    Iy_c = float(ch.text())
                elif ch.objectName() == 'Label_15':
                    rz_c = float(ch.text())
                elif ch.objectName() == 'Label_16':
                    ry_c = float(ch.text())
                elif ch.objectName() == 'Label_17':
                    Zz_c = float(ch.text())
                elif ch.objectName() == 'Label_18':
                    Zy_c = float(ch.text())
                elif ch.objectName() == 'Label_19':
                    if ch.text() == "":
                        ch.setText("0")
                    Zpz_c = ch.text()
                elif ch.objectName() == 'Label_20':
                    if ch.text() == "":
                        ch.setText("0")
                    Zpy_c = ch.text()
                elif ch.objectName() == 'Label_21':
                    if ch.text() == "":
                        ch.setText("0")
                    It_c = ch.text()
                elif ch.objectName() == 'Label_22':
                    if ch.text() == "":
                        ch.setText("0")
                    Iw_c = ch.text()
                else:
                    pass
            elif isinstance(ch, QtWidgets.QComboBox):
                if ch.objectName() == 'Label_8':
                    Type = ch.currentText()

        if ch == tab_Column.children()[len(tab_Column.children())-1]:
            conn = sqlite3.connect(PATH_TO_DATABASE)
            c = conn.cursor()
            if table == "Beams":
                c.execute("SELECT count(*) FROM Beams WHERE Designation = ?", (Designation_c,))
                data = c.fetchone()[0]
            else:
                c.execute("SELECT count(*) FROM Columns WHERE Designation = ?", (Designation_c,))
                data = c.fetchone()[0]
            if data == 0:
                if table == "Beams":
                    c.execute('''INSERT INTO Beams (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Iz,Iy,rz,ry,
                        Zz,Zy,Zpz,Zpy,It,Iw,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                              (Designation_c, Mass_c, Area_c,
                               D_c, B_c, tw_c, T_c,FlangeSlope_c,
                               R1_c, R2_c, Iz_c, Iy_c, rz_c,
                               ry_c, Zz_c, Zy_c,
                               Zpz_c, Zpy_c, It_c,Iw_c,Source_c, Type))
                    conn.commit()
                else:
                    c.execute('''INSERT INTO Columns (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Iz,Iy,rz,ry,
                        Zz,Zy,Zpz,Zpy,It,Iw,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                              (Designation_c, Mass_c, Area_c,
                               D_c, B_c, tw_c, T_c,FlangeSlope_c,
                               R1_c, R2_c, Iz_c, Iy_c, rz_c,
                               ry_c, Zz_c, Zy_c,
                               Zpz_c, Zpy_c, It_c,Iw_c,Source_c, Type))
                    conn.commit()
                c.close()
                conn.close()
                QMessageBox.information(QMessageBox(), 'Information', 'Data is added successfully to the database!')

            else:
                QMessageBox.information(QMessageBox(), 'Warning', 'Designation is already exist in Database!')

    def add_tab_beam(self):
        '''
        @author: Umair
        '''
        tab_Beam = self.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_BEAMSEC)
        for ch in tab_Beam.children():
            if isinstance(ch, QtWidgets.QLineEdit) and ch.text() == "":
                QMessageBox.information(QMessageBox(), 'Warning', 'Please Fill all missing parameters!')
                add_bm = tab_Beam.findChild(QtWidgets.QWidget, 'pushButton_Add_'+KEY_DISP_BEAMSEC)
                add_bm.setDisabled(True)
                break

            elif isinstance(ch, QtWidgets.QLineEdit) and ch.text() != "":

                if ch.objectName() == KEY_SECSIZE or ch.objectName() == KEY_SUPTDSEC:
                    Designation_b = ch.text()
                elif ch.objectName() == 'Label_23':
                    Source_b = ch.text()
                elif ch.objectName() == 'Label_1':
                    D_b = float(ch.text())
                elif ch.objectName() == 'Label_2':
                    B_b = float(ch.text())
                elif ch.objectName() == 'Label_3':
                    T_b = float(ch.text())
                elif ch.objectName() == 'Label_4':
                    tw_b = float(ch.text())
                elif ch.objectName() == 'Label_5':
                    FlangeSlope_b = float(ch.text())
                elif ch.objectName() == 'Label_6':
                    R1_b = float(ch.text())
                elif ch.objectName() == 'Label_7':
                    R2_b = float(ch.text())
                elif ch.objectName() == 'Label_11':
                    Mass_b = float(ch.text())
                elif ch.objectName() == 'Label_12':
                    Area_b = float(ch.text())
                elif ch.objectName() == 'Label_13':
                    Iz_b = float(ch.text())
                elif ch.objectName() == 'Label_14':
                    Iy_b = float(ch.text())
                elif ch.objectName() == 'Label_15':
                    rz_b = float(ch.text())
                elif ch.objectName() == 'Label_16':
                    ry_b = float(ch.text())
                elif ch.objectName() == 'Label_17':
                    Zz_b = float(ch.text())
                elif ch.objectName() == 'Label_18':
                    Zy_b = float(ch.text())
                elif ch.objectName() == 'Label_19':
                    if ch.text() == "":
                        ch.setText("0")
                    Zpz_b = ch.text()
                elif ch.objectName() == 'Label_20':
                    if ch.text() == "":
                        ch.setText("0")
                    Zpy_b = ch.text()
                elif ch.objectName() == 'Label_21':
                    if ch.text() == "":
                        ch.setText("0")
                    I_t = ch.text()
                elif ch.objectName() == 'Label_22':
                    if ch.text() == "":
                        ch.setText("0")
                    I_w = ch.text()
                else:
                    pass
            elif isinstance(ch, QtWidgets.QComboBox):
                if ch.objectName() == 'Label_8':
                    Type = ch.currentText()

        if ch == tab_Beam.children()[len(tab_Beam.children())-1]:
            conn = sqlite3.connect(PATH_TO_DATABASE)

            c = conn.cursor()
            c.execute("SELECT count(*) FROM Beams WHERE Designation = ?", (Designation_b,))
            data = c.fetchone()[0]
            if data == 0:
                c.execute('''INSERT INTO Beams (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Iz,Iy,rz,ry,Zz,Zy,Zpz,Zpy,
                    It,Iw,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                          (Designation_b, Mass_b, Area_b,
                           D_b, B_b, tw_b, T_b, FlangeSlope_b,
                           R1_b, R2_b, Iz_b, Iy_b, rz_b,
                           ry_b, Zz_b, Zy_b,
                           Zpz_b, Zpy_b,I_t,I_w, Source_b, Type))
                conn.commit()
                c.close()
                conn.close()
                QMessageBox.information(QMessageBox(), 'Information', 'Data is added successfully to the database.')
            else:
                QMessageBox.information(QMessageBox(), 'Warning', 'Designation is already exist in Database!')

    def add_tab_angle(self):
        '''
        @author: Umair
        '''
        tab_Angle = self.tabWidget.findChild(QtWidgets.QWidget, DISP_TITLE_ANGLE)
        tab_name = DISP_TITLE_ANGLE
        if tab_Angle == None:
            tab_Angle = self.tabWidget.findChild(QtWidgets.QWidget, DISP_TITLE_CLEAT)
            tab_name  = DISP_TITLE_CLEAT
        if tab_Angle == None:
            tab_Angle = self.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_SEATED_ANGLE)
            tab_name  = KEY_DISP_SEATED_ANGLE
        if tab_Angle == None:
            tab_Angle = self.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_TOPANGLE)
            tab_name = KEY_DISP_TOPANGLE
        # tab_cleat_angle = self.tabWidget.findChild(QtWidgets.QWidget, DISP_TITLE_CLEAT)
        for ch in tab_Angle.children():
            if isinstance(ch, QtWidgets.QLineEdit) and ch.text() == "":
                QMessageBox.information(QMessageBox(), 'Warning', 'Please Fill all missing parameters!')
                add_bm = tab_Angle.findChild(QtWidgets.QWidget, 'pushButton_Add_'+tab_name)
                add_bm.setDisabled(True)
                break

            elif isinstance(ch, QtWidgets.QLineEdit) and ch.text() != "":

                if ch.objectName() == KEY_SECSIZE_SELECTED or ch.objectName() == KEY_ANGLE_SELECTED:
                    Designation_a = ch.text()
                elif ch.objectName() == 'Label_24':
                    Source = ch.text()
                elif ch.objectName() == 'Label_1':
                    a = ch.text()
                elif ch.objectName() == 'Label_2':
                    b = ch.text()
                elif ch.objectName() == 'Label_3':
                    t = float(ch.text())
                elif ch.objectName() == 'Label_4':
                    R1 = float(ch.text())
                elif ch.objectName() == 'Label_5':
                    R2 = float(ch.text())
                elif ch.objectName() == 'Label_7':
                    Cz = float(ch.text())
                elif ch.objectName() == 'Label_8':
                    Cy = float(ch.text())
                elif ch.objectName() == 'Label_9':
                    Mass = float(ch.text())
                elif ch.objectName() == 'Label_10':
                    Area = float(ch.text())
                elif ch.objectName() == 'Label_11':
                    I_z = float(ch.text())
                elif ch.objectName() == 'Label_12':
                    I_y = float(ch.text())
                elif ch.objectName() == 'Label_13':
                    I_u_max = float(ch.text())
                elif ch.objectName() == 'Label_14':
                    I_v_min = float(ch.text())
                elif ch.objectName() == 'Label_15':
                    rz = float(ch.text())
                elif ch.objectName() == 'Label_16':
                    ry = float(ch.text())
                elif ch.objectName() == 'Label_17':
                    if ch.text() == "":
                        ch.setText("0")
                    ru_max = float(ch.text())
                elif ch.objectName() == 'Label_18':
                    if ch.text() == "":
                        ch.setText("0")
                    rv_min = ch.text()
                elif ch.objectName() == 'Label_19':
                    if ch.text() == "":
                        ch.setText("0")
                    zz = ch.text()
                elif ch.objectName() == 'Label_20':
                    if ch.text() == "":
                        ch.setText("0")
                    zy = ch.text()
                elif ch.objectName() == 'Label_21':
                    if ch.text() == "":
                        ch.setText("0")
                    zpz = ch.text()
                elif ch.objectName() == 'Label_22':
                    if ch.text() == "":
                        ch.setText("0")
                    zpy = ch.text()
                elif ch.objectName() == 'Label_23':
                    if ch.text() == "":
                        ch.setText("0")
                    It = ch.text()

                else:
                    pass
            elif isinstance(ch, QtWidgets.QComboBox):
                if ch.objectName() == 'Label_6':
                    Type = ch.currentText()

        if ch == tab_Angle.children()[len(tab_Angle.children())-1]:
            conn = sqlite3.connect(PATH_TO_DATABASE)

            c = conn.cursor()

            c.execute("SELECT count(*) FROM Angles WHERE Designation = ?", (Designation_a,))
            data = c.fetchone()[0]
            if data == 0:
                c.execute('''INSERT INTO Angles (Designation,Mass,Area,a,b,t,R1,R2,Cz,Cy,Iz,Iy,Iumax,Ivmin,rz,ry,
                rumax,rvmin,Zz,Zy,Zpz,Zpy,It,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                          (Designation_a, Mass, Area,
                           a,b, t, R1, R2, Cz,Cy,I_z,I_y,I_u_max,
                           I_v_min, rz, ry, ru_max, rv_min,zz,zy,zpz,zpy,It,Source,Type))
                conn.commit()
                c.close()
                conn.close()
                QMessageBox.information(QMessageBox(), 'Information', 'Data is added successfully to the database.')
            else:
                QMessageBox.information(QMessageBox(), 'Warning', 'Designation is already exist in Database!')

    def add_tab_channel(self):
        '''
        @author: Umair
        '''
        tab_Channel = self.tabWidget.findChild(QtWidgets.QWidget, DISP_TITLE_CHANNEL)
        for ch in tab_Channel.children():
            if isinstance(ch, QtWidgets.QLineEdit) and ch.text() == "":
                QMessageBox.information(QMessageBox(), 'Warning', 'Please Fill all missing parameters!')
                add_bm = tab_Channel.findChild(QtWidgets.QWidget, 'pushButton_Add_'+DISP_TITLE_ANGLE)
                add_bm.setDisabled(True)
                break

            elif isinstance(ch, QtWidgets.QLineEdit) and ch.text() != "":

                if ch.objectName() == KEY_SECSIZE_SELECTED:
                    Designation_c = ch.text()
                elif ch.objectName() == 'Label_23':
                    Source = ch.text()
                elif ch.objectName() == 'Label_1':
                    B = float(ch.text())
                elif ch.objectName() == 'Label_2':
                    T = float(ch.text())
                elif ch.objectName() == 'Label_3':
                    D = float(ch.text())
                elif ch.objectName() == 'Label_13':
                    t_w = float(ch.text())
                elif ch.objectName() == 'Label_14':
                    Flange_Slope = float(ch.text())
                elif ch.objectName() == 'Label_4':
                    R1 = float(ch.text())
                elif ch.objectName() == 'Label_5':
                    R2 = float(ch.text())
                elif ch.objectName() == 'Label_9':
                    Mass = float(ch.text())
                elif ch.objectName() == 'Label_10':
                    Area = float(ch.text())
                elif ch.objectName() == 'Label_17':
                    if ch.text() == "":
                        ch.setText("0")
                    cy = float(ch.text())
                elif ch.objectName() == 'Label_11':
                    I_z = float(ch.text())
                elif ch.objectName() == 'Label_12':
                    I_y = float(ch.text())
                elif ch.objectName() == 'Label_15':
                    rz = float(ch.text())
                elif ch.objectName() == 'Label_16':
                    ry = float(ch.text())

                elif ch.objectName() == 'Label_19':
                    if ch.text() == "":
                        ch.setText("0")
                    zz = ch.text()
                elif ch.objectName() == 'Label_20':
                    if ch.text() == "":
                        ch.setText("0")
                    zy = ch.text()
                elif ch.objectName() == 'Label_21':
                    if ch.text() == "":
                        ch.setText("0")
                    zpz = ch.text()
                elif ch.objectName() == 'Label_22':
                    if ch.text() == "":
                        ch.setText("0")
                    zpy = ch.text()

                else:
                    pass
            elif isinstance(ch, QtWidgets.QComboBox):
                if ch.objectName() == 'Label_6':
                    Type = ch.currentText()

        if ch == tab_Channel.children()[len(tab_Channel.children())-1]:
            conn = sqlite3.connect(PATH_TO_DATABASE)

            c = conn.cursor()
            c.execute("SELECT count(*) FROM Channels WHERE Designation = ?", (Designation_c,))
            data = c.fetchone()[0]
            if data == 0:
                c.execute('''INSERT INTO Channels (Designation,Mass, Area,D,B,tw,T,FlangeSlope, R1, R2,Cy,Iz,Iy,
                 rz, ry,Zz,Zy,Zpz,Zpy,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                          (Designation_c, Mass, Area,D,B,t_w,T,
                           Flange_Slope, R1, R2,cy,I_z,I_y, rz, ry,zz,zy,zpz,zpy,Source,Type))
                conn.commit()
                c.close()
                conn.close()
                QMessageBox.information(QMessageBox(), 'Information', 'Data is added successfully to the database.')
            else:
                QMessageBox.information(QMessageBox(), 'Warning', 'Designation is already exist in Database!')

    def retranslateUi(self, DesignPreferences):
        _translate = QtCore.QCoreApplication.translate
        DesignPreferences.setWindowTitle(_translate("DesignPreferences", "Design preferences"))
        self.btn_defaults.setText(_translate("DesignPreferences", "Defaults"))
        self.btn_save.setText(_translate("DesignPreferences", "Save"))
        self.btn_close.setText(_translate("DesignPreferences", "Save"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Column), _translate("DesignPreferences", "Column"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Beam), _translate("DesignPreferences", "Beam"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Bolt), _translate("DesignPreferences", "Bolt"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Weld), _translate("DesignPreferences", "Weld"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Detailing), _translate("DesignPreferences", "Detailing"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Design), _translate("DesignPreferences", "Design"))


class DesignPreferences(QDialog):

    def __init__(self, module_window, main, input_dictionary, parent=None):

        QDialog.__init__(self, parent)
        self.ui = Ui_Dialog()
        self.ui.setupUi(self, main, input_dictionary)
        self.main_controller = parent
        #self.uiobj = self.main_controller.uiObj
        self.module_window = module_window
        self.saved = None
        self.flag = False
        self.sectionalprop = I_sectional_Properties()
        self.ui.btn_save.hide()
        self.ui.btn_close.clicked.connect(self.close_designPref)
        self.ui.btn_defaults.clicked.connect(self.default_fn)
        self.module = main.module_name(main)
        self.main = main
        self.window_close_flag = True


    def default_fn(self):
        '''
        @author: Umair
        '''
        tab_Bolt = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Bolt")
        tab_Weld = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Weld")
        tab_Detailing = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Detailing")
        tab_Design = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Design")

        try:
            for children in tab_Bolt.children():
                if children.objectName() == KEY_DP_BOLT_TYPE:
                    children.setCurrentIndex(0)
                elif children.objectName() == KEY_DP_BOLT_HOLE_TYPE:
                    children.setCurrentIndex(0)
                elif children.objectName() == KEY_DP_BOLT_MATERIAL_G_O:
                    children.setText('410')
                elif children.objectName() == KEY_DP_BOLT_SLIP_FACTOR:
                    children.setCurrentIndex(4)
                else:
                    pass
        except:
            pass

        try:
            for children in tab_Weld.children():
                if children.objectName() == KEY_DP_WELD_FAB:
                    children.setCurrentIndex(0)
                elif children.objectName() == KEY_DP_WELD_MATERIAL_G_O:
                    children.setText('410')
                else:
                    pass
        except:
            pass

        try:
            for children in tab_Detailing.children():
                if children.objectName() == KEY_DP_DETAILING_EDGE_TYPE:
                    children.setCurrentIndex(0)
                elif children.objectName() == KEY_DP_DETAILING_GAP:
                    children.setText('10')
                elif children.objectName() == KEY_DP_DETAILING_CORROSIVE_INFLUENCES:
                    children.setCurrentIndex(0)
                else:
                    pass
        except:
            pass
        try:
            for children in tab_Design.children():
                if children.objectName() == KEY_DP_DESIGN_METHOD:
                    children.setCurrentIndex(0)
                else:
                    pass
        except:
            pass

    #
    # def save_designPref_para(self, module):
    #     """This routine is responsible for saving all design preferences selected by the user
    #     """
    #     '''
    #     @author: Umair
    #     '''
    #     tab_Bolt = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Bolt")
    #     tab_Weld = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Weld")
    #     tab_Detailing = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Detailing")
    #     tab_Design = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Design")
    #     tab_Connector = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Connector")
    #
    #     #
    #     # key_boltHoleType = tab_Bolt.findChild(QtWidgets.QWidget, KEY_DP_BOLT_HOLE_TYPE)
    #     # combo_boltHoleType = key_boltHoleType.currentText()
    #     # key_boltFu = tab_Bolt.findChild(QtWidgets.QWidget, KEY_DP_BOLT_MATERIAL_G_O)
    #     # line_boltFu = key_boltFu.text()
    #     # key_slipfactor = tab_Bolt.findChild(QtWidgets.QWidget, KEY_DP_BOLT_SLIP_FACTOR)
    #     # combo_slipfactor = key_slipfactor.currentText()
    #
    #     tab_Anchor_Bolt = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Anchor Bolt")
    #
    #     if module != KEY_DISP_BASE_PLATE:
    #         key_boltTensioning = tab_Bolt.findChild(QtWidgets.QWidget, KEY_DP_BOLT_TYPE)
    #         combo_boltTensioning = key_boltTensioning.currentText()
    #         key_boltHoleType = tab_Bolt.findChild(QtWidgets.QWidget, KEY_DP_BOLT_HOLE_TYPE)
    #         combo_boltHoleType = key_boltHoleType.currentText()
    #         key_boltFu = tab_Bolt.findChild(QtWidgets.QWidget, KEY_DP_BOLT_MATERIAL_G_O)
    #         line_boltFu = key_boltFu.text()
    #         key_slipfactor = tab_Bolt.findChild(QtWidgets.QWidget, KEY_DP_BOLT_SLIP_FACTOR)
    #         combo_slipfactor = key_slipfactor.currentText()
    #         key_detailingGap = tab_Detailing.findChild(QtWidgets.QWidget, KEY_DP_DETAILING_GAP)
    #         line_detailingGap = key_detailingGap.text()
    #     elif module == KEY_DISP_BASE_PLATE:
    #         key_boltDesignation = tab_Anchor_Bolt.findChild(QtWidgets.QWidget, KEY_DP_ANCHOR_BOLT_DESIGNATION)
    #         line_boltDesignation = key_boltDesignation.text()
    #         key_boltHoleType = tab_Anchor_Bolt.findChild(QtWidgets.QWidget, KEY_DP_ANCHOR_BOLT_HOLE_TYPE)
    #         combo_boltHoleType = key_boltHoleType.currentText()
    #         key_boltType = tab_Anchor_Bolt.findChild(QtWidgets.QWidget, KEY_DP_ANCHOR_BOLT_TYPE)
    #         combo_boltType = key_boltType.text()
    #         key_boltFu = tab_Anchor_Bolt.findChild(QtWidgets.QWidget, KEY_DP_ANCHOR_BOLT_MATERIAL_G_O)
    #         line_boltFu = key_boltFu.text()
    #         key_boltFriction = tab_Anchor_Bolt.findChild(QtWidgets.QWidget, KEY_DP_ANCHOR_BOLT_FRICTION)
    #         line_boltFriction = key_boltFriction.text()
    #
    #     key_weldType = tab_Weld.findChild(QtWidgets.QWidget, KEY_DP_WELD_FAB)
    #     combo_weldType = key_weldType.currentText()
    #     key_weldFu = tab_Weld.findChild(QtWidgets.QWidget, KEY_DP_WELD_MATERIAL_G_O)
    #     line_weldFu = key_weldFu.text()
    #     key_detailingEdgeType = tab_Detailing.findChild(QtWidgets.QWidget, KEY_DP_DETAILING_EDGE_TYPE)
    #     combo_detailingEdgeType = key_detailingEdgeType.currentText()
    #
    #     key_detailing_memebers = tab_Detailing.findChild(QtWidgets.QWidget, KEY_DP_DETAILING_CORROSIVE_INFLUENCES)
    #     combo_detailing_memebers = key_detailing_memebers.currentText()
    #     key_design_method = tab_Design.findChild(QtWidgets.QWidget, KEY_DP_DESIGN_METHOD)
    #     combo_design_method = key_design_method.currentText()
    #
    #     key_design_baseplate = tab_Design.findChild(QtWidgets.QWidget, KEY_DP_DESIGN_BASE_PLATE)
    #     if module not in [KEY_DISP_BASE_PLATE, KEY_DISP_TENSION_BOLTED, KEY_DISP_TENSION_WELDED,KEY_DISP_COMPRESSION]:
    #         key_plate_material = tab_Connector.findChild(QtWidgets.QWidget, KEY_PLATE_MATERIAL)
    #         combo_plate_material = key_plate_material.currentText()
    #         key_plate_material_fu = tab_Connector.findChild(QtWidgets.QWidget, KEY_PLATE_FU)
    #         line_plate_material_fu = key_plate_material_fu.text()
    #         key_plate_material_fy = tab_Connector.findChild(QtWidgets.QWidget, KEY_PLATE_FY)
    #         line_plate_material_fy = key_plate_material_fy.text()
    #         d1 = {KEY_DP_BOLT_TYPE: combo_boltTensioning,
    #               KEY_DP_BOLT_HOLE_TYPE: combo_boltHoleType,
    #               KEY_DP_BOLT_MATERIAL_G_O: line_boltFu,
    #               KEY_DP_BOLT_SLIP_FACTOR: combo_slipfactor,
    #               KEY_DP_WELD_FAB: combo_weldType,
    #               KEY_DP_WELD_MATERIAL_G_O: line_weldFu,
    #               KEY_DP_DETAILING_EDGE_TYPE: combo_detailingEdgeType,
    #               KEY_DP_DETAILING_GAP: line_detailingGap,
    #               KEY_DP_DETAILING_CORROSIVE_INFLUENCES: combo_detailing_memebers,
    #               KEY_DP_DESIGN_METHOD: combo_design_method,
    #               KEY_PLATE_MATERIAL: combo_plate_material if combo_plate_material != "Custom" else
    #               "Custom " + str(line_plate_material_fu) + " " + str(line_plate_material_fy),
    #               }
    #     elif module == KEY_DISP_BASE_PLATE:
    #         if self.flag:
    #             d1 = {KEY_DP_ANCHOR_BOLT_DESIGNATION: line_boltDesignation,
    #                   KEY_DP_ANCHOR_BOLT_TYPE: combo_boltType,
    #                   KEY_DP_ANCHOR_BOLT_HOLE_TYPE: combo_boltHoleType,
    #                   KEY_DP_ANCHOR_BOLT_MATERIAL_G_O: line_boltFu,
    #                   KEY_DP_ANCHOR_BOLT_FRICTION: line_boltFriction,
    #                   KEY_DP_WELD_FAB: combo_weldType,
    #                   KEY_DP_WELD_MATERIAL_G_O: line_weldFu,
    #                   KEY_DP_DETAILING_EDGE_TYPE: combo_detailingEdgeType,
    #                   KEY_DP_DETAILING_CORROSIVE_INFLUENCES: combo_detailing_memebers,
    #                   KEY_DP_DESIGN_METHOD: combo_design_method,
    #                   KEY_DP_DESIGN_BASE_PLATE: key_design_baseplate.currentText()
    #                   }
    #         else:
    #             d1 = {KEY_DP_WELD_FAB: combo_weldType,
    #                   KEY_DP_WELD_MATERIAL_G_O: line_weldFu,
    #                   KEY_DP_DETAILING_EDGE_TYPE: combo_detailingEdgeType,
    #                   KEY_DP_DETAILING_CORROSIVE_INFLUENCES: combo_detailing_memebers,
    #                   KEY_DP_DESIGN_METHOD: combo_design_method,
    #                   KEY_DP_DESIGN_BASE_PLATE: key_design_baseplate.currentText()
    #                   }
    #     # else:
    #     #     d1 = {KEY_DP_BOLT_HOLE_TYPE: combo_boltHoleType,
    #     #           KEY_DP_BOLT_MATERIAL_G_O: line_boltFu,
    #     #           KEY_DP_BOLT_SLIP_FACTOR: combo_slipfactor,
    #     #           KEY_DP_WELD_FAB: combo_weldType,
    #     #           KEY_DP_WELD_MATERIAL_G_O: line_weldFu,
    #     #           KEY_DP_DETAILING_EDGE_TYPE: combo_detailingEdgeType,
    #     #           KEY_DP_DETAILING_GAP: line_detailingGap,
    #     #           KEY_DP_DETAILING_CORROSIVE_INFLUENCES: combo_detailing_memebers,
    #     #           KEY_DP_DESIGN_METHOD: combo_design_method
    #     #           }
    #     return d1

    def highlight_slipfactor_description(self):
        """Highlight the description of currosponding slipfactor on selection of inputs
        Note : This routine is not in use in current version
        :return:
        """
        slip_factor = str(self.ui.combo_slipfactor.currentText())
        self.textCursor = QTextCursor(self.ui.textBrowser.document())
        cursor = self.textCursor
        # Setup the desired format for matches
        format = QTextCharFormat()
        format.setBackground(QBrush(QColor("red")))
        # Setup the regex engine
        pattern = str(slip_factor)
        regex = QRegExp(pattern)
        # Process the displayed document
        pos = 0
        index = regex.indexIn(self.ui.textBrowser.toPlainText(), pos)
        while (index != -1):
            # Select the matched text and apply the desired format
            cursor.setPosition(index)
            cursor.movePosition(QTextCursor.EndOfLine, 1)
            # cursor.movePosition(QTextCursor.EndOfWord, 1)
            cursor.mergeCharFormat(format)
            # Move to the next match
            pos = index + regex.matchedLength()
            index = regex.indexIn(self.ui.textBrowser.toPlainText(), pos)

    # def connect_to_database_update_other_attributes(self, table, designation):
    #     self.path_to_database = "ResourceFiles/Database/Intg_osdag.sqlite"
    #     conn = sqlite3.connect(self.path_to_database)
    #     db_query = "SELECT * FROM " + table + " WHERE Designation = ?"
    #     cur = conn.cursor()
    #     cur.execute(db_query, (designation,))
    #     row = cur.fetchone()
    #     self.mass = row[2]
    #     self.area = row[3]
    #     self.depth = row[4]
    #     self.flange_width = row[5]
    #     self.web_thickness = row[6]
    #     self.flange_thickness = row[7]
    #     self.flange_slope = row[8]
    #     self.root_radius = row[9]
    #     self.toe_radius = row[10]
    #     self.mom_inertia_z = row[11]
    #     self.mom_inertia_y = row[12]
    #     self.rad_of_gy_z = row[13]
    #     self.rad_of_gy_y = row[14]
    #     self.elast_sec_mod_z = row[15]
    #     self.elast_sec_mod_y = row[16]
    #     self.plast_sec_mod_z = row[17]
    #     self.plast_sec_mod_y = row[18]
    #     self.source = row[19]
    #
    #     conn.close()
    # def column_preferences(self, designation, table, material_grade):
    #     '''
    #     @author: Umair
    #     '''
    #     # designation = designation_table_material_grade[0]
    #     # table = designation_table_material_grade[1]
    #     # material_grade = designation_table_material_grade[2]
    #     tab_Column = self.ui.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_COLSEC)
    #     if designation == 'Select Section':
    #         self.flag = False
    #         self.ui.clear_tab("Column")
    #         return
    #
    #     col_list = []
    #     fu_fy_list = []
    #     col_attributes = Section(designation, material_grade)
    #     Section.connect_to_database_update_other_attributes(col_attributes, table, designation)
    #     for ch in tab_Column.children():
    #         if ch.objectName() == KEY_SUPTNGSEC_DESIGNATION:
    #             ch.setText(designation)
    #         elif ch.objectName() == KEY_SUPTNGSEC_SOURCE:
    #             ch.setText(col_attributes.source)
    #         elif ch.objectName() == KEY_SUPTNGSEC_MATERIAL:
    #             if self.module == KEY_DISP_BASE_PLATE:
    #                 ch.setText(material_grade)
    #             else:
    #                 indx = ch.findText(material_grade, QtCore.Qt.MatchFixedString)
    #                 if indx >= 0:
    #                     ch.setCurrentIndex(indx)
    #         elif ch.objectName() == KEY_SUPTNGSEC_FU:
    #             if self.module == KEY_DISP_BASE_PLATE:
    #                 ch.setText(str(col_attributes.fu))
    #                 fu_fy_list.append(ch)
    #             else:
    #                 ch.setText(str(col_attributes.fu))
    #                 ch.setEnabled(True if material_grade == 'Custom' else False)
    #         elif ch.objectName() == KEY_SUPTNGSEC_FY:
    #             if self.module == KEY_DISP_BASE_PLATE:
    #                 ch.setText(str(col_attributes.fy))
    #                 fu_fy_list.append(ch)
    #             else:
    #                 ch.setText(str(col_attributes.fy))
    #                 ch.setEnabled(True if material_grade == 'Custom' else False)
    #         elif ch.objectName() == KEY_SUPTNGSEC_DEPTH:
    #             ch.setText(str(col_attributes.depth))
    #             col_list.append(ch)
    #         elif ch.objectName() == KEY_SUPTNGSEC_FLANGE_W:
    #             ch.setText(str(col_attributes.flange_width))
    #             col_list.append(ch)
    #         elif ch.objectName() == KEY_SUPTNGSEC_FLANGE_T:
    #             ch.setText(str(col_attributes.flange_thickness))
    #             col_list.append(ch)
    #         elif ch.objectName() == KEY_SUPTNGSEC_WEB_T:
    #             ch.setText(str(col_attributes.web_thickness))
    #             col_list.append(ch)
    #         elif ch.objectName() == KEY_SUPTNGSEC_FLANGE_S:
    #             ch.setText(str(col_attributes.flange_slope))
    #         elif ch.objectName() == KEY_SUPTNGSEC_ROOT_R:
    #             ch.setText(str(col_attributes.root_radius))
    #         elif ch.objectName() == KEY_SUPTNGSEC_TOE_R:
    #             ch.setText(str(col_attributes.toe_radius))
    #         elif ch.objectName() == KEY_SUPTNGSEC_MOD_OF_ELAST:
    #             ch.setText("200")
    #             ch.setDisabled(True)
    #         elif ch.objectName() == KEY_SUPTNGSEC_MOD_OF_RIGID:
    #             ch.setText("76.9")
    #             ch.setDisabled(True)
    #         elif ch.objectName() == KEY_SUPTNGSEC_POISSON_RATIO:
    #             ch.setText("0.3")
    #             ch.setDisabled(True)
    #         elif ch.objectName() == KEY_SUPTNGSEC_THERMAL_EXP:
    #             ch.setText("12")
    #             ch.setDisabled(True)
    #         elif ch.objectName() == KEY_SUPTNGSEC_MASS:
    #             ch.setText(str(col_attributes.mass))
    #         elif ch.objectName() == KEY_SUPTNGSEC_SEC_AREA:
    #             ch.setText(str(col_attributes.area))
    #         elif ch.objectName() == KEY_SUPTNGSEC_MOA_LZ:
    #             ch.setText(str(col_attributes.mom_inertia_z))
    #         elif ch.objectName() == KEY_SUPTNGSEC_MOA_LY:
    #             ch.setText(str(col_attributes.mom_inertia_y))
    #         elif ch.objectName() == KEY_SUPTNGSEC_ROG_RZ:
    #             ch.setText(str(col_attributes.rad_of_gy_z))
    #         elif ch.objectName() == KEY_SUPTNGSEC_ROG_RY:
    #             ch.setText(str(col_attributes.rad_of_gy_y))
    #         elif ch.objectName() == KEY_SUPTNGSEC_EM_ZZ:
    #             ch.setText(str(col_attributes.elast_sec_mod_z))
    #         elif ch.objectName() == KEY_SUPTNGSEC_EM_ZY:
    #             ch.setText(str(col_attributes.elast_sec_mod_y))
    #         elif ch.objectName() == KEY_SUPTNGSEC_PM_ZPZ:
    #             ch.setText(str(col_attributes.plast_sec_mod_z))
    #         elif ch.objectName() == KEY_SUPTNGSEC_PM_ZPY:
    #             ch.setText(str(col_attributes.plast_sec_mod_y))
    #         elif ch.objectName() == 'pushButton_Add_Column':
    #             ch.setEnabled(True)
    #         else:
    #             pass
    #
    #     for e in col_list:
    #         if e.text() != "":
    #             e.textChanged.connect(lambda: self.new_sectionalprop_Column(col_list))
    #
    #     for f in fu_fy_list:
    #         if f.text() != "":
    #             self.fu_fy_validation_connect(fu_fy_list, f)
    #             # f.textChanged.connect(lambda: self.fu_fy_validation(fu_fy_list, f))
    #
    #     # def f():
    #     #     found = False
    #     #     material_key = tab_Column.findChild(QtWidgets.QWidget, KEY_SUPTNGSEC_MATERIAL)
    #     #     for i in range(material_key.count()):
    #     #         if material_key.itemText(i) == "Custom":
    #     #             found = True
    #     #         if i == material_key.count() - 1:
    #     #             if found:
    #     #                 material_key.setCurrentText("Custom")
    #     #                 return
    #     #             else:
    #     #                 material_key.addItem("Custom")
    #     #                 material_key.setCurrentText("Custom")
    #     #                 return
    #     # for m in material_list:
    #     #     if m.text() != "":
    #     #         m.textChanged.connect(f)
    #
    # def beam_preferences(self, designation, material_grade):
    #     '''
    #     @author: Umair
    #     '''
    #     # designation = designation_material_grade[0]
    #     # material_grade = designation_material_grade[1]
    #     tab_Beam = self.ui.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_BEAMSEC)
    #     if designation == 'Select Section':
    #         self.flag = False
    #         self.ui.clear_tab("Beam")
    #         return
    #
    #     beam_attributes = Section(designation, material_grade)
    #     Section.connect_to_database_update_other_attributes(beam_attributes, "Beams", designation)
    #     beam_list = []
    #     for ch in tab_Beam.children():
    #         if ch.objectName() == KEY_SUPTDSEC_DESIGNATION:
    #             ch.setText(designation)
    #         elif ch.objectName() == KEY_SUPTDSEC_SOURCE:
    #             ch.setText(beam_attributes.source)
    #         elif ch.objectName() == KEY_SUPTDSEC_MATERIAL:
    #             indx = ch.findText(material_grade, QtCore.Qt.MatchFixedString)
    #             if indx >= 0:
    #                 ch.setCurrentIndex(indx)
    #         elif ch.objectName() == KEY_SUPTDSEC_FU:
    #             ch.setText(str(beam_attributes.fu))
    #             ch.setEnabled(True if material_grade == 'Custom' else False)
    #         elif ch.objectName() == KEY_SUPTDSEC_FY:
    #             ch.setText(str(beam_attributes.fy))
    #             ch.setEnabled(True if material_grade == 'Custom' else False)
    #         elif ch.objectName() == KEY_SUPTDSEC_DEPTH:
    #             ch.setText(str(beam_attributes.depth))
    #             beam_list.append(ch)
    #         elif ch.objectName() == KEY_SUPTDSEC_FLANGE_W:
    #             ch.setText(str(beam_attributes.flange_width))
    #             beam_list.append(ch)
    #         elif ch.objectName() == KEY_SUPTDSEC_FLANGE_T:
    #             ch.setText(str(beam_attributes.flange_thickness))
    #             beam_list.append(ch)
    #         elif ch.objectName() == KEY_SUPTDSEC_WEB_T:
    #             ch.setText(str(beam_attributes.web_thickness))
    #             beam_list.append(ch)
    #         elif ch.objectName() == KEY_SUPTDSEC_FLANGE_S:
    #             ch.setText(str(beam_attributes.flange_slope))
    #         elif ch.objectName() == KEY_SUPTDSEC_ROOT_R:
    #             ch.setText(str(beam_attributes.root_radius))
    #         elif ch.objectName() == KEY_SUPTDSEC_TOE_R:
    #             ch.setText(str(beam_attributes.toe_radius))
    #         elif ch.objectName() == KEY_SUPTDSEC_MOD_OF_ELAST:
    #             ch.setText("200")
    #             ch.setDisabled(True)
    #         elif ch.objectName() == KEY_SUPTDSEC_MOD_OF_RIGID:
    #             ch.setText("76.9")
    #             ch.setDisabled(True)
    #         elif ch.objectName() == KEY_SUPTDSEC_POISSON_RATIO:
    #             ch.setText("0.3")
    #             ch.setDisabled(True)
    #         elif ch.objectName() == KEY_SUPTDSEC_THERMAL_EXP:
    #             ch.setText("12")
    #             ch.setDisabled(True)
    #         elif ch.objectName() == KEY_SUPTDSEC_MASS:
    #             ch.setText(str(beam_attributes.mass))
    #         elif ch.objectName() == KEY_SUPTDSEC_SEC_AREA:
    #             ch.setText(str(beam_attributes.area))
    #         elif ch.objectName() == KEY_SUPTDSEC_MOA_LZ:
    #             ch.setText(str(beam_attributes.mom_inertia_z))
    #         elif ch.objectName() == KEY_SUPTDSEC_MOA_LY:
    #             ch.setText(str(beam_attributes.mom_inertia_y))
    #         elif ch.objectName() == KEY_SUPTDSEC_ROG_RZ:
    #             ch.setText(str(beam_attributes.rad_of_gy_z))
    #         elif ch.objectName() == KEY_SUPTDSEC_ROG_RY:
    #             ch.setText(str(beam_attributes.rad_of_gy_y))
    #         elif ch.objectName() == KEY_SUPTDSEC_EM_ZZ:
    #             ch.setText(str(beam_attributes.elast_sec_mod_z))
    #         elif ch.objectName() == KEY_SUPTDSEC_EM_ZY:
    #             ch.setText(str(beam_attributes.elast_sec_mod_y))
    #         elif ch.objectName() == KEY_SUPTDSEC_PM_ZPZ:
    #             ch.setText(str(beam_attributes.plast_sec_mod_z))
    #         elif ch.objectName() == KEY_SUPTDSEC_PM_ZPY:
    #             ch.setText(str(beam_attributes.plast_sec_mod_y))
    #         elif ch.objectName() == 'pushButton_Add_Beam':
    #             ch.setEnabled(True)
    #         else:
    #             pass
    #
    #     for e in beam_list:
    #         if e.text() != "":
    #             e.textChanged.connect(lambda: self.new_sectionalprop_Beam(beam_list))
    #
    # def angle_preferences(self,designation,material_grade):
    #     tab_Angle = self.ui.tabWidget.findChild(QtWidgets.QWidget, DISP_TITLE_ANGLE)
    #
    #     # if designation == 'Select Section':
    #     #     self.ui.clear_tab("Angle")
    #     #     return
    #     ch=tab_Angle.findChild(QtWidgets.QWidget, KEY_ANGLE_DESIGNATION)
    #     ch.setText(designation)

    def fu_fy_validation_connect(self, fu_fy_list, f, m):
        f.textChanged.connect(lambda: self.fu_fy_validation(fu_fy_list, f, m))

    def fu_fy_validation(self, fu_fy_list, textbox, material_key):
        self.window_close_flag = False
        # self.rejected.disconnect()
        # self.rejected.connect(self.closeEvent_accept)
        print(fu_fy_list[0].text(), fu_fy_list[1].text())
        if "" not in [fu_fy_list[0].text(), fu_fy_list[1].text()]:
            fu = float(fu_fy_list[0].text())
            fy = float(fu_fy_list[1].text())
            material = material_key.currentText()

        else:
            textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: black;")
            return

        if fu and fy:
            if 'Ultimate_Strength' in textbox.objectName():

                if fu < 290 or fu > 639:
                    textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: red;")
                    self.window_close_flag = False
                    self.rejected.connect(self.closeEvent)
                    return
                else:

                    fu_limits = self.get_limits_for_fu(str(material))

                    if fu_limits['lower'] <= fu <= fu_limits['upper']:
                        textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: black;")
                        self.window_close_flag = True
                        self.rejected.connect(self.closeEvent)
                        return
                    else:
                        textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: red;")
                        self.window_close_flag = False
                        self.rejected.connect(self.closeEvent)
                        return

            if 'Yield_Strength' in textbox.objectName():
                if fy < 165 or fy > 499:
                    textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: red;")
                    self.window_close_flag = False
                    self.rejected.connect(self.closeEvent)
                    return

                else:

                    fy_limits = self.get_limits_for_fy(str(material))
                    if fy_limits['lower'] <= fy <= fy_limits['upper']:
                        textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: black;")
                        self.window_close_flag = True
                        self.rejected.connect(self.closeEvent)
                        return
                    else:
                        textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: red;")
                        self.window_close_flag = False
                        self.rejected.connect(self.closeEvent)
                        return

    def get_limits_for_fu(self, material_grade):

        lower_fu = {'E 165 (Fe 290)': 290,
                    'E 250 (Fe 410 W)A': 410,
                    'E 250 (Fe 410 W)B': 410,
                    'E 250 (Fe 410 W)C': 410,
                    'E 300 (Fe 440)': 440,
                    'E 350 (Fe 490)': 490,
                    'E 410 (Fe 540)': 540,
                    'E 450 (Fe 570)D': 570,
                    'E 450 (Fe 590) E': 590}[material_grade]

        upper_fu = {'E 165 (Fe 290)': 409,
                    'E 250 (Fe 410 W)A': 439,
                    'E 250 (Fe 410 W)B': 439,
                    'E 250 (Fe 410 W)C': 439,
                    'E 300 (Fe 440)': 489,
                    'E 350 (Fe 490)': 539,
                    'E 410 (Fe 540)': 569,
                    'E 450 (Fe 570)D': 589,
                    'E 450 (Fe 590) E': 639}[material_grade]

        return {'lower': lower_fu, 'upper': upper_fu}

    def get_limits_for_fy(self, material_grade):

        lower_fy = {'E 165 (Fe 290)': 165,
                    'E 250 (Fe 410 W)A': 230,
                    'E 250 (Fe 410 W)B': 230,
                    'E 250 (Fe 410 W)C': 230,
                    'E 300 (Fe 440)': 280,
                    'E 350 (Fe 490)': 320,
                    'E 410 (Fe 540)': 380,
                    'E 450 (Fe 570)D': 420,
                    'E 450 (Fe 590) E': 420}[material_grade]

        upper_fy = {'E 165 (Fe 290)': 249,
                    'E 250 (Fe 410 W)A': 299,
                    'E 250 (Fe 410 W)B': 299,
                    'E 250 (Fe 410 W)C': 299,
                    'E 300 (Fe 440)': 349,
                    'E 350 (Fe 490)': 409,
                    'E 410 (Fe 540)': 449,
                    'E 450 (Fe 570)D': 499,
                    'E 450 (Fe 590) E': 499}[material_grade]

        return {'lower': lower_fy, 'upper': upper_fy}

    # def anchor_bolt_designation(self, d):
    #     length = str(self.main.anchor_length_provided if self.main.design_button_status else 0)
    #     designation = str(d) + "X" + length + " IS5624 GALV"
    #     return designation, length

    # def anchor_bolt_preferences(self, d, typ):
    #
    #     change_list = []
    #     tab_anchor_bolt = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Anchor Bolt")
    #     # length = IS_5624_1993.table1(d)
    #     # length = str(length[1])
    #     # designation = str(d)+"X"+length+" IS5624 GALV"
    #     designation = self.anchor_bolt_designation(d)[0]
    #     # initial_designation = designation
    #     for ch in tab_anchor_bolt.children():
    #         if ch.objectName() == KEY_DP_ANCHOR_BOLT_DESIGNATION:
    #             ch.setText(designation)
    #             ch.setReadOnly(True)
    #         elif ch.objectName() == KEY_DP_ANCHOR_BOLT_LENGTH:
    #             ch.setText(str(self.main.anchor_length_provided) if self.main.design_button_status else '0')
    #             change_list.append(ch)
    #             # ch.setReadOnly(True)
    #         elif ch.objectName() == KEY_DP_ANCHOR_BOLT_TYPE:
    #             ch.setText(typ)
    #             ch.setReadOnly(True)
    #         elif ch.objectName() == KEY_DP_ANCHOR_BOLT_GALVANIZED:
    #             change_list.append(ch)
    #         elif ch.objectName() == KEY_DP_ANCHOR_BOLT_MATERIAL_G_O:
    #             ch.setText(str(self.main.anchor_fu_fy[0]) if self.main.design_button_status else '0')
    #
    #     for c in change_list:
    #         if isinstance(c, QtWidgets.QComboBox):
    #             c.currentIndexChanged.connect(lambda: self.anchor_bolt_designation_change(change_list[0], change_list))
    #         elif isinstance(c, QtWidgets.QLineEdit):
    #             c.textChanged.connect(lambda: self.anchor_bolt_designation_change(c, change_list))

    # def anchor_bolt_designation_change(self, e, e_list):
    #     des = self.ui.tabWidget.findChild(QtWidgets.QWidget, "Anchor Bolt").findChild(QtWidgets.QWidget,
    #                                                                                   KEY_DP_ANCHOR_BOLT_DESIGNATION)
    #     initial_des = des.text()
    #     if isinstance(e, QtWidgets.QComboBox):
    #         des_list = initial_des.split(' ')
    #         new_des = des_list[0]+" "+des_list[1]
    #         if e.currentText() == 'Yes':
    #             des.setText(str(initial_des + " GALV"))
    #         elif e.currentText() == 'No':
    #             des.setText(new_des)
    #     elif isinstance(e, QtWidgets.QLineEdit):
    #         des_list = initial_des.split('X')
    #         des_list_2 = des_list[1].split(' ')
    #         if e.text() == "":
    #             if e_list[0].currentText() == 'Yes':
    #                 new_des = str(des_list[0])+'X '+str(des_list_2[1])+' '+str(des_list_2[2])
    #             else:
    #                 new_des = str(des_list[0]) + 'X ' + str(des_list_2[1])
    #         elif e_list[0].currentText() == 'Yes':
    #             new_des = str(des_list[0])+'X'+str(e.text())+' '+str(des_list_2[1])+' '+str(des_list_2[2])
    #         else:
    #             new_des = str(des_list[0]) + 'X' + str(e.text()) + ' ' + str(des_list_2[1])
    #         des.setText(new_des)

    def closeEvent(self, event):
        if self.window_close_flag:
            event.accept()
            self.module_window.prev_inputs = self.module_window.input_dock_inputs
        else:
            QMessageBox.warning(self, "Error", "Select correct values for fu and fy!")
            event.ignore()

    # def new_sectionalprop_Column(self, col_list):
    #     '''
    #     @author: Umair
    #     '''
    #
    #     for e in col_list:
    #         if e.text() != "":
    #             if e.objectName() == KEY_SUPTNGSEC_DEPTH:
    #                 D = float(e.text())
    #             elif e.objectName() == KEY_SUPTNGSEC_FLANGE_W:
    #                 B = float(e.text())
    #             elif e.objectName() == KEY_SUPTNGSEC_FLANGE_T:
    #                 t_w = float(e.text())
    #             elif e.objectName() == KEY_SUPTNGSEC_WEB_T:
    #                 t_f = float(e.text())
    #             else:
    #                 pass
    #         else:
    #             return
    #     if col_list:
    #         tab_Column = self.ui.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_COLSEC)
    #         for c in tab_Column.children():
    #             if c.objectName() == KEY_SUPTNGSEC_MASS:
    #                 c.setText(str(self.sectionalprop.calc_Mass(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTNGSEC_SEC_AREA:
    #                 c.setText(str(self.sectionalprop.calc_Area(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTNGSEC_MOA_LZ:
    #                 c.setText(str(self.sectionalprop.calc_MomentOfAreaZ(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTNGSEC_MOA_LY:
    #                 c.setText(str(self.sectionalprop.calc_MomentOfAreaY(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTNGSEC_ROG_RZ:
    #                 c.setText(str(self.sectionalprop.calc_RogZ(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTNGSEC_ROG_RY:
    #                 c.setText(str(self.sectionalprop.calc_RogY(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTNGSEC_EM_ZZ:
    #                 c.setText(str(self.sectionalprop.calc_ElasticModulusZz(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTNGSEC_EM_ZY:
    #                 c.setText(str(self.sectionalprop.calc_ElasticModulusZy(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTNGSEC_PM_ZPZ:
    #                 c.setText(str(self.sectionalprop.calc_PlasticModulusZpz(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTNGSEC_PM_ZPY:
    #                 c.setText(str(self.sectionalprop.calc_PlasticModulusZpy(D, B, t_w, t_f)))
    #             elif c.objectName() == 'pushButton_Add_Column':
    #                 c.setEnabled(True)
    #             else:
    #                 pass
    #
    # def new_sectionalprop_Beam(self, beam_list):
    #     '''
    #     @author: Umair
    #     '''
    #
    #     for e in beam_list:
    #         if e.text() != "":
    #             if e.objectName() == KEY_SUPTDSEC_DEPTH:
    #                 D = float(e.text())
    #             elif e.objectName() == KEY_SUPTDSEC_FLANGE_W:
    #                 B = float(e.text())
    #             elif e.objectName() == KEY_SUPTDSEC_FLANGE_T:
    #                 t_w = float(e.text())
    #             elif e.objectName() == KEY_SUPTDSEC_WEB_T:
    #                 t_f = float(e.text())
    #             else:
    #                 pass
    #         else:
    #             return
    #     if beam_list:
    #         tab_Beam = self.ui.tabWidget.findChild(QtWidgets.QWidget, KEY_DISP_BEAMSEC)
    #         for c in tab_Beam.children():
    #             if c.objectName() == KEY_SUPTDSEC_MASS:
    #                 c.setText(str(self.sectionalprop.calc_Mass(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTDSEC_SEC_AREA:
    #                 c.setText(str(self.sectionalprop.calc_Area(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTDSEC_MOA_LZ:
    #                 c.setText(str(self.sectionalprop.calc_MomentOfAreaZ(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTDSEC_MOA_LY:
    #                 c.setText(str(self.sectionalprop.calc_MomentOfAreaY(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTDSEC_ROG_RZ:
    #                 c.setText(str(self.sectionalprop.calc_RogZ(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTDSEC_ROG_RY:
    #                 c.setText(str(self.sectionalprop.calc_RogY(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTDSEC_EM_ZZ:
    #                 c.setText(str(self.sectionalprop.calc_ElasticModulusZz(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTDSEC_EM_ZY:
    #                 c.setText(str(self.sectionalprop.calc_ElasticModulusZy(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTDSEC_PM_ZPZ:
    #                 c.setText(str(self.sectionalprop.calc_PlasticModulusZpz(D, B, t_w, t_f)))
    #             elif c.objectName() == KEY_SUPTDSEC_PM_ZPY:
    #                 c.setText(str(self.sectionalprop.calc_PlasticModulusZpy(D, B, t_w, t_f)))
    #             elif c.objectName() == 'pushButton_Add_Beam':
    #                 c.setEnabled(True)
    #             else:
    #                 pass

    def download_Database_Column(self):
        file_path = os.path.abspath(os.path.join(os.getcwd(), os.path.join("ResourceFiles", "add_sections.xlsx")))
        shutil.copyfile(file_path, os.path.join(str(self.folder), "images_html", "add_sections.xlsx"))
        QMessageBox.information(QMessageBox(), 'Information', 'Your File is Downloaded in your selected workspace')
        #self.ui.pushButton_Import_Column.setEnabled(True)

    def download_Database_Beam(self):
        file_path = os.path.abspath(os.path.join(os.getcwd(), os.path.join("ResourceFiles", "add_sections.xlsx")))
        shutil.copyfile(file_path, os.path.join(str(self.folder), "images_html", "add_sections.xlsx"))
        QMessageBox.information(QMessageBox(), 'Information', 'Your File is Downloaded in your selected workspace')
        #self.ui.pushButton_Import_Beam.setEnabled(True)

    def import_ColumnPref(self):
        wb = openpyxl.load_workbook(os.path.join(str(self.folder), "images_html", "add_sections.xlsx"))
        sheet = wb['First Sheet']
        conn = sqlite3.connect('ResourceFiles/Database/Intg_osdag.sqlite')

        for rowNum in range(2, sheet.max_row + 1):
            designation = sheet.cell(row=rowNum, column=2).value
            mass = sheet.cell(row=rowNum, column=3).value
            area = sheet.cell(row=rowNum, column=4).value
            d = sheet.cell(row=rowNum, column=5).value
            b = sheet.cell(row=rowNum, column=6).value
            tw = sheet.cell(row=rowNum, column=7).value
            t = sheet.cell(row=rowNum, column=8).value
            flangeSlope = sheet.cell(row=rowNum, column=9).value
            r1 = sheet.cell(row=rowNum, column=10).value
            r2 = sheet.cell(row=rowNum, column=11).value
            iz = sheet.cell(row=rowNum, column=12).value
            iy = sheet.cell(row=rowNum, column=13).value
            rz = sheet.cell(row=rowNum, column=14).value
            ry = sheet.cell(row=rowNum, column=15).value
            zz = sheet.cell(row=rowNum, column=16).value
            zy = sheet.cell(row=rowNum, column=17).value
            zpz = sheet.cell(row=rowNum, column=18).value
            zpy = sheet.cell(row=rowNum, column=19).value
            source = sheet.cell(row=rowNum, column=20).value
            c = conn.cursor()
            c.execute("SELECT count(*) FROM Columns WHERE Designation = ?", (designation,))
            data = c.fetchone()[0]
            if data == 0:
                c.execute('''INSERT INTO Columns (Designation,Mass,Area,D,B,tw,T,R1,R2,Iz,Iy,rz,ry,
    				                           Zz,zy,Zpz,Zpy,FlangeSlope,Source) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                          (designation, mass, area,
                           d, b, tw, t,
                           r1, r2, iz, iy, rz, ry,
                           zz, zy
                           ,
                           zpz, zpy, flangeSlope, source))
                conn.commit()
                c.close()

        conn.close()
        QMessageBox.information(QMessageBox(), 'Successful', ' File data is imported successfully to the database.')
        self.ui.pushButton_Import_Column.setDisabled(True)

    def import_BeamPref(self):
        wb = openpyxl.load_workbook(os.path.join(str(self.folder), "images_html", "add_sections.xlsx"))
        sheet = wb['First Sheet']
        conn = sqlite3.connect('ResourceFiles/Database/Intg_osdag.sqlite')

        for rowNum in range(2, sheet.max_row + 1):
            designation = sheet.cell(row=rowNum, column=2).value
            mass = sheet.cell(row=rowNum, column=3).value
            area = sheet.cell(row=rowNum, column=4).value
            d = sheet.cell(row=rowNum, column=5).value
            b = sheet.cell(row=rowNum, column=6).value
            tw = sheet.cell(row=rowNum, column=7).value
            t = sheet.cell(row=rowNum, column=8).value
            flangeSlope = sheet.cell(row=rowNum, column=9).value
            r1 = sheet.cell(row=rowNum, column=10).value
            r2 = sheet.cell(row=rowNum, column=11).value
            iz = sheet.cell(row=rowNum, column=12).value
            iy = sheet.cell(row=rowNum, column=13).value
            rz = sheet.cell(row=rowNum, column=14).value
            ry = sheet.cell(row=rowNum, column=15).value
            zz = sheet.cell(row=rowNum, column=16).value
            zy = sheet.cell(row=rowNum, column=17).value
            zpz = sheet.cell(row=rowNum, column=18).value
            zpy = sheet.cell(row=rowNum, column=19).value
            source = sheet.cell(row=rowNum, column=20).value

            c = conn.cursor()
            c.execute("SELECT count(*) FROM Beams WHERE Designation = ?", (designation,))
            data = c.fetchone()[0]
            if data == 0:
                c.execute('''INSERT INTO Beams (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Iz,Iy,rz,ry,
            				                           Zz,zy,Zpz,Zpy,Source) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                          (designation, mass, area,
                           d, b, tw, t,
                           flangeSlope, r1
                           ,
                           r2, iz, iy, rz, ry,
                           zz, zy
                           ,
                           zpz, zpy, source))
                conn.commit()
                c.close()

        conn.close()
        QMessageBox.information(QMessageBox(), 'Successful', ' File data is imported successfully to the database.')
        self.ui.pushButton_Import_Beam.setDisabled(True)

    def close_designPref(self):
        self.close()

    # def closeEvent(self, QCloseEvent):
    #     self.save_designPref_para()
    #     QCloseEvent.accept()


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    DesignPreferences = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(DesignPreferences)
    DesignPreferences.exec()
    sys.exit(app.exec_())
