""" Design of compression members (column)

@author: Rutvik Joshi,N swaroop,Adnan

Module: Compression Member Design

Reference:
            1) IS 800: 2007 General construction in steel - Code of practice (Third revision)

"""
import math
#from utils.common.is800_2007 import *
#from subprocess import call
#from C://Osdag-IITB//Osdag//utils//common//is800_2007 import *

import openpyxl
wb = openpyxl.load_workbook("C:/Users/Rutvik N Joshi/Desktop/py4e/Columns_Details - Copy.xlsx")
#print(wb.active.title)

class CompressionColumn(object):
    """
    A Function to determine whether to design for a column for any section(Y)
    or obtain axial load on certain column section type(N)
    or have custom section details and want to find strength(S)
    """

    def __init__(self):
        self.var1 = input('Want to design a column for any section(Y) or obtain axial load on certain column section type(N) or have custom section details and want to find strength(S) =' )
        print(self.var1)

#    def call_IS800_2007(self, path ='utils/common/is800_2007.py'):
#        self.path = path
#        call(["Python3", "{}".format(self.path)])

    '''Input Function'''
    def commoninput(self):
        if self.var1.lower() == 'y':  # b==var1
            self.var2 = int(input('Factored load(kilonewton) = '))  # a=var2
            self.flag = 1  # flag=var3
        elif self.var1.lower() == 'n':
            self.section = input('Section Name=').lower()
            self.var2 = int(input('Factored load(kilonewton) = '))
            self.flag = 0
        self.iteration = 0
        self.V1 = str(input(' Joint 1 displacement :Free(f) or Restrained(r):')).lower()
        self.theta1 = str(input('Joint 1 rotation:Free(f) or Restrained(r):')).lower()
        self.V2 = str(input('Joint 2 displacement :Free(f) or Restrained(r):')).lower()
        self.theta2 = str(input('Joint 2 rotation :Free(f) or Restrained(r):')).lower()

    def specialinput(self):
        if self.var1.lower() != 's':
            self.d = int(input('Yield stress{fy in MPa}= '))
            self.l = int(input('Length  in mm:'))
            self.e = int(input('Ultimate stress{fu in Kilonewton/sq.meters}= '))
            if self.flag == 1:
                print(
                    f'Factored load={self.var2}KN,Length of column={self.l}mm,Yield stress={self.d}MPa,Ultimate stress={self.e}KN/SQM,V1:{self.V1},theta1:{self.theta1},V2:{self.V2},theta2:{self.theta2}')  # ,End condition1={V1},End condition2={theta1},End condition3={V2},End condition4={theta2}')
            elif self.flag == 0:
                print(
                    f'section={self.section.upper()},Length of column={self.l}mm,Yield stress={self.d}MPa,Ultimate stress={self.e}KN/SQM,V1:{self.V1},theta1:{self.theta1},V2:{self.V2},theta2:{self.theta2}')  # ,End condition1={V1},End condition2={theta1},End condition3={V2},End condition4={theta2}
        else:
            self.mass = float(input('Enter mass kg/m:'))
            self.Aeff = float(input('Area cm2:'))
            self.h = float(input('D in mm:'))
            self.bf = float(input('B in mm:'))
            self.tw = float(input('tw in mm:'))
            self.tf = float(input('tf mm:'))
            self.rz = float(input('rz cm:'))
            self.ry = float(input('ry cm:'))
            self.r1 = float(input('r1 mm:'))
            self.d = float(input('Yield strength N/mm2:'))
            self.l = float(input('Eff. Length mm'))
            self.b1 = float(input('b1 in mm:'))
            self.d1 = float(input('d  of web mm:'))

    def initially_assume(self):
        self.g = 0.8
        self.Areacolumn = 4
        self.Areapr = 0
        self.sh1 = wb['Columns']

    def first_loop(self):
        if self.var1.lower() != 's':
            while self.d != 0 and self.var2 > 0:
                if self.g > 0.3:
                    self.g -= 0.1
                    self.iteration += 1
                    if self.iteration > 6 and self.g > 0.3:
                        print('Make changes structure unstable and run again')
                        break
                    print(f'g:{self.g}')
                    self.fcd = self.g * self.d
                    print(f'fcd={self.fcd}MPa')          #KN/M2
                    self.Areq = (self.var2 / (self.fcd * 0.1))  # approximate area required                                        step1
                    print(f'Areq:{self.Areq}cm2')
                    self.check = 1
                    self.mass = []
                    print(self.mass)
                    if self.flag == 5:
                        self.flag = 1
                    while True:
                        print('here2')
                        if self.flag != 0:
                            if self.flag == 1:
                                self.section = 'hb'
                            elif self.flag == 2:
                                self.section = 'pbp'
                                self.check = 1
                            elif self.flag == 3:
                                self.section = 'sc'
                            elif self.flag == 4:
                                self.section = 'uc'
                                self.check = 1
                            else:
                                break
                        while True:
                            print('here3')
                            if self.section == 'hb':
                                self.row = 1
                                for i in range(2, 19):
                                    self.excelvar = self.sh1.cell(i, 4).value
                                    self.excelmassvar = self.sh1.cell(i, 3).value
                                    self.row += 1
                                    if float(self.Areq) < float(self.excelvar):
                                        self.Areapr = float(self.excelvar)
                                        self.mass.append(self.excelmassvar)
                                # print(f'Areapr:{self.Areapr}')
                                break
                            if self.section == 'pbp':
                                if self.check == 1:
                                    self.check += 1
                                    self.row = 18
                                    for i in range(19, 24):
                                        self.excelvar = self.sh1.cell(i, 4).value
                                        self.excelmassvar = self.sh1.cell(i, 3).value
                                        self.row += 1
                                        if float(self.Areq) < float(self.excelvar):
                                            self.Areapr = float(self.excelvar)
                                            self.mass.append(self.excelmassvar)
                                if self.check == 2:
                                    self.check += 1
                                    self.row = 24
                                    for i in range(24, 33):
                                        self.excelvar = self.sh1.cell(i, 4).value
                                        self.excelmassvar = self.sh1.cell(i, 3).value
                                        self.row += 1
                                        if float(self.Areq) < float(self.excelvar):
                                            self.Areapr = float(self.excelvar)
                                            self.mass.append(self.excelmassvar)
                                if self.check == 3:
                                    self.check += 1
                                    self.row = 33
                                    for i in range(33, 38):
                                        self.excelvar = self.sh1.cell(i, 4).value
                                        self.excelmassvar = self.sh1.cell(i, 3).value
                                        self.row += 1
                                        if float(self.Areq) < float(self.excelvar):
                                            self.Areapr = float(self.excelvar)
                                            self.mass.append(self.excelmassvar)
                                if self.check == 4:
                                    self.check += 1
                                    self.row = 38
                                    for i in range(38, 41):
                                        self.excelvar = self.sh1.cell(i, 4).value
                                        self.excelmassvar = self.sh1.cell(i, 3).value
                                        self.row += 1
                                        if float(self.Areq) < float(self.excelvar):
                                            self.Areapr = float(self.excelvar)
                                            self.mass.append(self.excelmassvar)
                                if self.check == 5:
                                    self.check += 1
                                    self.row = 24
                                    for i in range(41, 48):
                                        self.excelvar = self.sh1.cell(i, 4).value
                                        self.excelmassvar = self.sh1.cell(i, 3).value
                                        self.row += 1
                                        if float(self.Areq) < float(self.excelvar):
                                            self.Areapr = float(self.excelvar)
                                            self.mass.append(self.excelmassvar)
                                break
                            if self.section == 'sc':
                                self.row = 47
                                for i in range(48, 57):
                                    self.excelvar = self.sh1.cell(i, 4).value
                                    self.excelmassvar = self.sh1.cell(i, 3).value
                                    self.row += 1
                                    if float(self.Areq) < float(self.excelvar):
                                        self.Areapr = float(self.excelvar)
                                        self.mass.append(self.excelmassvar)
                                # print(Areapr)
                                break
                            if self.section == 'uc':
                                # print('uc')
                                if self.check == 1:
                                    self.check += 1
                                    self.row = 56
                                    for i in range(57, 68):
                                        self.excelvar = self.sh1.cell(i, 4).value
                                        self.excelmassvar = self.sh1.cell(i, 3).value
                                        self.row += 1
                                        if float(self.Areq) < float(self.excelvar):
                                            self.Areapr = float(self.excelvar)
                                            # print(f'Areapr{Areapr}')
                                            self.mass.append(self.excelmassvar)
                                if self.check == 2:
                                    self.check += 1
                                    self.row = 67
                                    for i in range(68, 76):
                                        self.excelvar = self.sh1.cell(i, 4).value
                                        self.excelmassvar = self.sh1.cell(i, 3).value
                                        self.row += 1
                                        if float(self.Areq) < float(self.excelvar):
                                            self.Areapr = float(self.excelvar)
                                            self.mass.append(self.excelmassvar)
                                if self.check == 3:
                                    self.check += 1
                                    self.row = 75
                                    for i in range(76, 88):
                                        self.excelvar = self.sh1.cell(i, 4).value
                                        self.excelmassvar = self.sh1.cell(i, 3).value
                                        self.row += 1
                                        if float(self.Areq) < float(self.excelvar):
                                            self.Areapr = float(self.excelvar)
                                            self.mass.append(self.excelmassvar)
                            break
                        self.mass.sort()
                        if self.flag == 0:
                            if len(self.mass) == 0:
                                print('check again')
                                quit()
                            else:
                                # print('here')
                                break
                        if self.flag != 0:
                            print('Working on...')
                            print(self.Areq)
                            self.flag += 1
                            print(self.flag)
                            print(self.mass)
                            continue
                    if len(self.mass) == 0:
                        continue
                        # print(mass)
                if len(self.mass) != 0:
                    row = 1
                    # print(mass)
                    for i in range(2, 88):
                        row += 1
                        # print(row, sh1.cell(i,3).value)
                        if self.mass[0] == self.sh1.cell(i, 3).value:
                            break
                else:
                    continue
                self.Aeff = self.sh1.cell(row, 4).value
                self.h = self.sh1.cell(row, 5).value
                self.bf = self.sh1.cell(row, 6).value
                self.tf = self.sh1.cell(row, 8).value
                self.tw = self.sh1.cell(row, 7).value
                self.rz = self.sh1.cell(row, 14).value
                self.ry = self.sh1.cell(row, 15).value
                self.r1 = self.sh1.cell(row, 10).value
                if self.h / self.bf > 1.2:
                    if self.tf <= 40:
                        self.alphaxx = 0.21
                        self.alphayy = 0.34
                        self.buklingclass = 0.34
                    if 40 <= self.tf <= 100:
                        self.alphaxx = 0.34
                        self.alphayy = 0.49
                        self.buklingclass = 0.49
                if self.h / self.bf <= 1.2:
                    if self.tf <= 100:
                        self.alphaxx = 0.34
                        self.alphayy = 0.49
                        self.buklingclass = 0.49
                    if self.tf > 100:
                        self.alphaxx = 0.76
                        self.alphayy = 0.76
                        self.buklingclass = 0.76
#                self.alphaxx, self.alphayy , self.buklingclass = is800_2007.IS800_2007.cl_7_1_2_2_buckling_class_of_crosssections(self.h, self.bf, self.tf)
                self.epsilon = math.sqrt(250 / self.d)
                self.b1 = self.bf / 2
                self.d1 = self.h - 2 * (self.r1 + self.tf)
                if self.V1 == 'r' and self.theta1 == 'r' and self.V2 == 'r' and self.theta2 == 'r':  # l in mm
                    self.KL = 0.65 * self.l
                elif self.V1 == 'r' and self.theta1 == 'r' and self.V2 == 'r' and self.theta2 == 'f':
                    self.KL = 0.8 * self.l
                elif self.V1 == 'r' and self.theta1 == 'r' and self.V2 == 'f' and self.theta2 == 'r':
                    self.KL = 1.2 * self.l
                elif self.V1 == 'r' and self.theta1 == 'f' and self.V2 == 'r' and self.theta2 == 'f':
                    self.KL = 1 * self.l
                elif self.V1 == 'f' and self.theta1 == 'r' and self.V2 == 'f' and self.theta2 == 'r':
                    self.KL = 2 * self.l
                elif self.V1 == 'r' and self.theta1 == 'r' and self.V2 == 'f' and self.theta2 == 'f':
                    self.KL = 2 * self.l
                self.lamba = self.KL / (self.ry * 10)  # ry in cm
                if self.lamba <= 180:
                    self.lamba = self.lamba
                else:
                    self.lamba = 180
#                self.lamba= IS800_2007.cl_7_2_2_effective_length_of_prismatic_compression_members(self.V1, self.theta1, self.V2, self.theta2, self.l, self.ry)
#                self.facd= IS800_2007.cl_7_1_2_1_design_compressisive_stress(self.d, self.lamba, self.buklingclass)
                self.gammamo = 1.1  # step6   calculatin of fcd
                self.Esteel = 2 * (10 ** 5)
                self.fcc = (3.14 ** 2 * self.Esteel) / self.lamba ** 2
                self.lambda1 = (self.d / self.fcc) ** 0.5
                print(f'lambda1={self.lambda1}')
                self.phi = 0.5 * (1 + self.buklingclass * (self.lambda1 - 0.2) + self.lambda1 ** 2)
                self.facd = (self.d / self.gammamo) / (self.phi + ((self.phi ** 2 - self.lambda1 ** 2) ** 0.5))
                if self.facd <= self.d / self.gammamo and self.Aeff * self.facd / 10 > self.var2:
                    print(f'Design compressive strength={self.Aeff*self.facd/10}KN')
                    print(f'Aeff={self.Aeff}cm^2')
                    print(f'g:{self.g}')
                    print(f'fcd={self.fcd}MPa')  # KN/M2
                    print(f'Areq:{self.Areq}cm2')
                    print(f'Aeff={self.Aeff}cm^2')
                    print(f'row:{self.row}')
                    print(f'h:{self.h}mm')
                    print(f'bf:{self.bf}mm')
                    print(f'tf:{self.tf}mm')
                    print(f'tw:{self.tw}mm')
                    print(f'rz:{self.rz}cm')
                    print(f'ry:{self.ry}cm')
                    print(f'r1:{self.r1}mm')
                    print(f'buklingclass{self.buklingclass}')
                    print(f'alphaxx{self.alphaxx}')
                    print(f'alphayy{self.alphayy}')
                    print(
                        f'ryy is minor axis buckling class will happen first about y-y axis.So buckling class would be{self.buklingclass}')
                    if self.b1 / self.tf < 9.4 * self.epsilon and self.d / self.tw <= 42 * self.epsilon:  # doubts
                        print('plastic')
                    elif 9.4 * self.epsilon < self.b1 / self.tf < 10.5 * self.epsilon and self.d / self.tw <= 42 * self.epsilon:
                        print('compact')
                    elif 10.5 * self.epsilon < self.b1 / self.tf < 15.7 * self.epsilon and self.d / self.tw <= 42 * self.epsilon:
                        print('semi-compact')
                    print(f'lamba={self.lamba} mm')
                    print(f'b1:{self.b1}')
                    print(f'd1:{self.d1}')
                    print(f'epsilon:{self.epsilon}')
                    print(f'Esteel:{self.Esteel}')
                    print(f'fcc:{self.fcc}')
                    print(f'phi:{self.phi}')
                    print(f'facd={self.facd}KN/cm^2')
                    print(f'Design compressive strength={self.Aeff * self.facd / 10}kN')
                    print(f'Aeff={self.Aeff}cm^2')
                    break
        else:
            if self.h / self.bf > 1.2:
                if self.tf <= 40:
                    self.alphaxx = 0.21
                    self.alphayy = 0.34
                    self.buklingclass = 0.34
                if 40 <= self.tf <= 100:
                    self.alphaxx = 0.34
                    self.alphayy = 0.49
                    self.buklingclass = 0.49
            if self.h / self.bf <= 1.2:
                if self.tf <= 100:
                    self.alphaxx = 0.34
                    self.alphayy = 0.49
                    self.buklingclass = 0.49
                if self.tf > 100:
                    self.alphaxx = 0.76
                    self.alphayy = 0.76
                    self.buklingclass = 0.76
            print(f'buklingclass:{self.buklingclass}')
            print(f'alphaxx{self.alphaxx}')
            print(f'alphayy{self.alphayy}')
            print(
                f'ryy is minor axis buckling class will happen first about y-y axis.So buckling class would be:{self.buklingclass}')
            # step4
            self.epsilon = math.sqrt(250 / self.d)
            print(f'b1:{self.b1}')
            print(f'd1:{self.d1}')
            print(f'epsilon:{self.epsilon}')
            if self.b1 / self.tf < 9.4 * self.epsilon and self.d / self.tw <= 42 * self.epsilon:  # doubts
                print('plastic section')
            elif 9.4 * self.epsilon < self.b1 / self.tf < 10.5 * self.epsilon and self.d / self.tw <= 42 * self.epsilon:
                print('compact section')
            elif 10.5 * self.epsilon < self.b1 / self.tf < 15.7 * self.epsilon and self.d / self.tw <= 42 * self.epsilon:
                print('semi-compact section')
                # step5
            if self.V1 == 'r' and self.theta1 == 'r' and self.V2 == 'r' and self.theta2 == 'r':  # l in mm
                self.KL = 0.65 * self.l
            elif self.V1 == 'r' and self.theta1 == 'r' and self.V2 == 'r' and self.theta2 == 'f':
                self.KL = 0.8 * self.l
            elif self.V1 == 'r' and self.theta1 == 'r' and self.V2 == 'f' and self.theta2 == 'r':
                self.KL = 1.2 * self.l
            elif self.V1 == 'r' and self.theta1 == 'f' and self.V2 == 'r' and self.theta2 == 'f':
                self.KL = 1 * self.l
            elif self.V1 == 'f' and self.theta1 == 'r' and self.V2 == 'f' and self.theta2 == 'r':
                self.KL = 2 * self.l
            elif self.V1 == 'r' and self.theta1 == 'r' and self.V2 == 'f' and self.theta2 == 'f':
                self.KL = 2 * self.l
                # l in mm
            self.lamba = self.KL / (self.ry * 10)  # ry in cm
            if self.lamba <= 180:
                self.lamba = self.lamba
            else:
                self.lamba = 180
            print(f'lamba={self.lamba}')
            self.gammamo = 1.1  # step6   calculatin of fcd
            self.Esteel = 2 * (10 ** 5)
            print(f'Esteel{self.Esteel}')
            self.fcc = (3.14 ** 2 * self.Esteel) / self.lamba ** 2
            print(f'fcc:{self.fcc}')
            self.lambda1 = (self.d / self.fcc) ** 0.5
            print(f'lambda1={self.lambda1}')
            self.phi = 0.5 * (1 + self.buklingclass * (self.lambda1 - 0.2) + self.lambda1 ** 2)
            print(f'phi:{self.phi}')
            self.facd = (self.d / self.gammamo) / (self.phi + ((self.phi ** 2 - self.lambda1 ** 2) ** 0.5))
            print(f'facd={self.facd}KN/cm^2')
            print(f'Design compressive strength={self.Aeff * self.facd / 10}KN')
            print(f'Aeff={self.Aeff}cm^2')

    def output(self):
        print(self.var1)            #self.a , self.b
        print(wb.active.title)
        try:
            print(self.section)
        except :
           pass
        try:
            print(self.var2)
        except:
            pass
        try:
            print(self.flag)
        except :
           quit()
        quit()

#        def output2(self):
#            try:
#                print(f'section:{self.section}')
#            except ValueError:
#                quit()
#if __name__=="__main__" :
#    c=CompressionColumn()
#    c.call_python_file()
user=CompressionColumn()
user.commoninput()
user.specialinput()
user.initially_assume()
user.first_loop()
user.output()
#user=CompressionColumn.from_input()
#a=input(print('Hi'))
#user = CompressionColumn(a)
#print(user.b)
