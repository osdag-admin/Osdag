"""
Main application window for Osdag GUI.
Handles tab management, docking icons, and main window controls.
"""
import osdag_gui.resources.resources_rc

from PySide6.QtWidgets import QMainWindow
from PySide6.QtCore import QThread, Signal

import sys
import os, yaml
from pathlib import Path
from PySide6.QtWidgets import (
    QWidget, QPushButton, QVBoxLayout, QHBoxLayout, QApplication, QGridLayout, QFileDialog,
    QLabel, QMainWindow, QSizePolicy, QFrame, QScrollArea, QButtonGroup, QTabBar, QTabWidget,
)
from PySide6.QtSvgWidgets import QSvgWidget
from PySide6.QtCore import Qt, Signal, QSize, QEvent, QTimer, QPropertyAnimation, QEasingCurve, Property
from PySide6.QtGui import QFont, QIcon, QPainter, QColor, QGuiApplication, QPixmap
from PySide6.QtSvg import QSvgRenderer

from osdag_gui.ui.windows.home_window import HomeWindow
from osdag_gui.ui.windows.template_page import CustomWindow
from osdag_gui.ui.components.dialogs.custom_messagebox import CustomMessageBox, MessageBoxType

from osdag_gui.data.database.database_config import PROJECT_PATH, ID, update_project_path, delete_project_record

from osdag_gui.data.database.database_config import get_module_function
from osdag_core.Common import *
from osdag_core.design_type.connection.fin_plate_connection import FinPlateConnection
import openpyxl

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.main_widget_instance = None
        self.setWindowIcon(QIcon(":/images/osdag_logo.png"))
        self.setCursor(Qt.CursorShape.ArrowCursor)
        # Apply global QToolTip stylesheet here
        QApplication.instance().setStyleSheet("""
            QToolTip {
                background-color: #FFFFFF;
                color: #000000;
                border: 1px solid #90AF13;
                padding: 2px 2px;
                font-size: 10px;
                border-radius: 0px;
                qproperty-alignment: AlignVCenter;
            }
        """)

        screen = QGuiApplication.primaryScreen()
        screen_size = screen.availableGeometry()

        screen_width = screen_size.width()
        screen_height = screen_size.height()

        # Calculate window size
        window_width = int(7 * screen_width / 10)
        window_height = int((7 * screen_height) / 8)

        # Set window size
        self.resize(window_width, window_height)

        # Center the window
        x = int((screen_width - window_width) / 2)
        y = int((screen_height - window_height) / 2)

        self.setGeometry(x, y, window_width, window_height)
        self.setWindowFlags(Qt.FramelessWindowHint) # Make the window frameless for custom buttons
        self.current_tab_index = 0 # To keep track of the next tab index
        self.btn_size = QSize(46, 30)
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f4f4f4;
                border: 1px solid #90af13;
                margin: 0px;
                padding: 0px;
            } 
            QWidget#BottomLine {
                background-color: #90af13;
            }
        """)

        # Initialize UI first, as sidebar will overlay it
        self.init_ui() # Call init_ui before sidebar creation to ensure main content exists
        self.handle_add_tab("Home")

        # Using QTimer to delay maximizing until after the window is fully initialized
        # Before maximizing, so that when we click on Restore it comes to normal state.
        QTimer.singleShot(0, self.showMaximized)


    def init_ui(self):
        # Main Vertical Layout for the entire window's *content*
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_v_layout = QVBoxLayout(central_widget)
        main_v_layout.setContentsMargins(1, 0, 1, 1)
        main_v_layout.setSpacing(0)

        # --- Top HBox Layout (Contains logo, tabs, and window control buttons) ---
        top_h_layout = QHBoxLayout()
        top_h_layout.setContentsMargins(0, 0, 0, 0)
        top_h_layout.setSpacing(0)

        icon_label_widget = QWidget()
        icon_label_h_layout = QHBoxLayout(icon_label_widget)
        icon_label_h_layout.setContentsMargins(5, 0, 5, 0)
        icon_label_h_layout.setSpacing(0)

        # SVG Widget (Dummy SVG for demonstration)
        self.svg_widget = QSvgWidget()
        self.svg_widget.load(":/vectors/Osdag_logo.svg")
        self.svg_widget.setFixedSize(18, 18)

        icon_label_h_layout.addWidget(self.svg_widget)
        top_h_layout.addWidget(icon_label_widget)
        
        # Keep a reference for event filtering (double-click to maximize/restore)
        self.icon_label_widget = icon_label_widget

        tabs_h_layout = QHBoxLayout()
        tabs_h_layout.setSpacing(0)
        tabs_h_layout.setContentsMargins(0, 2, 0, 0)

        # QTabBar
        self.tab_bar = QTabBar()
        self.tab_bar.setExpanding(False)
        self.tab_bar.setTabsClosable(True)
        self.tab_bar.setMovable(False)
        self.tab_bar.tabCloseRequested.connect(self.handle_close_tab)
        # Custom tab style
        self.tab_bar.setStyleSheet('''
            QTabBar::tab {
                background: #F4F4F4;
                border-left: 1px solid #d9d7d7;
                border-right: 1px solid #d9d7d7;
                border-top: 1px solid #F4F4F4;
                border-bottom: 1px solid #F4F4F4;
                padding: 6px 18px 6px 18px;
                color: #000000;
                font-size: 11px;
                margin-left: 0px;
            }
            QTabBar::tab:selected {
                background: #ffffff;
                color: #000000;
                border: 1px solid #90AF13;
                border-bottom: 1px solid #ffffff;
                padding: 6px 18px 6px 18px;
            }
            QTabBar::tab:hover {
                border-top: 1px solid #90AF13;
                border-left: 1px solid #90AF13;
                border-right: 1px solid #90AF13;
            }
            QTabBar::close-button {
                image: url(:/vectors/window_close.svg);
                subcontrol-origin: padding;
                subcontrol-position: center right;
            }
            QTabBar::close-button:hover {
                image: url(:/vectors/window_close_hover.svg);
            }
        ''')
        tabs_h_layout.addWidget(self.tab_bar)
        top_h_layout.addLayout(tabs_h_layout)
        
        # Install event filters for double-click maximize/restore on title widgets
        self.tab_bar.installEventFilter(self)
        self.icon_label_widget.installEventFilter(self)

        # Stretch to push buttons to the right
        top_h_layout.addStretch(1)

        # Helper function to create a styled button
        def create_button(icon_svg, is_close=False):
            btn = QPushButton()
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #f4f4f4;
                    color: white;
                    border: none;
                    padding: 0px;
                }
                QPushButton:hover {
                    background-color: #d9d7d7;
                }
                QPushButton:pressed {
                    background-color: #cfcfcf;
                }
                QPushButton#close_button:hover {
                    background-color: #E81123;
                }
                QPushButton#close_button:pressed {
                    background-color: #F1707A;
                }
            """)
            btn.setFixedSize(self.btn_size)
            btn.setIcon(QIcon(QPixmap.fromImage(QPixmap(icon_svg).toImage())))
            btn.setIconSize(QSize(14, 14))
            if is_close:
                btn.setObjectName("close_button")
            return btn

        class ClickableSvgWidget(QSvgWidget):
            clicked = Signal()  # Define a custom clicked signal
            def __init__(self, parent=None):
                super().__init__(parent)
                self.setCursor(Qt.CursorShape.PointingHandCursor)

            def mousePressEvent(self, event):
                if event.button() == Qt.MouseButton.LeftButton:
                    self.clicked.emit()  # Emit the clicked signal on left-click
                super().mousePressEvent(event)

        # Control buttons
        control_button_layout = QHBoxLayout()
        control_button_layout.setSpacing(10)
        control_button_layout.setContentsMargins(5,5,5,5)

        self.input_dock_control = ClickableSvgWidget()
        self.input_dock_control.load(":/vectors/input_dock_active.svg")
        self.input_dock_control.setFixedSize(18, 18)
        self.input_dock_control.clicked.connect(self.toggle_input_dock)
        self.input_dock_active = True
        control_button_layout.addWidget(self.input_dock_control)

        self.log_dock_control = ClickableSvgWidget()
        self.log_dock_control.load(":/vectors/logs_dock_inactive.svg")
        self.log_dock_control.setFixedSize(18, 18)
        self.log_dock_control.clicked.connect(self.logs_dock_icon_toggle)
        self.log_dock_active = False
        control_button_layout.addWidget(self.log_dock_control)

        self.output_dock_control = ClickableSvgWidget()
        self.output_dock_control.load(":/vectors/output_dock_inactive.svg")
        self.output_dock_control.setFixedSize(18, 18)
        self.output_dock_control.clicked.connect(self.toggle_output_dock)
        self.output_dock_active = False
        control_button_layout.addWidget(self.output_dock_control)

        self.input_dock_control.hide()
        self.log_dock_control.hide()
        self.output_dock_control.hide()

        top_h_layout.addLayout(control_button_layout)

        self.minimize_button = create_button(":/vectors/window_minimize.svg")
        self.minimize_button.clicked.connect(self.showMinimized)
        top_h_layout.addWidget(self.minimize_button)

        self.maximize_button = create_button(":/vectors/window_maximize.svg")
        self.maximize_button.clicked.connect(self.toggle_maximize_restore)
        top_h_layout.addWidget(self.maximize_button)

        self.close_button = create_button(":/vectors/window_close.svg", is_close=True)
        self.close_button.clicked.connect(self.close)
        top_h_layout.addWidget(self.close_button)

        self.start_pos = None
        self.start_geometry = None

        # Add top HBox to main VBox
        main_v_layout.addLayout(top_h_layout)

        # QTabWidget
        self.tab_widget = QTabWidget()
        self.tab_widget.tabBar().hide()
        self.tab_widget.setTabsClosable(True) # Allow closing tabs
        self.tab_widget.setMovable(False) # Allow reordering tabs
        self.tab_widget.setStyleSheet("""
            QTabWidget {
                background-color: #ffffff;
                border: 0px;
            }
        """)
        self.tab_widget_content = []
        self.tab_widget.tabCloseRequested.connect(self.handle_close_tab)
        main_v_layout.addWidget(self.tab_widget)

        # Connect the QTabBar to custom handler
        self.tab_bar.currentChanged.connect(self.handle_tab_change)

        # Ensure initial synchronization
        if self.tab_bar.count() > 0:
            self.tab_widget.setCurrentIndex(self.tab_bar.currentIndex())

    def set_maximize_icon(self):
        self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_maximize.svg").toImage())))

    def set_restore_icon(self):
        self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_restore.svg").toImage())))

    def toggle_maximize_restore(self):
        """Toggles between maximized and normal window states and updates the icon."""
        if self.isMaximized():
            self.showNormal()
            self.set_maximize_icon()
        else:
            self.showMaximized()
            self.set_restore_icon()

    def add_new_tab(self, module):
        """Helper to add a new tab to QTabWidget."""
        body_widget = QWidget()
        
        # Create and set layout for body_widget first
        self.main_widget_layout = QHBoxLayout(body_widget)
        self.main_widget_layout.setContentsMargins(0, 0, 0, 0)
        self.main_widget_layout.setSpacing(0)

        # it initially sets the home on the Tab
        self.open_home_page(module)
        # False(dock icons show status), True(input dock show), False(logs dock show), False(output dock show)
        self.tab_widget_content.append([body_widget, False, True, False, False])
        self.tab_widget.addTab(body_widget, f"Tab {self.current_tab_index + 1}")
        # Update main_widget_layout to the layout of the new tab's body_widget
        if hasattr(body_widget, 'layout'):
            self.main_widget_layout = body_widget.layout()

    def handle_add_tab(self, module):
        """Handles the 'Add New Tab' button click."""
        self.current_tab_index += 1
        self.tab_bar.addTab("Home") # Add to tab bar
        # Set the newly added tab as current
        self.add_new_tab(module) # Add to tab widget
        
        new_index = self.tab_bar.count() - 1
        self.tab_bar.setCurrentIndex(new_index)
        self.tab_widget.setCurrentIndex(new_index)
        # Update docking icons for the newly added tab
        current_tab_data = self.tab_widget_content[new_index]
        self.update_docking_icons(current_tab_data[1], current_tab_data[2], current_tab_data[3], current_tab_data[4])
        
        # self.sidebar.raise_() # Ensure sidebar stays on top after new tab addition

    def handle_tab_change(self, index):
        # Switch the QTabWidget to the new tab
        if index < len(self.tab_widget_content) and index >= 0:
            self.tab_widget.setCurrentIndex(index)

            if self.tab_bar.tabText(index) == "Home":
                self.tab_widget_content[index][1] = False
            else:
                self.tab_widget_content[index][1] = True
                
            current_tab_data = self.tab_widget_content[index]
            # Update main_widget_instance to the main widget in the current tab
            body_widget = current_tab_data[0]
            if hasattr(body_widget, 'layout') and body_widget.layout().count() > 0:
                widget_item = body_widget.layout().itemAt(0)
                if widget_item is not None:
                    widget = widget_item.widget()
                    if widget is not None:
                        self.main_widget_instance = widget
            # Update main_widget_layout to the layout of the current tab's body_widget
            if hasattr(body_widget, 'layout'):
                self.main_widget_layout = body_widget.layout()

            # Update dock icons based on the new tab's state
            self.update_docking_icons(current_tab_data[1], current_tab_data[2], current_tab_data[3], current_tab_data[4])


    # This is triggered by Quit button in Menu bar on template_page
    def close_current_tab(self):
        current_index = self.tab_bar.currentIndex()
        self.handle_close_tab(current_index)

    # General closing function
    def handle_close_tab(self, index):

        tab_title = self.tab_bar.tabText(index) if index >= 0 else "Module"
        is_last_tab = self.tab_widget.count() == 1
        to_save = self._check_design_done(index)
        module = self._get_template_instance(index)
        
        if to_save and is_last_tab:
            result = CustomMessageBox(
                title="Confirm Exit",
                text=(
                    f"'{tab_title}' is the last tab.\n"
                     "Closing it will exit Osdag.\n"
                    f"Do you want to save your '{tab_title}' design before closing?"
                ),
                buttons=["Save and Exit", "Exit Without Saving", "Cancel"]
            ).exec()
            
            if result == "Save and Exit":
                # Call Save Function
                module.saveDesign_inputs()
                # Exit Osdag
                self.close()
            elif result == "Exit Without Saving":
                # Exit Osdag
                self.close()
        
        elif to_save:
            result = CustomMessageBox(
                title="Save Design",
                text=f" Do you want to Save Your '{tab_title}' design before closing?",
                buttons=["Yes", "No"],
                dialogType=MessageBoxType.Warning,
            ).exec()

            if result == "Yes":
                # Call Save Function
                module.saveDesign_inputs()
                self._close_tab(index)
            elif result == "No":
                # Close Tab
                self._close_tab(index)

        elif is_last_tab:
            result = CustomMessageBox(
                title="Confirm Exit",
                text=f"'{tab_title}' is the last tab.\nClosing it will exit Osdag.\nDo you really want to close this tab?",
                buttons=["Yes", "No"],
                dialogType=MessageBoxType.Warning,
            ).exec()

            # Handle result
            if result == "Yes":
                self.close()  # Close the main window (exit Osdag)
        else:
            self._close_tab(index)

    # Check if design is created in the module or not
    def _check_design_done(self, index) -> bool:
        module = self._get_template_instance(index)
        if hasattr(module, 'backend'):
            return module.backend.design_status
        else:
            return False
    
    def _get_template_instance(self, index) -> object:
        return self.tab_widget_content[index][0].layout().itemAt(0).widget()

    def _close_tab(self, index):
        """Handles closing of tabs."""
        self.tab_widget.removeTab(index)
        self.tab_bar.removeTab(index)
        self.tab_widget_content.pop(index)
        # synchronize with tab_bar
        self._synchronize_tab_widget()
        
    def _synchronize_tab_widget(self):
        current_index = self.tab_bar.currentIndex()
        self.tab_widget.setCurrentIndex(current_index)
        # Update global variables and icons
        current_tab_data = self.tab_widget_content[current_index]
        body_widget = current_tab_data[0]
        if hasattr(body_widget, 'layout') and body_widget.layout().count() > 0:
            widget = body_widget.layout().itemAt(0).widget()
            self.main_widget_instance = widget
        # Ensure main_widget_layout points to the currently active tab's layout
        if hasattr(body_widget, 'layout'):
            self.main_widget_layout = body_widget.layout()
        # Update docking icons using the current active tab index (not the closed one)
        self.update_docking_icons(current_tab_data[1], current_tab_data[2], current_tab_data[3], current_tab_data[4])

    # Allow dragging the window when frameless
    def mousePressEvent(self, event):
        # The draggable area is the combined height of the top_h_layout (tab bar + buttons) and the menu_bar
        if self.isMaximized():
            return
        draggable_height = self.tab_bar.height() + (self.layout().contentsMargins().top() * 2) # Account for potential margins/spacing
        # A more robust way might be to check if the cursor is within the bounding box of top_h_layout or menu_bar
        if event.button() == Qt.LeftButton and event.position().y() < draggable_height:
            self.old_pos = event.globalPosition().toPoint()

    def mouseMoveEvent(self, event):
        if self.isMaximized():
            return
        if hasattr(self, 'old_pos'):
            delta = event.globalPosition().toPoint() - self.old_pos
            self.move(self.x() + delta.x(), self.y() + delta.y())
            self.old_pos = event.globalPosition().toPoint()

    def mouseReleaseEvent(self, event):
        if self.isMaximized():
            return
        if event.button() == Qt.LeftButton:
            if hasattr(self, 'old_pos'):
                del self.old_pos

    def mouseDoubleClickEvent(self, event):
        # Toggle maximize/restore when double-clicking in the draggable title area
        if event.button() == Qt.LeftButton:
            draggable_height = self.tab_bar.height() + (self.layout().contentsMargins().top() * 2)
            if event.position().y() < draggable_height:
                self.toggle_maximize_restore()

    def eventFilter(self, obj, event):
        # Handle double-click on title widgets (e.g., tab bar, logo area)
        if event.type() == QEvent.MouseButtonDblClick:
            if event.button() == Qt.LeftButton:
                self.toggle_maximize_restore()
                return True
        return super().eventFilter(obj, event)

    def handle_card_open_clicked(self, card_title):
        if card_title == "Fin Plate":
            self.open_fin_plate_page()

    #-------------Functions-to-load-modules-in-Tabwidget-START---------------------------

    def open_fin_plate_page(self):
        title = "Fin Plate Connection"
        self.clear_layout(self.main_widget_layout)
        fin_plate = CustomWindow(title, FinPlateConnection, parent=self)

        # Load the last Design Inputs-start------------------------------------
        last_design_folder = os.path.join('ResourceFiles', 'last_designs')
        last_design_file = str(fin_plate.backend.module_name()).replace(' ', '') + ".osi"
        last_design_file = os.path.join(last_design_folder, last_design_file)
        last_design_dictionary = {}

        # Create folder if it doesn't exist
        if not os.path.isdir(last_design_folder):
            os.makedirs(last_design_folder)

        # Load previous design if file exists
        if os.path.isfile(last_design_file):
            with open(str(last_design_file), 'r') as last_design:
                last_design_dictionary = yaml.safe_load(last_design)
                fin_plate.setDictToUserInputs(last_design_dictionary)
        # Load the last Design Inputs-end------------------------------------

        self.main_widget_instance = fin_plate
        fin_plate.openNewTab.connect(self.handle_add_tab)
        fin_plate.downloadDatabase.connect(self.download_Database)
        self.main_widget_layout.addWidget(fin_plate)
        index = self.tab_bar.currentIndex()
        self.tab_bar.setTabText(index, title)
        # Show docking Icons
        self.tab_widget_content[index][1] = True
        current_tab_data = self.tab_widget_content[index]
        self.update_docking_icons(current_tab_data[1], current_tab_data[2], current_tab_data[3], current_tab_data[4])
    
    def open_home_page(self, module):
        self.clear_layout(self.main_widget_layout)
        home_window = HomeWindow()
        home_window.triggerLoadOsi.connect(self.common_osi_load)
        home_window.openProject.connect(self.handle_open_project)
        home_window.openModule.connect(self.handle_open_module)
        home_window.downloadDatabase.connect(self.download_Database)
        self.main_widget_instance = home_window
        home_window.set_active_button(module)
        home_window.cardOpenClicked.connect(self.handle_card_open_clicked)
        self.main_widget_layout.addWidget(home_window)

    # To open the recent module
    def handle_open_module(self, key:str):
        func = get_module_function(key)
        if func != 'None':
            func = getattr(self, func)
            func() # Open the Releated Module

    # To handle the click on open project of any recent project
    def handle_open_project(self, record: dict):
        self.common_osi_load(osi_path=record.get(PROJECT_PATH), id=record.get(ID))

    # Common function to load osi file and also to open recent project
    # If osi_path=None -> it triggers Load Osi else trigger open recent project
    def common_osi_load(self, osi_path=None, id=None):
        if osi_path is None:
            osi_path, _ = QFileDialog.getOpenFileName(self, "Open Design", os.path.join(str(' ')),
                                                  "InputFiles(*.osi)")
            
        else:
            if not Path(osi_path).exists():
                result = CustomMessageBox(
                    title="Warning",
                    text="Osi File has been moved, File does not exist!",
                    dialogType=MessageBoxType.Warning,
                    buttons=["Locate Osi", "Remove Record"]
                ).exec()
                if result == "Locate Osi":
                    file_dialog_path, _ = QFileDialog.getOpenFileName(self, "Locate Osi File", os.path.expanduser("~"), "InputFiles(*.osi)")
                    if file_dialog_path and id is not None:
                        osi_path = file_dialog_path
                        new_name = Path(osi_path).stem
                        try:
                            update_project_path(id, osi_path, new_name)
                        except Exception as e:
                            print(f"Failed to update project path: {e}")
                    else:
                        print("No file selected for relocation.")
                        return
                elif result == "Remove Record":
                    print("Remove Record")
                    if id is not None:
                        try:
                            delete_project_record(id)
                            # Also delete the latex files for this project
                            import shutil
                            report_folder = f"osdag_gui/data/reports/file_{id}"
                            try:
                                shutil.rmtree(report_folder)
                            except FileNotFoundError:
                                pass
                            except Exception as e:
                                print(f"Failed to delete report folder: {e}")
                            CustomMessageBox(
                                title="Record Removed",
                                text="The record has been removed from recent projects.",
                                dialogType=MessageBoxType.Information
                            ).exec()
                            # Update Home page with deleted project
                            self.main_widget_instance.show_home()

                        except Exception as e:
                            CustomMessageBox(
                                title="Error",
                                text=f"Failed to remove record: {e}",
                                dialogType=MessageBoxType.Critical
                            ).exec()
                    else:
                        print("No ID provided for record removal.")
                    return

        if not osi_path:
            print("No Path selected!")
            return
        try:
            in_file = str(osi_path)
            with open(in_file, 'r') as fileObject:
                uiObj = yaml.safe_load(fileObject)
            module = uiObj[KEY_MODULE]

            print("Osi File Belongs to. ", module)

            func = get_module_function(module)
            if func == 'None':
                CustomMessageBox(
                    title="Information",
                    text="Please load the appropriate Input",
                    dialogType=MessageBoxType.Information
                ).exec()
                print("Module Not Implemented yet.")
                return
            func = getattr(self, func)
            func()
            # Set variables in template page because it is opened project
            self.main_widget_instance.setDictToUserInputs(uiObj)
            self.main_widget_instance.project_id = id
            self.main_widget_instance.save_state = True

        except IOError:
            CustomMessageBox(
                title="Unable to open file",
                text="There was an error opening \"%s\"" % osi_path,
                dialogType=MessageBoxType.Critical
            ).exec()
            return    

    #-------------Functions-to-load-modules-in-Tabwidget-END------------------------------------

    #----------------------------Download-Database/Excel-END-----------------------------------------
    def download_Database(self, table, call_type="database"):

        fileName, _ = QFileDialog.getSaveFileName(QFileDialog(), "Download File", os.path.join(os.getcwd(), str(table+"_Details.xlsx")),
                                                  "SectionDetails(*.xlsx)")
        if not fileName:
            return
        try:
            conn = sqlite3.connect(PATH_TO_DATABASE)
            c = conn.cursor()
            header = get_db_header(table)
            wb = openpyxl.Workbook()
            sheet = wb.create_sheet(table, 0)

            col = 1
            for head in header:
                sheet.cell(row=1, column=col).value = head
                col += 1
            if call_type != "header":
                if table == 'Columns':
                    c.execute("SELECT * FROM Columns")
                elif table == 'Beams':
                    c.execute("SELECT * FROM Beams")
                elif table == 'Angles':
                    c.execute("SELECT * FROM Angles")
                elif table == 'Channels':
                    c.execute("SELECT * FROM Channels")
                data = c.fetchall()
                conn.commit()
                c.close()
                row = 2
                for rows in data:
                    col = 1
                    for cols in range(len(header)):
                        sheet.cell(row=row, column=col).value = rows[col - 1]
                        col += 1
                    row += 1
            wb.save(fileName)
            CustomMessageBox(
                title='Information',
                text='Your File is Downloaded.',
                dialogType=MessageBoxType.Information
            ).exec()

        except IOError:
            CustomMessageBox(
                title='Information',
                text='Unable to save file',
                informativeText="There was an error saving \"%s\"" % fileName,
                dialogType=MessageBoxType.Information
            ).exec()
            return
    #----------------------------Download-Database/Excel--END----------------------------------------

    def clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.setParent(None)
            else:
                self.clear_layout(item.layout())

    def update_docking_icons(self, docking_icons_active=None, input_is_active=None, log_is_active=None, output_is_active=None):
        index = self.tab_bar.currentIndex()
        current_tab_data = self.tab_widget_content[index]
        if(docking_icons_active is None):
            docking_icons_active = current_tab_data[1]

        # Update input dock icon
        if docking_icons_active:
            self.input_dock_control.show()
            self.output_dock_control.show()
            self.log_dock_control.show()

            if(input_is_active is not None):
                self.input_dock_active = input_is_active
                # Update and save control state
                self.tab_widget_content[index][2] = input_is_active
                if self.input_dock_active:
                    self.input_dock_control.load(":/vectors/input_dock_active.svg")
                else:
                    self.input_dock_control.load(":/vectors/input_dock_inactive.svg")
                            
            # Update output dock icon
            if(output_is_active is not None):
                self.output_dock_active = output_is_active
                # Update and save control state
                self.tab_widget_content[index][4] = output_is_active
                if self.output_dock_active:
                    self.output_dock_control.load(":/vectors/output_dock_active.svg")
                else:
                    self.output_dock_control.load(":/vectors/output_dock_inactive.svg")

            # Update log dock icon
            if(log_is_active is not None):
                self.log_dock_active = log_is_active
                # Update and save control state
                self.tab_widget_content[index][3] = log_is_active
                if self.log_dock_active:
                    self.log_dock_control.load(":/vectors/logs_dock_active.svg")
                else:
                    self.log_dock_control.load(":/vectors/logs_dock_inactive.svg")
        else:
            self.input_dock_control.hide()
            self.output_dock_control.hide()
            self.log_dock_control.hide()

    def toggle_input_dock(self):
        if self.main_widget_instance:
            self.main_widget_instance.input_dock_toggle()
    
    def toggle_output_dock(self):
        if self.main_widget_instance:
            self.main_widget_instance.output_dock_toggle()

    def logs_dock_icon_toggle(self):
        self.log_dock_active = not self.log_dock_active
        self.tab_widget_content[self.tab_bar.currentIndex()][3] = self.log_dock_active
        if self.log_dock_active:
            self.log_dock_control.load(":/vectors/logs_dock_active.svg")
        else:
            self.log_dock_control.load(":/vectors/logs_dock_inactive.svg")    

        if self.main_widget_instance:
            self.main_widget_instance.logs_dock_toggle(self.log_dock_active)

    def print_n_test(self):
        print(self)

if __name__ == "__main__":
    import sys, os
    sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
    from PySide6.QtWidgets import QApplication
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec())

