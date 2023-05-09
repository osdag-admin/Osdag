
from PyQt5 import QtGui, QtWidgets, QtCore
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *

from Common import *
from utils.common.Section_Properties_Calculator import *
from utils.common.component import *
from utils.common.other_standards import *
from design_type.connection.fin_plate_connection import FinPlateConnection
from design_type.connection.connection import Connection

import os
from drawing_2D.Svg_Window import SvgWindow
import sys
import sqlite3
import shutil
import openpyxl
from get_DPI_scale import scale,width,height


class MyTableWidget(QWidget):
    def __init__(self, parent):
        super(QWidget, self).__init__(parent)
        self.layout = QVBoxLayout(self)
        self.tabs = QTabWidget(self)
        #self.tabs.setStyleSheet("QTabBar::tab { height: 40px; width: 150px;}")
        self.layout.addWidget(self.tabs)
        #self.setLayout(self.layout)


    def addTab(self, widget, text):
        self.tabs.addTab(widget, text)
        widget.setAutoFillBackground(True)


class Window(QDialog):

    def __init__(self, main, input_dictionary):
        super().__init__()
        self.input_dictionary = input_dictionary
        self.do_not_clear_list = []
        self.save_changes_list = []
        self.values_changed = False
        for t in main.input_dictionary_design_pref(main):
            self.save_changes_list.extend(t[2])
        self.initUI(main,input_dictionary)
        # self.rejected.connect(self.close_message)

    def center(self):
        frameGm = self.frameGeometry()
        screen = QtWidgets.QApplication.desktop().screenNumber(QtWidgets.QApplication.desktop().cursor().pos())
        centerPoint = QtWidgets.QApplication.desktop().screenGeometry(screen).center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def closeEvent(self, event):
        if self.values_changed:
            popup = QMessageBox(self)
            popup.setIcon(QMessageBox.Information)
            popup.setWindowTitle("Save")
            popup.setText('Do you want to save the changes?')
            popup.setStandardButtons(QMessageBox.Yes |
                                     QMessageBox.No |
                                     QMessageBox.Cancel)
            popup.setDefaultButton(QMessageBox.Cancel)
            answer = popup.exec_()
            if answer == QMessageBox.Yes:
                self.accept()
                event.accept()
            elif answer == QMessageBox.No:
                self.reject()
                event.accept()
            elif answer == QMessageBox.Cancel:
                event.ignore()
        else:
            QDialog.closeEvent(self, event)

    def connect_widget_for_change(self, widget):
        if isinstance(widget, QComboBox):
            widget.currentIndexChanged.connect(self.something_changed)
        elif isinstance(widget, QLineEdit):
            widget.textChanged.connect(self.something_changed)

    def something_changed(self):
        self.values_changed = True

    def initUI(self,main,input_dictionary):

        button_size_x = scale*190
        button_size_y = scale*30
        #self.statusBar().showMessage('')
        #self.setGeometry(300, 300, 1170, 710)
        self.setObjectName("DesignPreferences")
        self.setWindowTitle('Design Preference')
        self.tabWidget = MyTableWidget(self)
        self.setLayout(self.tabWidget.layout)
        hlayout = QHBoxLayout()
        self.tabWidget.layout.addLayout(hlayout)
        self.btn_defaults = QPushButton()
        self.btn_defaults.setText("Defaults")
        self.btn_save = QPushButton()
        self.btn_save.setText("Save")
        hlayout.addWidget(self.btn_defaults)
        hlayout.addWidget(self.btn_save)
        self.btn_defaults.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
        self.btn_save.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
        self.btn_defaults.setFixedSize(button_size_x,button_size_y)
        self.btn_save.setFixedSize(button_size_x,button_size_y)

        tab_index = 0
        for tab_details in main.tab_list(main):
            last_title = ""
            tab_name = tab_details[0]
            tab_elements = tab_details[2]
            tab_type = tab_details[1]

            scrollArea = QScrollArea()
            scrollArea.setWidgetResizable(True)
            scrollAreaWidgetContents = QWidget()
            scrollArea.setWidget(scrollAreaWidgetContents)

            if tab_type == TYPE_TAB_1:

                tab = QWidget()
                self.tabWidget.addTab(tab, tab_name)
                tab_index +=1
                self.tabWidget.tabs.setTabText(tab_index, tab_name)
                tab.setObjectName(tab_name)

                lay = QVBoxLayout(tab)
                lay.addWidget(scrollArea)


                vertical = QVBoxLayout(scrollAreaWidgetContents)
                horizontalLayout = QHBoxLayout()
                vertical.addLayout(horizontalLayout)

                horizontal = QHBoxLayout()
                #hl1 = QFrame(tab)
                #hl1.setFrameShape(QFrame.HLine)
                #vertical.addWidget(hl1)
                lay.addLayout(horizontal)

                buttons = [(str("pushButton_Add_" + tab_name), 'Add'), (str("pushButton_Clear_" + tab_name), 'Clear'),
                            (str("pushButton_Import_" + tab_name), "Import xlsx file"), (str("pushButton_Download_" + tab_name), "Download xlsx file")]

                elements = tab_elements(main, input_dictionary)
                #elements = list(lmao)
                for i in range(len(buttons)):
                    object_name = buttons[i][0]
                    btn_text = buttons[i][1]
                    button = QPushButton(tab)
                    button.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
                    horizontal.addWidget(button)
                    button.setObjectName(object_name)
                    button.setText(btn_text)
                    button.setFixedSize(button_size_x, button_size_y)
                    if input_dictionary != {}:
                        if main.module_name(main) == KEY_DISP_BASE_PLATE and input_dictionary[KEY_CONN] == VALUES_CONN_BP[2]:
                            button.setEnabled(False)

                r = 1
                grid = QGridLayout()
                horizontalLayout.addLayout(grid)
                grid.setAlignment(Qt.AlignTop|Qt.AlignLeft)
                grid.setHorizontalSpacing(10)
                grid.setVerticalSpacing(10)

                for element in elements:
                    type = element[2]
                    lable = element[1]
                    if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
                        label = QLabel(tab)
                        label.setObjectName(element[0] + "_label")
                        label.setText("<html><head/><body><p>" + lable + "</p></body></html>")
                        grid.addWidget(label,r,1)
                        label.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))

                    if type ==TYPE_TEXTBOX:
                        line = QLineEdit(tab)
                        grid.addWidget(line,r,2)
                        line.setObjectName(element[0])
                        line.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
                        line.setFixedSize(85, 20)
                        if lable == 'Designation' or lable == KEY_DISP_SEC_PROFILE:
                            line.textChanged.connect(self.manage_designation_size(line))

                        if input_dictionary:
                            line.setText(str(element[4]))

                        if lable in [KEY_DISP_FU, KEY_DISP_FY, KEY_DISP_POISSON_RATIO, KEY_DISP_THERMAL_EXP,
                                     KEY_DISP_MOD_OF_ELAST, KEY_DISP_MOD_OF_RIGID, 'Source']:
                            line.setReadOnly(True)
                            self.do_not_clear_list.append(line)
                        if main.module_name(main) in [KEY_DISP_TENSION_BOLTED, KEY_DISP_TENSION_WELDED] and lable in \
                                [KEY_DISP_LOCATION, KEY_DISP_SEC_PROFILE]:
                            line.setReadOnly(True)
                            self.do_not_clear_list.append(line)
                        if last_title == KEY_DISP_DIMENSIONS:
                            if element[1] in [KEY_DISP_ROOT_R, KEY_DISP_TOE_R]:
                                regex_validator = QtCore.QRegExp("[0-9]*[.][0-9]*|[.][0-9]*|0")
                            else:
                                regex_validator = QtCore.QRegExp("[1-9][0-9]*[.][0-9]*|[.][0-9]*")
                            line.setValidator(QtGui.QRegExpValidator(regex_validator, line))
                        if last_title == KEY_DISP_SEC_PROP:
                            regex_validator = QtCore.QRegExp("[1-9][0-9]*[.][0-9]*|[.][0-9]*|N/A|-")
                            line.setValidator(QtGui.QRegExpValidator(regex_validator, line))

                        if element[0] in self.save_changes_list:
                            self.connect_widget_for_change(line)

                        r += 1

                    if type == TYPE_COMBOBOX:
                        combo = QComboBox(tab)
                        grid.addWidget(combo,r,2)
                        # combo.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
                        combo.setMaxVisibleItems(5)
                        combo.setObjectName(element[0])
                        combo.view().setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
                        combo.addItems(element[3])
                        if input_dictionary:
                            combo.setCurrentText(str(element[4]))
                        font = combo.font()
                        metrices = QtGui.QFontMetrics(font)
                        item_width = 0
                        item_width = max([metrices.boundingRect(item).width() for item in element[3]],default = 0)
                        combo.view().setMinimumWidth(item_width + 30)

                        combo.setStyleSheet("QComboBox { combobox-popup: 0; }")

                        if lable == KEY_DISP_MATERIAL:
                            combo.setFixedSize(115, 20)
                            self.do_not_clear_list.append(combo)
                        else:
                            combo.setFixedSize(85,20)

                        if element[0] in self.save_changes_list:
                            self.connect_widget_for_change(combo)
                        r += 1

                    if type == TYPE_TITLE:
                        title = QLabel(tab)
                        title.setText(lable)
                        grid.addWidget(title,r,1,1,2)
                        title.setObjectName("_title")
                        title.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
                        last_title = lable
                        r += 1

                    if type == TYPE_IMAGE:
                        img = QLabel(tab)
                        img.setObjectName(element[0])
                        grid.addWidget(img,r,1,10,2)
                        pmap = QPixmap(element[4])
                        img.setPixmap(pmap.scaled(300,300,Qt.KeepAspectRatio, Qt.FastTransformation)) # you can also use IgnoreAspectRatio
                        r += 10

                    if type == TYPE_BREAK:
                        r = 1
                        grid = QGridLayout()
                        horizontalLayout.addLayout(grid)
                        grid.setAlignment(Qt.AlignTop|Qt.AlignLeft)
                        grid.setHorizontalSpacing(10)
                        grid.setVerticalSpacing(10)
                        continue

            elif tab_type == TYPE_TAB_2:


                tab = QWidget()
                self.tabWidget.addTab(tab, tab_name)
                tab_index +=1
                self.tabWidget.tabs.setTabText(tab_index, tab_name)
                tab.setObjectName(tab_name)

                lay = QVBoxLayout(tab)
                lay.addWidget(scrollArea)


                vertical = QVBoxLayout(scrollAreaWidgetContents)
                horizontalLayout = QHBoxLayout()
                vertical.addLayout(horizontalLayout)


                r = 1
                grid = QGridLayout()
                horizontalLayout.addLayout(grid)
                grid.setHorizontalSpacing(10)
                grid.setVerticalSpacing(10)
                grid.setAlignment(Qt.AlignTop|Qt.AlignLeft)

                label_1 = QLabel(tab)
                label_1.setObjectName("_title")
                label_1.setText("Inputs")
                grid.addWidget(label_1,r,1)


                r += 3

                Notes = []
                elements = tab_elements(main, input_dictionary)
                for element in elements:
                    type = element[2]
                    lable = element[1]
                    if type in [TYPE_COMBOBOX, TYPE_TEXTBOX]:
                        label = QLabel(tab)
                        #label.setWordWrap(True)
                        label.setText("<html><head/><body><p>" + lable + "</p></body></html>")
                        label.setObjectName(element[0] + "_label")
                        grid.addWidget(label,r,1)
                        label.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))

                    if type == TYPE_TEXTBOX:
                        line = QLineEdit(tab)
                        grid.addWidget(line,r,2)
                        line.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
                        line.setObjectName(element[0])
                        line.setFixedSize(130, 22)
                        if element[3]:
                            line.setText(element[3])
                        dbl_validator = QDoubleValidator()
                        if element[0] in [KEY_DP_WELD_MATERIAL_G_O]:
                            line.setValidator(dbl_validator)
                            line.setMaxLength(7)
                        if element[0] in [KEY_DP_DETAILING_GAP] and main.module_name(main) in [KEY_DISP_TENSION_BOLTED, KEY_DISP_TENSION_WELDED]:
                            line.setReadOnly(True)
                            self.do_not_clear_list.append(line)
                        if element[0] in [KEY_BASE_PLATE_FU, KEY_BASE_PLATE_FY, KEY_DP_ANCHOR_BOLT_DESIGNATION_OCF,
                                          KEY_DP_ANCHOR_BOLT_DESIGNATION_ICF, KEY_DP_ANCHOR_BOLT_MATERIAL_G_O_OCF,
                                          KEY_DP_ANCHOR_BOLT_MATERIAL_G_O_ICF, KEY_DP_ANCHOR_BOLT_TYPE_OCF,
                                          KEY_DP_ANCHOR_BOLT_TYPE_ICF]:
                            line.setReadOnly(True)
                        if input_dictionary:
                            line.setText(str(element[4]))

                        if element[0] in self.save_changes_list:
                            self.connect_widget_for_change(line)
                        r += 1

                    if type == TYPE_COMBOBOX:
                        combo = QComboBox(tab)
                        grid.addWidget(combo,r,2)
                        combo.setMaxVisibleItems(5)
                        combo.setObjectName(element[0])
                        combo.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
                        combo.view().setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
                        combo.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
                        combo.setStyleSheet("QComboBox { combobox-popup: 0; }")
                        combo.setFixedSize(130, 22)
                        combo.addItems(element[3])
                        font = combo.font()
                        metrices = QtGui.QFontMetrics(font)
                        item_width = max([metrices.boundingRect(item).width() for item in element[3]],default = 0)
                        combo.view().setMinimumWidth(item_width + 30)
                        if element[0] == KEY_DP_DESIGN_METHOD:
                            combo.model().item(1).setEnabled(False)
                            combo.model().item(2).setEnabled(False)
                        if input_dictionary:
                            combo.setCurrentText(str(element[4]))
                        if element[0] in self.save_changes_list:
                            self.connect_widget_for_change(combo)
                        r += 1

                    if type == 'Title':
                        title = QLabel(tab)
                        title.setProperty("heading", True)
                        title.style().unpolish(title)
                        title.style().polish(title)
                        title.setText(element[1])
                        grid.addWidget(title,r,1)
                        title.setObjectName("_title")
                        title.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
                        r += 1

                    if type == 'Image':
                        img = QLabel(tab)
                        grid.addWidget(img,r,1,10,2)
                        pmap = QPixmap('C:/Users/nitin/Desktop/FOSSEE/Osdag3/ResourceFiles/images/Channel.png')
                        img.setPixmap(pmap.scaled(220,800,Qt.KeepAspectRatio, Qt.FastTransformation))
                        r += 10

                    if type == 'TextBrowser':
                        r = 1
                        grid = QGridLayout()
                        horizontalLayout.addLayout(grid)
                        grid.setHorizontalSpacing(10)
                        grid.setVerticalSpacing(10)
                        grid.setAlignment(Qt.AlignRight|Qt.AlignTop)
                        grid.setContentsMargins(50,0,0,0)
                        lbl = QLabel(tab)
                        lbl.setText('Description')
                        grid.addWidget(lbl,r,1)
                        lbl.setObjectName("label_3")
                        lbl.setSizePolicy(QSizePolicy(QSizePolicy.Maximum,QSizePolicy.Maximum))
                        r += 1

                        txt_browser = QTextBrowser(tab)
                        txt_browser.setHtml(element[3])
                        #txt_browser.setFixedHeight(480)
                        txt_browser.horizontalScrollBar().setVisible(False)
                        txt_browser.setObjectName(element[0])
                        grid.addWidget(txt_browser,r,1) # if using setMinimumSize
                        #grid.addWidget(txt_browser,r,10,2) # if using FixedSize, also use r+=10.


                    if type == 'Note':
                        Notes.append(lable)

                    if type == 'Break':
                        r = 1
                        grid = QGridLayout()
                        horizontalLayout.addLayout(grid)
                        grid.setHorizontalSpacing(10)
                        grid.setVerticalSpacing(10)
                        grid.setAlignment(Qt.AlignTop|Qt.AlignLeft)
                        continue

                if Notes:

                    hl1 = QFrame(tab)
                    hl1.setFrameShape(QFrame.HLine)
                    vertical.addWidget(hl1)
                    for lable in Notes:
                        lbl = QLabel(tab)
                        lbl.setWordWrap(True)
                        lbl.setText("<html><head/><body><p>" + lable + "</p></body></html>")
                        lbl.setObjectName("_title")
                        vertical.addWidget(lbl)


            scrollArea.setWidget(scrollAreaWidgetContents)

        # self.setCentralWidget(self.tabWidget)
        #self.tabWidget.resize(self.size())

        self.tabWidget.tabs.setCurrentIndex(2)
        #QtCore.QMetaObject.connectSlotsByName(DesignPreferences)
        module = main.module_name(main)

        if module in [KEY_DISP_FINPLATE, KEY_DISP_ENDPLATE, KEY_DISP_CLEATANGLE, KEY_DISP_SEATED_ANGLE, KEY_DISP_BCENDPLATE]:

            pushButton_Clear_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_COLSEC)
            pushButton_Clear_Column.clicked.connect(lambda: self.clear_tab(KEY_DISP_COLSEC))
            pushButton_Add_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_COLSEC)
            pushButton_Add_Column.clicked.connect(self.add_tab_column)
            pushButton_Import_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + KEY_DISP_COLSEC)
            pushButton_Import_Column.clicked.connect(lambda: self.import_section("Columns"))
            pushButton_Download_Column = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + KEY_DISP_COLSEC)
            pushButton_Download_Column.clicked.connect(lambda: self.download_Database(table="Columns", call_type="header"))
            pushButton_Clear_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_BEAMSEC)
            pushButton_Clear_Beam.clicked.connect(lambda: self.clear_tab(KEY_DISP_BEAMSEC))
            pushButton_Add_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_BEAMSEC)
            pushButton_Add_Beam.clicked.connect(self.add_tab_beam)
            pushButton_Import_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + KEY_DISP_BEAMSEC)
            pushButton_Import_Beam.clicked.connect(lambda: self.import_section("Beams"))
            pushButton_Download_Beam = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + KEY_DISP_BEAMSEC)
            pushButton_Download_Beam.clicked.connect(lambda: self.download_Database(table="Beams", call_type="header"))

            if module == KEY_DISP_CLEATANGLE:
                pushButton_Clear_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + DISP_TITLE_CLEAT)
                pushButton_Clear_Angle.clicked.connect(lambda: self.clear_tab(DISP_TITLE_CLEAT))
                pushButton_Add_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + DISP_TITLE_CLEAT)
                pushButton_Add_Angle.clicked.connect(self.add_tab_angle)
                pushButton_Import_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + DISP_TITLE_CLEAT)
                pushButton_Import_Angle.clicked.connect(lambda: self.import_section("Angles"))
                pushButton_Download_Angle = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + DISP_TITLE_CLEAT)
                pushButton_Download_Angle.clicked.connect(lambda: self.download_Database(table="Angles", call_type="header"))
            if module == KEY_DISP_SEATED_ANGLE:
                pushButton_Clear_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_SEATED_ANGLE)
                pushButton_Clear_Angle.clicked.connect(lambda: self.clear_tab(KEY_DISP_SEATED_ANGLE))
                pushButton_Add_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_SEATED_ANGLE)
                pushButton_Add_Angle.clicked.connect(self.add_tab_angle)
                pushButton_Import_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + KEY_DISP_SEATED_ANGLE)
                pushButton_Import_Angle.clicked.connect(lambda: self.import_section("Angles"))
                pushButton_Download_Angle = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + KEY_DISP_SEATED_ANGLE)
                pushButton_Download_Angle.clicked.connect(lambda: self.download_Database(table="Angles", call_type="header"))

        if module == KEY_DISP_COLUMNCOVERPLATE or module == KEY_DISP_COLUMNCOVERPLATEWELD or module == KEY_DISP_COLUMNENDPLATE:
            pushButton_Clear_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_COLSEC)
            pushButton_Clear_Column.clicked.connect(lambda: self.clear_tab(KEY_DISP_COLSEC))
            pushButton_Add_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_COLSEC)
            pushButton_Add_Column.clicked.connect(self.add_tab_column)
            pushButton_Import_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + KEY_DISP_COLSEC)
            pushButton_Import_Column.clicked.connect(lambda: self.import_section("Columns"))
            pushButton_Download_Column = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + KEY_DISP_COLSEC)
            pushButton_Download_Column.clicked.connect(lambda: self.download_Database(table="Columns", call_type="header"))

        if module == KEY_DISP_BEAMCOVERPLATE or module == KEY_DISP_BEAMCOVERPLATEWELD:
            pushButton_Clear_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_BEAMSEC)
            pushButton_Clear_Beam.clicked.connect(lambda: self.clear_tab(KEY_DISP_BEAMSEC))
            pushButton_Add_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_BEAMSEC)
            pushButton_Add_Beam.clicked.connect(self.add_tab_beam)
            pushButton_Import_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + KEY_DISP_BEAMSEC)
            pushButton_Import_Beam.clicked.connect(lambda: self.import_section("Beams"))
            pushButton_Download_Beam = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + KEY_DISP_BEAMSEC)
            pushButton_Download_Beam.clicked.connect(lambda: self.download_Database(table="Beams", call_type="header"))

        if module == KEY_DISP_BB_EP_SPLICE:
            pushButton_Clear_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_BEAMSEC)
            pushButton_Clear_Beam.clicked.connect(lambda: self.clear_tab(KEY_DISP_BEAMSEC))
            pushButton_Add_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_BEAMSEC)
            pushButton_Add_Beam.clicked.connect(self.add_tab_beam)
            pushButton_Import_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + KEY_DISP_BEAMSEC)
            pushButton_Import_Beam.clicked.connect(lambda: self.import_section("Beams"))
            pushButton_Download_Beam = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + KEY_DISP_BEAMSEC)
            pushButton_Download_Beam.clicked.connect(lambda: self.download_Database(table="Beams", call_type="header"))

        if module == KEY_DISP_COMPRESSION:
            pushButton_Clear_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_COLSEC)
            pushButton_Clear_Column.clicked.connect(lambda: self.clear_tab(KEY_DISP_COLSEC))
            pushButton_Add_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_COLSEC)
            pushButton_Add_Column.clicked.connect(self.add_tab_column)
            pushButton_Import_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + KEY_DISP_COLSEC)
            pushButton_Import_Column.clicked.connect(lambda: self.import_section("Columns"))
            pushButton_Download_Column = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + KEY_DISP_COLSEC)
            pushButton_Download_Column.clicked.connect(lambda: self.download_Database(table="Columns", call_type="header"))
            # pushButton_Clear_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_BEAMSEC)
            # pushButton_Clear_Beam.clicked.connect(lambda: self.clear_tab(KEY_DISP_BEAMSEC))
            # pushButton_Add_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_BEAMSEC)
            # pushButton_Add_Beam.clicked.connect(self.add_tab_beam)
            # pushButton_Import_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + KEY_DISP_BEAMSEC)
            # pushButton_Import_Beam.clicked.connect(lambda: self.import_section("Beams"))
            # pushButton_Download_Beam = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + KEY_DISP_BEAMSEC)
            # pushButton_Download_Beam.clicked.connect(lambda: self.download_Database(table="Beams", call_type="header"))

        if module == KEY_DISP_BASE_PLATE:
            pushButton_Clear_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + KEY_DISP_COLSEC)
            pushButton_Clear_Column.clicked.connect(lambda: self.clear_tab(KEY_DISP_COLSEC))
            pushButton_Add_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + KEY_DISP_COLSEC)
            pushButton_Add_Column.clicked.connect(self.add_tab_column)
            pushButton_Import_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + KEY_DISP_COLSEC)
            pushButton_Import_Column.clicked.connect(lambda: self.import_section("Columns"))
            pushButton_Download_Column = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + KEY_DISP_COLSEC)
            pushButton_Download_Column.clicked.connect(lambda: self.download_Database(table="Columns", call_type="header"))

        if module == KEY_DISP_TENSION_BOLTED or module == KEY_DISP_TENSION_WELDED:
            pushButton_Clear_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + DISP_TITLE_ANGLE)
            pushButton_Clear_Angle.clicked.connect(lambda: self.clear_tab(DISP_TITLE_ANGLE))
            pushButton_Add_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + DISP_TITLE_ANGLE)
            pushButton_Add_Angle.clicked.connect(self.add_tab_angle)
            pushButton_Import_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + DISP_TITLE_ANGLE)
            pushButton_Import_Angle.clicked.connect(lambda: self.import_section("Angles"))
            pushButton_Download_Angle = self.tabWidget.tabs.findChild(QWidget, "pushButton_Download_" + DISP_TITLE_ANGLE)
            pushButton_Download_Angle.clicked.connect(lambda: self.download_Database(table="Angles", call_type="header"))
            pushButton_Clear_Channel = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Clear_" + DISP_TITLE_CHANNEL)
            pushButton_Clear_Channel.clicked.connect(lambda: self.clear_tab(DISP_TITLE_CHANNEL))
            pushButton_Add_Channel = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Add_" + DISP_TITLE_CHANNEL)
            pushButton_Add_Channel.clicked.connect(self.add_tab_channel)
            pushButton_Import_Channel = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Import_" + DISP_TITLE_CHANNEL)
            pushButton_Import_Channel.clicked.connect(lambda: self.import_section("Channels"))
            pushButton_Download_Channel = self.tabWidget.tabs.findChild(QtWidgets.QWidget, "pushButton_Download_" + DISP_TITLE_CHANNEL)
            pushButton_Download_Channel.clicked.connect(lambda: self.download_Database(table="Channels", call_type="header"))

    def manage_designation_size(self,line_edit):
        def change_size():
            font = line_edit.font()
            text = line_edit.text()
            metrices = QtGui.QFontMetrics(font)
            width = metrices.boundingRect(text).width()
            width += 25
            if width > 91:
                line_edit.setFixedWidth(width)
            else:
                line_edit.setFixedWidth(91)
        return change_size

    def clear_tab(self, tab_name):
        '''
        @author: Umair
        '''
        tab = self.tabWidget.tabs.findChild(QtWidgets.QWidget, tab_name)

        if tab:
            for c in tab.findChildren(QtWidgets.QWidget):
                if c in self.do_not_clear_list:
                    continue

                if isinstance(c, QtWidgets.QComboBox):
                    c.setCurrentIndex(0)
                elif isinstance(c, QtWidgets.QLineEdit):
                    c.clear()

    def add_baseplate_tab_column(self):
        '''
        @author: Umair
        '''
        tab_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, KEY_DISP_COLSEC)
        rhs = connectdb("RHS", call_type="popup")
        shs = connectdb("SHS", call_type="popup")
        chs = connectdb("CHS", call_type="popup")
        hs = rhs + shs
        input_section = self.input_dictionary[KEY_SECSIZE]

        if input_section in hs:
            table = "RHS" if input_section in rhs else "SHS"
            values = {KEY_SECSIZE: '', 'Label_21': ''}
            for i in [1, 2, 3, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]:
                key = "Label_HS_"+str(i)
                values.update({key: ''})
        elif input_section in chs:
            table = "CHS"
            values = {KEY_SECSIZE: '', 'Label_21': ''}
            for i in [1, 2, 3, 11, 12, 13, 14, 15, 16]:
                key = "Label_CHS_" + str(i)
                values.update({key: ''})
        else:
            table = "Columns"
            values = {KEY_SECSIZE: '', 'Label_8': '', 'Label_21': ''}
            for i in [1, 2, 3, 11, 12, 13, 14, 15, 16]:
                key = "Label_" + str(i)
                values.update({key: ''})

        keys_to_add = values.keys()

        for ch in tab_Column.findChildren(QtWidgets.QWidget):
            if isinstance(ch, QtWidgets.QLineEdit) and ch.text() == "":
                QMessageBox.information(QMessageBox(), 'Warning', 'Please fill all the missing parameters!')
                return
            elif isinstance(ch, QtWidgets.QLineEdit) and ch.text() != "":
                if ch.objectName() in keys_to_add:
                    values[ch.objectName()] = ch.text()
            elif isinstance(ch, QtWidgets.QComboBox):
                if ch.objectName() in keys_to_add:
                    values[ch.objectName()] = ch.currentText()

        for k in keys_to_add:
            if k in [KEY_SECSIZE, "Label_21", "Label_8"]:
                continue
            else:
                values[key] = float(values[key])

        if ch:
            conn = sqlite3.connect(PATH_TO_DATABASE)
            c = conn.cursor()
            query = "SELECT count(*) FROM "+table+" WHERE Designation = ?"
            c.execute(query, (values[KEY_SECSIZE],))
            data = c.fetchone()[0]
            if data == 0:
                if table == "RHS":
                    c.execute('''INSERT INTO RHS (Designation,D,B,T,W,A,Izz,Iyy,Rzz,Ryy,
                        Zzz,Zyy,Zpz,Zpy,Source) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                              (values[KEY_SECSIZE], values["Label_HS_1"], values["Label_HS_2"],
                               values["Label_HS_3"], values["Label_HS_11"], values["Label_HS_12"],
                               values["Label_HS_13"], values["Label_HS_14"], values["Label_HS_15"],
                               values["Label_HS_16"], values["Label_HS_17"], values["Label_HS_18"],
                               values["Label_HS_19"], values["Label_HS_20"], values["Label_HS_21"],
                               ))
                    conn.commit()
                elif table == "SHS":
                    c.execute('''INSERT INTO SHS (Designation,D,B,T,W,A,Izz,Iyy,Rzz,Ryy,
                        Zzz,Zyy,Zpz,Zpy,Source) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                              (values[KEY_SECSIZE], values["Label_HS_1"], values["Label_HS_2"],
                               values["Label_HS_3"], values["Label_HS_11"], values["Label_HS_12"],
                               values["Label_HS_13"], values["Label_HS_14"], values["Label_HS_15"],
                               values["Label_HS_16"], values["Label_HS_17"], values["Label_HS_18"],
                               values["Label_HS_19"], values["Label_HS_20"], values["Label_HS_21"],
                               ))
                    conn.commit()
                elif table == "CHS":
                    c.execute('''INSERT INTO CHS (Designation,NB,OD,T,W,A,V,Ves,Vis,I,
                        Z,R,Rsq,Source) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                              (values[KEY_SECSIZE], values["Label_CHS_1"], values["Label_CHS_2"],
                               values["Label_CHS_3"], values["Label_HS_11"], values["Label_HS_12"],
                               values["Label_HS_13"], values["Label_HS_14"], values["Label_HS_15"],
                               values["Label_HS_16"], values["Label_HS_17"], values["Label_HS_18"],
                               values["Label_HS_19"], values["Label_HS_20"], values["Label_HS_21"],
                               ))
                    conn.commit()
                else:
                    c.execute('''INSERT INTO Columns (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Iz,Iy,rz,ry,
                        Zz,Zy,Zpz,Zpy,It,Iw,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                              (Designation_c, Mass_c, Area_c,
                               D_c, B_c, tw_c, T_c,FlangeSlope_c,
                               R1_c, R2_c, Iz_c, Iy_c, rz_c,
                               ry_c, Zz_c, Zy_c,
                               Zpz_c, Zpy_c, It_c,Iw_c,Source_c, Type))
                    conn.commit()
                c.close()
                conn.close()
                QMessageBox.information(QMessageBox(), 'Information', 'Data is added successfully to the database!')

            else:
                QMessageBox.information(QMessageBox(), 'Warning', 'Designation already exists in the database!')

    def add_tab_column(self):
        '''
        @author: Umair
        '''
        tab_Column = self.tabWidget.tabs.findChild(QtWidgets.QWidget, KEY_DISP_COLSEC)
        name = self.tabWidget.tabs.tabText(self.tabWidget.tabs.indexOf(tab_Column))
        #print(tab_Column.findChildren(QtWidgets.QWidget))
        if name in [KEY_DISP_COLSEC, KEY_DISP_SECSIZE]:
            table = "Columns"
        elif name == KEY_DISP_PRIBM:
            table = "Beams"
        else:
            pass
        for ch in tab_Column.findChildren(QtWidgets.QWidget):
            if isinstance(ch, QtWidgets.QLineEdit) and ch.text() == "":
                QMessageBox.information(QMessageBox(), 'Warning', 'Please fill all the missing parameters!')
                # add_col = tab_Column.findChild(QtWidgets.QWidget, 'pushButton_Add_'+KEY_DISP_COLSEC)
                # add_col.setDisabled(True)
                return
            elif isinstance(ch, QtWidgets.QLineEdit) and ch.text() != "":
                if ch.objectName() == KEY_SECSIZE or ch.objectName() == KEY_SUPTNGSEC:
                    Designation_c = ch.text()
                elif ch.objectName() == KEY_SOURCE:
                    Source_c = ch.text()
                elif ch.objectName() == 'Label_1':
                    D_c = float(ch.text())
                elif ch.objectName() == 'Label_2':
                    B_c = float(ch.text())
                elif ch.objectName() == 'Label_3':
                    T_c = float(ch.text())
                elif ch.objectName() == 'Label_4':
                    tw_c = float(ch.text())
                elif ch.objectName() == 'Label_5':
                    FlangeSlope_c = float(ch.text())
                elif ch.objectName() == 'Label_6':
                    R1_c = float(ch.text())
                elif ch.objectName() == 'Label_7':
                    R2_c = float(ch.text())
                elif ch.objectName() == 'Label_11':
                    Mass_c = float(ch.text())
                elif ch.objectName() == 'Label_12':
                    Area_c = float(ch.text())
                elif ch.objectName() == 'Label_13':
                    Iz_c = float(ch.text())
                elif ch.objectName() == 'Label_14':
                    Iy_c = float(ch.text())
                elif ch.objectName() == 'Label_15':
                    rz_c = float(ch.text())
                elif ch.objectName() == 'Label_16':
                    ry_c = float(ch.text())
                elif ch.objectName() == 'Label_17':
                    Zz_c = float(ch.text())
                elif ch.objectName() == 'Label_18':
                    Zy_c = float(ch.text())
                elif ch.objectName() == 'Label_19':
                    if ch.text() == "":
                        ch.setText("0")
                    Zpz_c = ch.text()
                elif ch.objectName() == 'Label_20':
                    if ch.text() == "":
                        ch.setText("0")
                    Zpy_c = ch.text()
                elif ch.objectName() == 'Label_21':
                    if ch.text() == "":
                        ch.setText("0")
                    It_c = ch.text()
                elif ch.objectName() == 'Label_22':
                    if ch.text() == "":
                        ch.setText("0")
                    Iw_c = ch.text()
                else:
                    pass
            elif isinstance(ch, QtWidgets.QComboBox):
                if ch.objectName() == 'Label_8':
                    Type = ch.currentText()

        # if ch.objectName() == "pushButton_Download_" + name:   # If Download button
        if ch:
            conn = sqlite3.connect(PATH_TO_DATABASE)
            c = conn.cursor()
            if table == "Beams":
                c.execute("SELECT count(*) FROM Beams WHERE Designation = ?", (Designation_c,))
                data = c.fetchone()[0]
            else:
                c.execute("SELECT count(*) FROM Columns WHERE Designation = ?", (Designation_c,))
                data = c.fetchone()[0]
            if data == 0:
                if table == "Beams":
                    c.execute('''INSERT INTO Beams (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Iz,Iy,rz,ry,
                        Zz,Zy,Zpz,Zpy,It,Iw,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                              (Designation_c, Mass_c, Area_c,
                               D_c, B_c, tw_c, T_c,FlangeSlope_c,
                               R1_c, R2_c, Iz_c, Iy_c, rz_c,
                               ry_c, Zz_c, Zy_c,
                               Zpz_c, Zpy_c, It_c,Iw_c,Source_c, Type))
                    conn.commit()
                else:
                    c.execute('''INSERT INTO Columns (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Iz,Iy,rz,ry,
                        Zz,Zy,Zpz,Zpy,It,Iw,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                              (Designation_c, Mass_c, Area_c,
                               D_c, B_c, tw_c, T_c,FlangeSlope_c,
                               R1_c, R2_c, Iz_c, Iy_c, rz_c,
                               ry_c, Zz_c, Zy_c,
                               Zpz_c, Zpy_c, It_c,Iw_c,Source_c, Type))
                    conn.commit()
                c.close()
                conn.close()
                QMessageBox.information(QMessageBox(), 'Information', 'Data is added successfully to the database!')

            else:
                QMessageBox.information(QMessageBox(), 'Warning', 'Designation already exists in the database!')

    def add_tab_beam(self):
        '''
        @author: Umair
        '''
        tab_Beam = self.tabWidget.tabs.findChild(QtWidgets.QWidget, KEY_DISP_BEAMSEC)
        name = self.tabWidget.tabs.tabText(self.tabWidget.tabs.indexOf(tab_Beam))
        for ch in tab_Beam.findChildren(QtWidgets.QWidget):
            if isinstance(ch, QtWidgets.QLineEdit) and ch.text() == "":
                QMessageBox.information(QMessageBox(), 'Warning', 'Please fill all the missing parameters!')
                add_bm = tab_Beam.findChild(QtWidgets.QWidget, 'pushButton_Add_'+KEY_DISP_BEAMSEC)
                add_bm.setDisabled(True)
                return

            elif isinstance(ch, QtWidgets.QLineEdit) and ch.text() != "":

                if ch.objectName() == KEY_SECSIZE or ch.objectName() == KEY_SUPTDSEC:
                    Designation_b = ch.text()
                elif ch.objectName() == KEY_SOURCE:
                    Source_b = ch.text()
                elif ch.objectName() == 'Label_1':
                    D_b = float(ch.text())
                elif ch.objectName() == 'Label_2':
                    B_b = float(ch.text())
                elif ch.objectName() == 'Label_3':
                    T_b = float(ch.text())
                elif ch.objectName() == 'Label_4':
                    tw_b = float(ch.text())
                elif ch.objectName() == 'Label_5':
                    FlangeSlope_b = float(ch.text())
                elif ch.objectName() == 'Label_6':
                    R1_b = float(ch.text())
                elif ch.objectName() == 'Label_7':
                    R2_b = float(ch.text())
                elif ch.objectName() == 'Label_11':
                    Mass_b = float(ch.text())
                elif ch.objectName() == 'Label_12':
                    Area_b = float(ch.text())
                elif ch.objectName() == 'Label_13':
                    Iz_b = float(ch.text())
                elif ch.objectName() == 'Label_14':
                    Iy_b = float(ch.text())
                elif ch.objectName() == 'Label_15':
                    rz_b = float(ch.text())
                elif ch.objectName() == 'Label_16':
                    ry_b = float(ch.text())
                elif ch.objectName() == 'Label_17':
                    Zz_b = float(ch.text())
                elif ch.objectName() == 'Label_18':
                    Zy_b = float(ch.text())
                elif ch.objectName() == 'Label_19':
                    if ch.text() == "":
                        ch.setText("0")
                    Zpz_b = ch.text()
                elif ch.objectName() == 'Label_20':
                    if ch.text() == "":
                        ch.setText("0")
                    Zpy_b = ch.text()
                elif ch.objectName() == 'Label_21':
                    if ch.text() == "":
                        ch.setText("0")
                    I_t = ch.text()
                elif ch.objectName() == 'Label_22':
                    if ch.text() == "":
                        ch.setText("0")
                    I_w = ch.text()
                else:
                    pass
            elif isinstance(ch, QtWidgets.QComboBox):
                if ch.objectName() == 'Label_8':
                    Type = ch.currentText()

        # if ch.objectName() ==  "pushButton_Download_" + name:
        if ch:
            conn = sqlite3.connect(PATH_TO_DATABASE)

            c = conn.cursor()
            c.execute("SELECT count(*) FROM Beams WHERE Designation = ?", (Designation_b,))
            data = c.fetchone()[0]
            if data == 0:
                c.execute('''INSERT INTO Beams (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Iz,Iy,rz,ry,Zz,Zy,Zpz,Zpy,
                    It,Iw,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                          (Designation_b, Mass_b, Area_b,
                           D_b, B_b, tw_b, T_b, FlangeSlope_b,
                           R1_b, R2_b, Iz_b, Iy_b, rz_b,
                           ry_b, Zz_b, Zy_b,
                           Zpz_b, Zpy_b,I_t,I_w, Source_b, Type))
                conn.commit()
                c.close()
                conn.close()
                QMessageBox.information(QMessageBox(), 'Information', 'Data is added successfully to the database.')
            else:
                QMessageBox.information(QMessageBox(), 'Warning', 'Designation already exists in the database!')

    def add_tab_angle(self):
        '''
        @author: Umair
        '''

        tab_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, DISP_TITLE_ANGLE)
        tab_name = DISP_TITLE_ANGLE
        if tab_Angle == None:
            tab_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, DISP_TITLE_CLEAT)
            tab_name  = DISP_TITLE_CLEAT
        if tab_Angle == None:
            tab_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, KEY_DISP_SEATED_ANGLE)
            tab_name  = KEY_DISP_SEATED_ANGLE
        if tab_Angle == None:
            tab_Angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, KEY_DISP_TOPANGLE)
            tab_name = KEY_DISP_TOPANGLE
        # tab_cleat_angle = self.tabWidget.tabs.findChild(QtWidgets.QWidget, DISP_TITLE_CLEAT)
        # name = self.tabWidget.tabs.tabText(self.tabWidget.tabs.indexOf(tab_Angle))
        if self.add_compound_section(tab_Angle):
            return
        for ch in tab_Angle.findChildren(QtWidgets.QWidget):
            if isinstance(ch, QtWidgets.QLineEdit) and ch.text() == "":
                QMessageBox.information(QMessageBox(), 'Warning', 'Please fill all the missing parameters!')
                add_bm = tab_Angle.findChild(QtWidgets.QWidget, 'pushButton_Add_'+tab_name)
                add_bm.setDisabled(True)
                return

            elif isinstance(ch, QtWidgets.QLineEdit) and ch.text() != "":

                if ch.objectName() == KEY_SECSIZE_SELECTED or ch.objectName() == KEY_ANGLE_SELECTED:
                    Designation_a = ch.text()
                elif ch.objectName() == KEY_SOURCE:
                    Source = ch.text()
                elif ch.objectName() == 'Label_1':
                    a = ch.text()
                elif ch.objectName() == 'Label_2':
                    b = ch.text()
                elif ch.objectName() == 'Label_3':
                    t = float(ch.text())
                elif ch.objectName() == 'Label_4':
                    R1 = float(ch.text())
                elif ch.objectName() == 'Label_5':
                    R2 = float(ch.text())
                elif ch.objectName() == 'Label_7':
                    Cz = float(ch.text())
                elif ch.objectName() == 'Label_8':
                    Cy = float(ch.text())
                elif ch.objectName() == 'Label_9':
                    Mass = float(ch.text())
                elif ch.objectName() == 'Label_10':
                    Area = float(ch.text())
                elif ch.objectName() == 'Label_11':
                    I_z = float(ch.text())
                elif ch.objectName() == 'Label_12':
                    I_y = float(ch.text())
                elif ch.objectName() == 'Label_13':
                    I_u_max = float(ch.text())
                elif ch.objectName() == 'Label_14':
                    I_v_min = float(ch.text())
                elif ch.objectName() == 'Label_15':
                    rz = float(ch.text())
                elif ch.objectName() == 'Label_16':
                    ry = float(ch.text())
                elif ch.objectName() == 'Label_17':
                    if ch.text() == "":
                        ch.setText("0")
                    ru_max = float(ch.text())
                elif ch.objectName() == 'Label_18':
                    if ch.text() == "":
                        ch.setText("0")
                    rv_min = ch.text()
                elif ch.objectName() == 'Label_19':
                    if ch.text() == "":
                        ch.setText("0")
                    zz = ch.text()
                elif ch.objectName() == 'Label_20':
                    if ch.text() == "":
                        ch.setText("0")
                    zy = ch.text()
                elif ch.objectName() == 'Label_21':
                    if ch.text() == "":
                        ch.setText("0")
                    zpz = ch.text()
                elif ch.objectName() == 'Label_22':
                    if ch.text() == "":
                        ch.setText("0")
                    zpy = ch.text()
                elif ch.objectName() == 'Label_23':
                    if ch.text() == "":
                        ch.setText("0")
                    It = ch.text()

                else:
                    pass
            elif isinstance(ch, QtWidgets.QComboBox):
                if ch.objectName() == 'Label_6':
                    Type = ch.currentText()

        # if ch.objectName() ==  "pushButton_Download_" + name:
        if ch:
            conn = sqlite3.connect(PATH_TO_DATABASE)

            c = conn.cursor()

            c.execute("SELECT count(*) FROM Angles WHERE Designation = ?", (Designation_a,))
            data = c.fetchone()[0]
            if data == 0:
                c.execute('''INSERT INTO Angles (Designation,Mass,Area,a,b,t,R1,R2,Cz,Cy,Iz,Iy,Iumax,Ivmin,rz,ry,
                rumax,rvmin,Zz,Zy,Zpz,Zpy,It,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                          (Designation_a, Mass, Area,
                           a,b, t, R1, R2, Cz,Cy,I_z,I_y,I_u_max,
                           I_v_min, rz, ry, ru_max, rv_min,zz,zy,zpz,zpy,It,Source,Type))
                conn.commit()
                c.close()
                conn.close()
                QMessageBox.information(QMessageBox(), 'Information', 'Data is added successfully to the database.')
            else:
                QMessageBox.information(QMessageBox(), 'Warning', 'Designation already exists in the database!')

    def add_tab_channel(self):
        '''
        @author: Umair
        '''
        tab_Channel = self.tabWidget.tabs.findChild(QtWidgets.QWidget, DISP_TITLE_CHANNEL)

        if self.add_compound_section(tab_Channel):
            return

        name = self.tabWidget.tabs.tabText(self.tabWidget.tabs.indexOf(tab_Channel))
        for ch in tab_Channel.findChildren(QtWidgets.QWidget):
            if isinstance(ch, QtWidgets.QLineEdit) and ch.text() == "":
                QMessageBox.information(QMessageBox(), 'Warning', 'Please fill all the missing parameters!')
                add_bm = tab_Channel.findChild(QtWidgets.QWidget, 'pushButton_Add_'+DISP_TITLE_ANGLE)
                add_bm.setDisabled(True)
                return

            elif isinstance(ch, QtWidgets.QLineEdit) and ch.text() != "":

                if ch.objectName() == KEY_SECSIZE_SELECTED:
                    Designation_c = ch.text()
                elif ch.objectName() == KEY_SOURCE:
                    Source = ch.text()
                elif ch.objectName() == 'Label_1':
                    B = float(ch.text())
                elif ch.objectName() == 'Label_2':
                    T = float(ch.text())
                elif ch.objectName() == 'Label_3':
                    D = float(ch.text())
                elif ch.objectName() == 'Label_13':
                    t_w = float(ch.text())
                elif ch.objectName() == 'Label_14':
                    Flange_Slope = float(ch.text())
                elif ch.objectName() == 'Label_4':
                    R1 = float(ch.text())
                elif ch.objectName() == 'Label_5':
                    R2 = float(ch.text())
                elif ch.objectName() == 'Label_9':
                    Mass = float(ch.text())
                elif ch.objectName() == 'Label_10':
                    Area = float(ch.text())
                elif ch.objectName() == 'Label_17':
                    if ch.text() == "":
                        ch.setText("0")
                    cy = float(ch.text())
                elif ch.objectName() == 'Label_11':
                    I_z = float(ch.text())
                elif ch.objectName() == 'Label_12':
                    I_y = float(ch.text())
                elif ch.objectName() == 'Label_15':
                    rz = float(ch.text())
                elif ch.objectName() == 'Label_16':
                    ry = float(ch.text())

                elif ch.objectName() == 'Label_19':
                    if ch.text() == "":
                        ch.setText("0")
                    zz = ch.text()
                elif ch.objectName() == 'Label_20':
                    if ch.text() == "":
                        ch.setText("0")
                    zy = ch.text()
                elif ch.objectName() == 'Label_21':
                    if ch.text() == "":
                        ch.setText("0")
                    zpz = ch.text()
                elif ch.objectName() == 'Label_22':
                    if ch.text() == "":
                        ch.setText("0")
                    zpy = ch.text()
                elif ch.objectName() == 'Label_26':
                    if ch.text() == "":
                        ch.setText("0")
                    It = ch.text()
                elif ch.objectName() == 'Label_27':
                    if ch.text() == "":
                        ch.setText("0")
                    Iw = ch.text()

                else:
                    pass
            elif isinstance(ch, QtWidgets.QComboBox):
                if ch.objectName() == 'Label_6':
                    Type = ch.currentText()

        # if ch.objectName() ==  "pushButton_Download_" + name:
        if ch:
            conn = sqlite3.connect(PATH_TO_DATABASE)

            c = conn.cursor()
            c.execute("SELECT count(*) FROM Channels WHERE Designation = ?", (Designation_c,))
            data = c.fetchone()[0]
            if data == 0:
                c.execute('''INSERT INTO Channels (Designation,Mass, Area,D,B,tw,T,FlangeSlope, R1, R2,Cy,Iz,Iy,
                 rz, ry,Zz,Zy,Zpz,Zpy,It,Iw,Source,Type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                          (Designation_c, Mass, Area,D,B,t_w,T,
                           Flange_Slope, R1, R2,cy,I_z,I_y, rz, ry,zz,zy,zpz,zpy,It, Iw,Source,Type))
                conn.commit()
                c.close()
                conn.close()
                QMessageBox.information(QMessageBox(), 'Information', 'Data is added successfully to the database.')
            else:
                QMessageBox.information(QMessageBox(), 'Warning', 'Designation already exists in the database!')

    def add_compound_section(self, tab):
        if tab.findChild(QWidget, KEY_SEC_PROFILE):
            if tab.findChild(QWidget, KEY_SEC_PROFILE).text() in ['Back to Back Angles', 'Star Angles', 'Back to Back Channels']:
                QMessageBox.information(QMessageBox(), "Information", "To create new compound section please add as single section")
                return True
            else:
                return False
        else:
            return False

    def download_Database(self, table, call_type="database"):

        fileName, _ = QFileDialog.getSaveFileName(QFileDialog(), "Download File", os.path.join(os.getcwd(), str(table+"_Details.xlsx")),
                                                  "SectionDetails(*.xlsx)")
        if not fileName:
            return
        try:
            conn = sqlite3.connect(PATH_TO_DATABASE)
            c = conn.cursor()
            header = get_db_header(table)
            wb = openpyxl.Workbook()
            sheet = wb.create_sheet(table, 0)

            col = 1
            for head in header:
                sheet.cell(row=1, column=col).value = head
                col += 1
            if call_type != "header":
                if table == 'Columns':
                    c.execute("SELECT * FROM Columns")
                elif table == 'Beams':
                    c.execute("SELECT * FROM Beams")
                elif table == 'Angles':
                    c.execute("SELECT * FROM Angles")
                elif table == 'Channels':
                    c.execute("SELECT * FROM Channels")
                data = c.fetchall()
                conn.commit()
                c.close()
                row = 2
                for rows in data:
                    col = 1
                    for cols in range(len(header)):
                        sheet.cell(row=row, column=col).value = rows[col - 1]
                        col += 1
                    row += 1
            wb.save(fileName)
            QMessageBox.information(QMessageBox(), 'Information', 'Your File is Downloaded.')

        except IOError:
            QMessageBox.information(QMessageBox(), "Unable to save file",
                                    "There was an error saving \"%s\"" % fileName)
            return

    # def download_Database_Beam(self):
    #     file_path = os.path.abspath(os.path.join(os.getcwd(), os.path.join("ResourceFiles", "add_sections.xlsx")))
    #     shutil.copyfile(file_path, os.path.join(str(self.folder), "images_html", "add_sections.xlsx"))
    #     QMessageBox.information(QMessageBox(), 'Information', 'Your File is Downloaded in your selected workspace')
    #     # self.ui.pushButton_Import_Beam.setEnabled(True)

    def import_section(self, tab_name):
        fileName, _ = QFileDialog.getOpenFileName(QFileDialog(), "Open File", os.getcwd(),
                                                  "SectionDetails(*.xlsx)")
        if not fileName:
            return
        try:
            wb = openpyxl.load_workbook(fileName)
            if tab_name in wb.sheetnames:
                if wb.sheetnames.count(tab_name) > 1:
                    QMessageBox.information(QMessageBox(), 'Information',
                                            str(' File contains multiple ' + tab_name + ' Sheet.'))
                    return

                sheet = wb[tab_name]
                header = []
                for cell in sheet[1]:
                    header.append(str(cell.value))
                if header == get_db_header(tab_name):
                    conn = sqlite3.connect(PATH_TO_DATABASE)
                    discarded = []
                    ignored = []
                    values = {}
                    for rows in range(2, sheet.max_row + 1):
                        for cols in range(1, len(header)+1):
                            key = header[cols - 1]
                            val = sheet.cell(row=rows, column=cols).value
                            if self.import_db_validation(tab_name, key, val):
                                values.update({key: val})
                            else:
                                discarded.append(sheet[rows][1].value)
                                break
                        c = conn.cursor()
                        if tab_name == 'Columns':
                            c.execute("SELECT count(*) FROM Columns WHERE Designation = ?", (values['Designation'],))
                        elif tab_name == 'Beams':
                            c.execute("SELECT count(*) FROM Beams WHERE Designation = ?", (values['Designation'],))
                        elif tab_name == 'Angles':
                            c.execute("SELECT count(*) FROM Angles WHERE Designation = ?", (values['Designation'],))
                        elif tab_name == 'Channels':
                            c.execute("SELECT count(*) FROM Channels WHERE Designation = ?", (values['Designation'],))

                        data = c.fetchone()[0]
                        if data == 0:
                            values['Source'] = 'Custom'
                            if tab_name == 'Columns':
                                c.execute('''INSERT INTO Columns (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,
                                Iz,Iy,rz,ry,Zz,Zy,Zpz,Zpy,It,Iw,Source,Type) VALUES 
                                (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                                          (values['Designation'], values['Mass'], values['Area'], values['D'],
                                           values['B'], values['tw'], values['T'], values['FlangeSlope'],
                                           values['R1'], values['R2'], values['Iz'], values['Iy'], values['rz'],
                                           values['ry'], values['Zz'], values['Zy'], values['Zpz'], values['Zpy'],
                                           values['It'], values['Iw'], values['Source'], values['Type']))
                            elif tab_name == 'Beams':
                                c.execute('''INSERT INTO Beams (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,
                                Iz,Iy,rz,ry,Zz,Zy,Zpz,Zpy,It,Iw,Source,Type) VALUES
                                (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                                          (values['Designation'], values['Mass'], values['Area'], values['D'],
                                           values['B'], values['tw'], values['T'], values['FlangeSlope'],
                                           values['R1'], values['R2'], values['Iz'], values['Iy'], values['rz'],
                                           values['ry'], values['Zz'], values['Zy'], values['Zpz'], values['Zpy'],
                                           values['It'], values['Iw'], values['Source'], values['Type']))
                            elif tab_name == 'Angles':
                                c.execute('''INSERT INTO Angles (Designation,Mass,Area,a,b,t,R1,R2,Cz,Cy,Iz,Iy,Iumax,
                                Ivmin,rz,ry,rumax,rvmin,Zz,Zy,Zpz,Zpy,It,Source,Type) VALUES
                                (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                                          (values['Designation'], values['Mass'], values['Area'], values['a'],
                                           values['b'], values['t'], values['R1'], values['R2'], values['Cz'],
                                           values['Cy'], values['Iz'], values['Iy'], values['Iumax'], values['Ivmin'],
                                           values['rz'], values['ry'], values['rumax'], values['rvmin'], values['Zz'],
                                           values['Zy'], values['Zpz'], values['Zpy'], values['It'], values['Source'],
                                           values['Type']))
                            elif tab_name == 'Channels':
                                c.execute('''INSERT INTO Channels (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Cy,
                                Iz,Iy,rz,ry,Zz,Zy,Zpz,Zpy,Source,Type) VALUES
                                (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                                          (values['Designation'], values['Mass'], values['Area'], values['D'],
                                           values['B'], values['tw'], values['T'], values['FlangeSlope'], values['R1'],
                                           values['R2'], values['Cy'], values['Iz'], values['Iy'], values['rz'],
                                           values['ry'], values['Zz'], values['Zy'], values['Zpz'], values['Zpy'],
                                           values['Source'], values['Type']))

                            conn.commit()
                            c.close()

                        else:
                            ignored.append(values['Designation'])

                    conn.close()
                    message = QMessageBox()
                    message.setWindowTitle('Successful')
                    message.addButton(message.Ok)
                    message.setText('File data is imported successfully to the database.')
                    if discarded or ignored:
                        rejected = message.addButton('Rejected Sections', message.ActionRole)
                        rejected.clicked.connect(lambda: self.import_validation_dialog(discarded, ignored))
                    message.exec()
                else:
                    QMessageBox.information(QMessageBox(), 'Information',
                                            str(str(tab_name) + ' Sheet has headers different than database.'))

            else:
                QMessageBox.information(QMessageBox(), 'Information', str(' File does not contain '+str(tab_name)+' Sheet.'))

        except IOError:
            QMessageBox.information(QMessageBox(), "Unable to open file",
                                    "There was an error opening \"%s\"" % fileName)
            return

    def import_db_validation(self, tab, key, value):

        if key in ['Mass', 'Area', 'D', 'B', 'tw', 'T', 'FlangeSlope', 'R1', 'R2', 'Iz', 'Iy', 'rz', 'ry', 'Zz', 'Zy',
                   'Zpz', 'Zpy', 'It', 'Iw']:
            return isinstance(value, int) or isinstance(value, float)
        else:
            return True

    def import_validation_dialog(self, discarded, ignored):

        dialog = QDialog()
        dialog.setWindowTitle('Rejected Sections')
        vlayout = QVBoxLayout(dialog)
        height = 200
        total = len(discarded)+len(ignored)
        if 0 < total < 30:
            height += total*10
        else:
            height = 500
        dialog.resize(400, height)
        dialog.setLayout(vlayout)
        if discarded:
            scroll_discarded = QScrollArea(dialog)
            vlayout.addWidget(scroll_discarded)
            scroll_discarded.setWidgetResizable(True)
            scroll_discarded.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
            widget_discarded = QWidget(scroll_discarded)
            layout_discarded = QVBoxLayout(widget_discarded)
            widget_discarded.setLayout(layout_discarded)
            label_discarded = QLabel("These values were rejected in the validation checks.")
            layout_discarded.addWidget(label_discarded)
            scroll_discarded.setWidget(widget_discarded)
            text_discarded = QTextBrowser()
            layout_discarded.addWidget(text_discarded)
            for d in discarded:
                text_discarded.append(d)
        if ignored:
            scroll_ignored = QScrollArea(dialog)
            vlayout.addWidget(scroll_ignored)
            scroll_ignored.setWidgetResizable(True)
            scroll_ignored.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
            widget_ignored = QWidget(scroll_ignored)
            layout_ignored = QVBoxLayout(widget_ignored)
            widget_ignored.setLayout(layout_ignored)
            label_ignored = QLabel("These values were ignored because they already exist in the database.")
            layout_ignored.addWidget(label_ignored)
            scroll_ignored.setWidget(widget_ignored)
            text_ignored = QTextBrowser()
            layout_ignored.addWidget(text_ignored)
            for i in ignored:
                text_ignored.append(i)
        dialog.exec()

        # self.ui.pushButton_Import_Column.setDisabled(True)

    # def import_BeamPref(self):
    #     wb = openpyxl.load_workbook(os.path.join(str(self.folder), "images_html", "add_sections.xlsx"))
    #     sheet = wb['First Sheet']
    #     conn = sqlite3.connect('ResourceFiles/Database/Intg_osdag.sqlite')
    #
    #     for rowNum in range(2, sheet.max_row + 1):
    #         designation = sheet.cell(row=rowNum, column=2).value
    #         mass = sheet.cell(row=rowNum, column=3).value
    #         area = sheet.cell(row=rowNum, column=4).value
    #         d = sheet.cell(row=rowNum, column=5).value
    #         b = sheet.cell(row=rowNum, column=6).value
    #         tw = sheet.cell(row=rowNum, column=7).value
    #         t = sheet.cell(row=rowNum, column=8).value
    #         flangeSlope = sheet.cell(row=rowNum, column=9).value
    #         r1 = sheet.cell(row=rowNum, column=10).value
    #         r2 = sheet.cell(row=rowNum, column=11).value
    #         iz = sheet.cell(row=rowNum, column=12).value
    #         iy = sheet.cell(row=rowNum, column=13).value
    #         rz = sheet.cell(row=rowNum, column=14).value
    #         ry = sheet.cell(row=rowNum, column=15).value
    #         zz = sheet.cell(row=rowNum, column=16).value
    #         zy = sheet.cell(row=rowNum, column=17).value
    #         zpz = sheet.cell(row=rowNum, column=18).value
    #         zpy = sheet.cell(row=rowNum, column=19).value
    #         source = sheet.cell(row=rowNum, column=20).value
    #
    #         c = conn.cursor()
    #         c.execute("SELECT count(*) FROM Beams WHERE Designation = ?", (designation,))
    #         data = c.fetchone()[0]
    #         if data == 0:
    #             c.execute('''INSERT INTO Beams (Designation,Mass,Area,D,B,tw,T,FlangeSlope,R1,R2,Iz,Iy,rz,ry,
    #                                                    Zz,zy,Zpz,Zpy,Source) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
    #                       (designation, mass, area,
    #                        d, b, tw, t,
    #                        flangeSlope, r1
    #                        ,
    #                        r2, iz, iy, rz, ry,
    #                        zz, zy
    #                        ,
    #                        zpz, zpy, source))
    #             conn.commit()
    #             c.close()
    #
    #     conn.close()
    #     QMessageBox.information(QMessageBox(), 'Successful', ' File data is imported successfully to the database.')
    #     self.ui.pushButton_Import_Beam.setDisabled(True)

        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Column), _translate("DesignPreferences", "Column"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Beam), _translate("DesignPreferences", "Beam"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Bolt), _translate("DesignPreferences", "Bolt"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Weld), _translate("DesignPreferences", "Weld"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Detailing), _translate("DesignPreferences", "Detailing"))
        # self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_Design), _translate("DesignPreferences", "Design"))


class DesignPreferences():

    def __init__(self, main, module_window, input_dictionary, parent=None):

        self.ui = Window( main, input_dictionary)

        #self.ui.setupUi(self, main, input_dictionary)
        #self.ui.show()
        self.main_controller = parent
        #self.uiobj = self.main_controller.uiObj
        self.module_window = module_window
        self.saved = None
        self.flag = False
        self.sectionalprop = I_sectional_Properties()
        #self.ui.btn_save.hide()
        self.ui.btn_save.clicked.connect(self.close_designPref)
        self.ui.btn_defaults.clicked.connect(lambda: self.default_fn(main, input_dictionary))
        self.module = main.module_name(main)
        self.main = main
        self.window_close_flag = True
        self.changes = None

    def show(self):
        resolution = QtWidgets.QDesktopWidget().screenGeometry()
        width = resolution.width()
        height = resolution.height()
        # self.ui.resize(width*(0.67),height*(0.60))
        self.ui.resize(width * 0.7, height * 0.6)
        # self.ui.center()
        # self.ui.tabWidget.resize(width * (0.67), height * (0.60))
        self.ui.setWindowFlag(Qt.WindowMinimizeButtonHint, True)
        self.ui.setWindowFlag(Qt.WindowMaximizeButtonHint, True)
        self.ui.center()
        self.changes = self.ui.exec_()
        if self.changes != QDialog.Accepted:
            self.flag = False
        self.module_window.prev_inputs = self.module_window.input_dock_inputs

    def default_fn(self, main, input_dictionary):
        '''
        @author: Umair
        '''
        tab_Bolt = self.ui.tabWidget.tabs.findChild(QtWidgets.QWidget, "Bolt")
        tab_Weld = self.ui.tabWidget.tabs.findChild(QtWidgets.QWidget, "Weld")
        tab_Detailing = self.ui.tabWidget.tabs.findChild(QtWidgets.QWidget, "Detailing")
        tab_Design = self.ui.tabWidget.tabs.findChild(QtWidgets.QWidget, "Design")

        bolt_values_dictionary = {}
        weld_values_dictionary = {}
        design_values_dictionary = {}
        detailing_values_dictionary = {}

        if tab_Bolt is not None:
            for i in main.bolt_values(main, input_dictionary):
                if i[2] in [TYPE_TEXTBOX, TYPE_COMBOBOX]:
                    bolt_values_dictionary.update(
                        {i[0]: str(main.get_values_for_design_pref(main, i[0], input_dictionary))})

            for children in tab_Bolt.findChildren(QtWidgets.QWidget):
                if children.objectName() in bolt_values_dictionary.keys():
                    if type(children) == QLineEdit:
                        children.setText(bolt_values_dictionary[children.objectName()])
                        # if bolt_values_dictionary[children.objectName()==0:
                        #     children.textEdit.setDisabled(True)

                    elif type(children) == QComboBox:
                        children.setCurrentText(bolt_values_dictionary[children.objectName()])
                    else:
                        pass

        if tab_Weld is not None:
            for i in main.weld_values(main, input_dictionary):
                if i[2] in [TYPE_TEXTBOX, TYPE_COMBOBOX]:
                    weld_values_dictionary.update(
                        {i[0]: str(main.get_values_for_design_pref(main, i[0], input_dictionary))})

            for children in tab_Weld.findChildren(QtWidgets.QWidget):
                if children.objectName() in weld_values_dictionary.keys():
                    if type(children) == QLineEdit:
                        children.setText(weld_values_dictionary[children.objectName()])
                    elif type(children) == QComboBox:
                        children.setCurrentText(weld_values_dictionary[children.objectName()])
                    else:
                        pass

        if tab_Detailing is not None:
            for i in main.detailing_values(main, input_dictionary):
                if i[2] in [TYPE_TEXTBOX, TYPE_COMBOBOX]:
                    detailing_values_dictionary.update(
                        {i[0]: str(main.get_values_for_design_pref(main, i[0], input_dictionary))})

            for children in tab_Detailing.findChildren(QtWidgets.QWidget):
                if children.objectName() in detailing_values_dictionary.keys():
                    if type(children) == QLineEdit:
                        children.setText(detailing_values_dictionary[children.objectName()])
                    elif type(children) == QComboBox:
                        children.setCurrentText(detailing_values_dictionary[children.objectName()])
                    else:
                        pass

        if tab_Design is not None:
            for i in main.design_values(main, input_dictionary):
                if i[2] in [TYPE_TEXTBOX, TYPE_COMBOBOX]:
                    design_values_dictionary.update(
                        {i[0]: str(main.get_values_for_design_pref(main, i[0], input_dictionary))})

            for children in tab_Design.findChildren(QtWidgets.QWidget):
                if children.objectName() in design_values_dictionary.keys():
                    if type(children) == QLineEdit:
                        children.setText(design_values_dictionary[children.objectName()])
                    elif type(children) == QComboBox:
                        children.setCurrentText(design_values_dictionary[children.objectName()])
                    else:
                        pass

    def highlight_slipfactor_description(self):
        """Highlight the description of currosponding slipfactor on selection of inputs
        Note : This routine is not in use in current version
        :return:
        """
        slip_factor = str(self.ui.combo_slipfactor.currentText())
        self.textCursor = QTextCursor(self.ui.textBrowser.document())
        cursor = self.textCursor
        # Setup the desired format for matches
        format = QTextCharFormat()
        format.setBackground(QBrush(QColor("red")))
        # Setup the regex engine
        pattern = str(slip_factor)
        regex = QRegExp(pattern)
        # Process the displayed document
        pos = 0
        index = regex.indexIn(self.ui.textBrowser.toPlainText(), pos)
        while (index != -1):
            # Select the matched text and apply the desired format
            cursor.setPosition(index)
            cursor.movePosition(QTextCursor.EndOfLine, 1)
            # cursor.movePosition(QTextCursor.EndOfWord, 1)
            cursor.mergeCharFormat(format)
            # Move to the next match
            pos = index + regex.matchedLength()
            index = regex.indexIn(self.ui.textBrowser.toPlainText(), pos)


    def fu_fy_validation_connect(self, fu_fy_list, f, m):
        f.textChanged.connect(lambda: self.fu_fy_validation(fu_fy_list, f, m))

    def fu_fy_validation(self, fu_fy_list, textbox, material_key):
        self.window_close_flag = False
        # self.rejected.disconnect()
        # self.rejected.connect(self.closeEvent_accept)
        #print(fu_fy_list[0].text(), fu_fy_list[1].text())
        if "" not in [fu_fy_list[0].text(), fu_fy_list[1].text()]:
            fu = float(fu_fy_list[0].text())
            fy = float(fu_fy_list[1].text())
            material = material_key.currentText()

        else:
            textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: black;")
            return

        if fu and fy:
            if 'Ultimate_Strength' in textbox.objectName():

                if fu < 290 or fu > 639:
                    textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: red;")
                    self.window_close_flag = False
                    self.rejected.connect(self.closeEvent)
                    return
                else:

                    fu_limits = self.get_limits_for_fu(str(material))

                    if fu_limits['lower'] <= fu <= fu_limits['upper']:
                        textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: black;")
                        self.window_close_flag = True
                        self.rejected.connect(self.closeEvent)
                        return
                    else:
                        textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: red;")
                        self.window_close_flag = False
                        self.rejected.connect(self.closeEvent)
                        return

            if 'Yield_Strength' in textbox.objectName():
                if fy < 165 or fy > 499:
                    textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: red;")
                    self.window_close_flag = False
                    self.rejected.connect(self.closeEvent)
                    return

                else:

                    fy_limits = self.get_limits_for_fy(str(material))
                    if fy_limits['lower'] <= fy <= fy_limits['upper']:
                        textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: black;")
                        self.window_close_flag = True
                        self.rejected.connect(self.closeEvent)
                        return
                    else:
                        textbox.setStyleSheet("border: 1 px; border-style: solid; border-color: red;")
                        self.window_close_flag = False
                        self.rejected.connect(self.closeEvent)
                        return

    def get_limits_for_fu(self, material_grade):

        lower_fu = {'E 165 (Fe 290)': 290,
                    'E 250 (Fe 410 W)A': 410,
                    'E 250 (Fe 410 W)B': 410,
                    'E 250 (Fe 410 W)C': 410,
                    'E 300 (Fe 440)': 440,
                    'E 350 (Fe 490)': 490,
                    'E 410 (Fe 540)': 540,
                    'E 450 (Fe 570)D': 570,
                    'E 450 (Fe 590) E': 590}[material_grade]

        upper_fu = {'E 165 (Fe 290)': 409,
                    'E 250 (Fe 410 W)A': 439,
                    'E 250 (Fe 410 W)B': 439,
                    'E 250 (Fe 410 W)C': 439,
                    'E 300 (Fe 440)': 489,
                    'E 350 (Fe 490)': 539,
                    'E 410 (Fe 540)': 569,
                    'E 450 (Fe 570)D': 589,
                    'E 450 (Fe 590) E': 639}[material_grade]

        return {'lower': lower_fu, 'upper': upper_fu}

    def get_limits_for_fy(self, material_grade):

        lower_fy = {'E 165 (Fe 290)': 165,
                    'E 250 (Fe 410 W)A': 230,
                    'E 250 (Fe 410 W)B': 230,
                    'E 250 (Fe 410 W)C': 230,
                    'E 300 (Fe 440)': 280,
                    'E 350 (Fe 490)': 320,
                    'E 410 (Fe 540)': 380,
                    'E 450 (Fe 570)D': 420,
                    'E 450 (Fe 590) E': 420}[material_grade]

        upper_fy = {'E 165 (Fe 290)': 249,
                    'E 250 (Fe 410 W)A': 299,
                    'E 250 (Fe 410 W)B': 299,
                    'E 250 (Fe 410 W)C': 299,
                    'E 300 (Fe 440)': 349,
                    'E 350 (Fe 490)': 409,
                    'E 410 (Fe 540)': 449,
                    'E 450 (Fe 570)D': 499,
                    'E 450 (Fe 590) E': 499}[material_grade]

        return {'lower': lower_fy, 'upper': upper_fy}

    def closeEvent(self, event):
        if self.window_close_flag:
            event.accept()
            # self.module_window.prev_inputs = self.module_window.input_dock_inputs
        else:
            QMessageBox.warning(self, "Error", "Select correct values for fu and fy!")
            event.ignore()

    def close_designPref(self):
        self.ui.accept()

    # def closeEvent(self, QCloseEvent):
    #     self.save_designPref_para()
    #     QCloseEvent.accept()
