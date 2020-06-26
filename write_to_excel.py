from design_type.connection.shear_connection import ShearConnection
from design_type.connection.fin_plate_connection import FinPlateConnection
from design_type.tension_member.tension_bolted import Tension_bolted
from design_type.tension_member.tension_welded import Tension_welded
from design_type.connection.beam_cover_plate import BeamCoverPlate
from design_type.connection.beam_cover_plate_weld import BeamCoverPlateWeld
from design_type.connection.column_cover_plate import ColumnCoverPlate
from design_type.connection.column_cover_plate_weld import ColumnCoverPlateWeld
from design_type.connection.column_end_plate import ColumnEndPlate
import yaml
from utils.common.component import Bolt, Plate, Weld
from Common import *
import os
import xlsxwriter
# import pandas as pd
from openpyxl import load_workbook, Workbook

path = r'.\ResourceFiles\design_example'

files = []
# r=root, d=directories, f = files
for r, d, f in os.walk(path):
    for file in f:
        if '.osi' in file:
            files.append(os.path.join(r, file))


for f in files:
    print(f)
    with open(f, 'r') as input_file:
        d = yaml.load(input_file, Loader=yaml.FullLoader)

    module = d['Module']


    if module == KEY_DISP_COLUMNCOVERPLATEWELD:
        main = ColumnCoverPlateWeld
        main.set_osdaglogger(None)
        main.set_input_values(main, d)
        base = os.path.basename(f)
        #TODO:comment below line for fin check
        test_in_list = d
        # filename = str(os.path.splitext(base)[0])+".txt"
        # test_out_list = main.results_to_test(main)
        # f = open(filename, "w")
        # f.write(str(test_out_list))
        # f.close()

        workbook_name = 'CCWP.xlsx'

        sheet_name = str(os.path.splitext(base)[0])
        try:
            wb = load_workbook(workbook_name)
            print('entered try')
        except Exception as e:
            print('creatingnew')
            wb = Workbook()
        # wb.create_sheet('test2')



        # writer = ExcelWriter(workbook_name, engine='openpyxl')
        # writer.book = wb
        test_out_list = main.results_to_test(main)
        # TODO:uncomment below line for fin check
        # [test_in_list,test_out_list] = main.results_to_test(main)

        # df = pd.DataFrame.from_records(list(test_out_list.items()), columns=['Check', 'Value'])

        input_keys = list(test_in_list.keys())
        input_values = list(map(str, list(test_in_list.values())))
        output_keys = list(test_out_list.keys())
        output_values = list(map(str, list(test_out_list.values())))

        while len(input_keys) != len(output_keys):
            if len(input_keys) < len(output_keys):
                for i in range(0,len(output_keys)-len(input_keys)):
                    input_keys.append('')
                    input_values.append('')
            else:
                for i in range(0,len(input_keys)-len(output_keys)):
                    output_keys.append('')
                    output_values.append('')

        sheet_name_as_list = [sheet_name]
        for i in range(0, len(input_keys)):
            sheet_name_as_list.append('')

        for row in zip(sheet_name_as_list,input_keys,input_values,output_keys,output_values):
            wb.active.append(row)

        # df.to_excel(writer, sheet_name=sheet_name, index=False)
        wb.save(workbook_name)
        wb.close()
        # writer.save()
        # writer.close()
    #
    # elif module == KEY_DISP_TENSION_BOLTED:
    #     print('filenAME', f)
    #     main = Tension_bolted
    #     main.set_osdaglogger(None)
    #     main.set_input_values(main, d)
    #
    # elif module == KEY_DISP_TENSION_WELDED:
    #     print('filenAME', f)
    #     main = Tension_welded
    #     main.set_osdaglogger(None)
    #     main.set_input_values(main, d)
    #
    # elif module == KEY_DISP_COLUMNCOVERPLATE:
    #     print('filenAME', f)
    #     main = ColumnCoverPlate
    #     main.set_osdaglogger(None)
    #     main.set_input_values(main, d)
    #
    # elif module == KEY_DISP_COLUMNCOVERPLATEWELD:
    #     print('filenAME', f)
    #     main = ColumnCoverPlateWeld
    #     main.set_osdaglogger(None)
    #     main.set_input_values(main, d)
    #
    # elif module == KEY_DISP_BEAMCOVERPLATE:
    #     print('filenAME', f)
    #     main = BeamCoverPlate
    #     main.set_osdaglogger(None)
    #     main.set_input_values(main, d)
    #
    # elif module == KEY_DISP_BEAMCOVERPLATEWELD:
    #     print('filenAME', f)
    #     main = BeamCoverPlateWeld
    #     main.set_osdaglogger(None)
    #     main.set_input_values(main, d)


    # elif module == 'Column Coverplate Connection':
    #     self = ColumnCoverPlate
    #     self.set_osdaglogger()
    #     self.set_input_values(self, d)




