from design_type.connection.shear_connection import ShearConnection
from design_type.connection.fin_plate_connection import FinPlateConnection
import yaml
from utils.common.component import Bolt, Plate, Weld
from Common import *
import os
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook, Workbook

path = r'.\ResourceFiles\design_example'

files = []
# r=root, d=directories, f = files
for r, d, f in os.walk(path):
    for file in f:
        if '.osi' in file:
            files.append(os.path.join(r, file))


for f in files:
    print(f)
    with open(f, 'r') as input_file:
        d = yaml.load(input_file, Loader=yaml.FullLoader)

    module = d['Module']

    if module == 'Fin Plate':
        main = FinPlateConnection
        main.set_osdaglogger(None)
        main.set_input_values(main, d)
        base = os.path.basename(f)
        # filename = str(os.path.splitext(base)[0])+".txt"
        # test_out_list = main.results_to_test(main)
        # f = open(filename, "w")
        # f.write(str(test_out_list))
        # f.close()
        workbook_name = 'Fin.xlsx'
        sheet_name = str(os.path.splitext(base)[0])
        try:
            wb = load_workbook(workbook_name)
            print('entered try')
        except Exception as e:
            print('creatingnew')
            wb = Workbook()
        # wb.create_sheet('test2')
        # wb.save('test.xls')


        writer = pd.ExcelWriter(workbook_name, engine='openpyxl')
        writer.book = wb
        test_out_list = main.results_to_test(main)

        df = pd.DataFrame.from_records(list(test_out_list.items()), columns=['Check', 'Value'])


        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()
        writer.close()


    # elif module == 'Column Coverplate Connection':
    #     self = ColumnCoverPlate
    #     self.set_osdaglogger()
    #     self.set_input_values(self, d)




